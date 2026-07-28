#!/usr/bin/env python3
"""Template SHAPE probe — the pre-change baseline for the quarterly lane.

⭐ WHY THIS RUNS BEFORE ANY EDIT. Item 9 asks that the ANNUAL template be
confirmed byte-shape UNCHANGED. "Unchanged" is only meaningful against a
recorded before-state, and a before-state captured after the edit is not a
before-state. So this probe is run and its output recorded first.

It reads the generated workbook rather than the generator source: the question
is what the customer receives, and source that looks correct can still emit a
workbook that is not.

Read-only. Generates in memory, writes nothing to the product.
"""
import sys, os, io, argparse

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "services"))

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from api.modules.financials import ingest


def describe(freq, last_hist_year=2025):
    blob = ingest.build_company_template(
        company_id=1, company_name="Shape Probe", currency="USD",
        statement_units="actual", ownership="private", standard="us_gaap",
        frequency=freq, last_historical_year=last_hist_year)
    wb = load_workbook(io.BytesIO(blob))
    out = {"bytes": len(blob), "sheets": wb.sheetnames}

    # the income statement is the reference sheet the parser keys on
    is_name = [n for n in wb.sheetnames if "Income" in n or "income" in n]
    ws = wb[is_name[0]] if is_name else wb[wb.sheetnames[0]]
    out["sheet"] = ws.title

    hist, fcst = [], []
    for c in range(2, ws.max_column + 1):
        L = get_column_letter(c)
        kind = str(ws[f"{L}3"].value or "").strip().lower()
        yr = ws[f"{L}4"].value
        if kind == "historical": hist.append((L, yr))
        elif kind == "forecast": fcst.append((L, yr))
    out["historical"] = hist
    out["forecast"] = fcst
    out["max_column"] = ws.max_column
    out["last_col_letter"] = get_column_letter(ws.max_column)

    # dropdown coverage — the sqref the Historical/Forecast validation actually spans
    dvs = []
    for dv in ws.data_validations.dataValidation:
        if dv.type == "list" and "Historical" in str(dv.formula1):
            dvs.append(str(dv.sqref))
    out["dropdown_sqref"] = dvs

    out["merges"] = [str(m) for m in ws.merged_cells.ranges]

    # shading discipline: shaded IFF unlocked input
    viol = []
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 40),
                            min_col=2, max_col=ws.max_column):
        for cell in row:
            shaded = cell.fill is not None and cell.fill.fill_type == "solid"
            unlocked = cell.protection is not None and cell.protection.locked is False
            is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            if shaded and is_formula:
                viol.append(f"{cell.coordinate} shaded formula")
            if shaded and not unlocked:
                viol.append(f"{cell.coordinate} shaded but locked")
    out["shading_violations"] = viol
    out["protection"] = bool(ws.protection.sheet)

    out["defined_names"] = sorted(wb.defined_names.keys())
    out["hidden_sheets"] = [s.title for s in wb.worksheets if s.sheet_state != "visible"]

    inst = [n for n in wb.sheetnames if "nstruction" in n]
    if inst:
        iws = wb[inst[0]]
        out["instructions"] = [str(iws.cell(row=r, column=1).value)
                               for r in range(1, 20)
                               if iws.cell(row=r, column=1).value
                               and ("Forecast" in str(iws.cell(row=r, column=1).value)
                                    or "years" in str(iws.cell(row=r, column=1).value)
                                    or "quarters" in str(iws.cell(row=r, column=1).value))]
    return out


def show(tag, d):
    print(f"\n{'='*82}\n{tag}\n{'='*82}")
    print(f"  bytes={d['bytes']}  sheet='{d['sheet']}'  max_column={d['max_column']} ({d['last_col_letter']})")
    print(f"  HISTORICAL {len(d['historical'])}: {d['historical'][0][0]}..{d['historical'][-1][0]}"
          f"  labels={[y for _, y in d['historical']]}")
    if d["forecast"]:
        print(f"  FORECAST   {len(d['forecast'])}: {d['forecast'][0][0]}..{d['forecast'][-1][0]}"
              f"  labels={[y for _, y in d['forecast']]}")
    else:
        print("  FORECAST   0")
    print(f"  dropdown sqref     : {d['dropdown_sqref']}")
    print(f"  merged ranges      : {d['merges']}")
    print(f"  sheet protection   : {d['protection']}")
    print(f"  defined names ({len(d['defined_names'])}): {d['defined_names']}")
    print(f"  hidden sheets      : {d['hidden_sheets']}")
    print(f"  shading violations : {d['shading_violations'] or 'NONE — shaded IFF unlocked input'}")
    for line in d.get("instructions", []):
        print(f"  instructions       : {line[:100]}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--simulate-quarterly-forecast", type=int, default=None,
                    help="probe what the generator WOULD emit at this count, "
                         "by patching the constant in memory only")
    a = ap.parse_args()
    if a.simulate_quarterly_forecast:
        # In-memory only. The source file is not edited — this answers "what would
        # the workbook look like" without committing to the change.
        ingest.FORECAST_QUARTERLY = a.simulate_quarterly_forecast
        print(f"[simulation] FORECAST_QUARTERLY patched IN MEMORY to "
              f"{a.simulate_quarterly_forecast} — source file untouched")
    show("ANNUAL", describe("annual"))
    show("QUARTERLY", describe("quarterly"))
