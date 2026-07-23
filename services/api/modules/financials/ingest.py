"""Phase 7a-2 — company-scoped data ingestion: a themed, pre-filled Excel
template generator and an upload validator that produces the canonical dataset.

Reuses the standard line-item vocabulary (LABELS / BLOCK_KEYS) and validation
(engines.validate_dataset) so uploaded company data is identical in shape to
the showcase datasets the engines already consume. A hidden _AXIOM sheet lets
an upload self-identify (company_id, frequency, template_version).
"""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import engines
from .templates import LABELS, COMPANY_ROWS, BLOCK_KEYS

TEMPLATE_VERSION = "7L-v5.1"   # §9 OKR v5.1: 200-row capacity, paste-safe validation, no parser cap

# §9 OKR strategy sheets — fixed-name, standard-independent. Builder + parser share
# these so the dropdown values and the accepted enums can never drift apart.
OBJECTIVES_SHEET = "Objectives"
KR_SHEET = "Key Results"
KPI_SHEET = "KPI Plan vs Actual"
GOALS_SHEET = "Organizational Goals"                 # legacy v4 sheet name (still parsed → objectives)
OBJ_PRIORITIES = ("High", "Medium", "Low")
OBJ_HORIZONS = ("Short", "Medium", "Long")           # canonical dropdown values; ranges live in the note
OBJ_STATUSES = ("Red", "Amber", "Green")             # optional current-health dropdown
# back-compat aliases (legacy goal parser reuses these enums)
GOAL_PRIORITIES = OBJ_PRIORITIES
GOAL_HORIZONS = OBJ_HORIZONS
GOAL_STATUSES = OBJ_STATUSES
OBJ_HEADER_ROW = 2                                     # row 1 = guidance note; row 2 = headers; row 3+ = data
OBJ_DATA_START = 3
KR_HEADER_ROW = 2
KR_DATA_START = 3
KPI_HEADER_ROW = 2
KPI_DATA_START = 3
KPI_SEED_ROWS = ("Revenue growth %", "EBITDA margin %", "Operating margin %", "Market share %")
MIN_KRS_PER_OBJECTIVE = 1                              # hard rule: ≥1 KR per objective (guidance — warns, never blocks)
ROW_CAPACITY = 200                                    # v5.1: pre-formatted input rows per strategy sheet
OBJ_ID_MAX = 200                                      # Objective IDs pre-seeded O1…O200
# statement_units -> factor that normalizes raw figures to the canonical internal
# unit (MILLIONS) the valuation engine + report builder assume. Honoring this is
# how "actual / thousands / millions" flows correctly end-to-end.
UNIT_SCALE = {"actual": 1e-6, "thousands": 1e-3, "millions": 1.0}
META_SIG = "AXIOM-COMPANY-TEMPLATE"
_LOCK_PWD = "AXIOM"

# AXIOM dark-green theme
_INK = "0D1B12"          # near-black green (headers bg)
_GREEN = "1F6F43"        # section fill
_ACCENT = "4ADE80"       # accent text
_INPUT = "EAF7EF"        # unlocked input cell tint
_SUBTOTAL = "D7E9DD"     # locked subtotal/formula tint
_thin = Side(style="thin", color="BBD9C6")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

FIRST_COL = 2            # period columns start at B
ANNUAL_COLS = 5
QUARTERLY_COLS = 12

# subtotal/derived rows shown per sheet as LOCKED formulas (guidance only —
# the server recomputes; these help the user sanity-check as they type).
SUBTOTALS = {
    "income_statement": [("Gross Profit", "={rev}-{cogs}"),
                         ("EBITDA", "={rev}-{cogs}-{opex}"),
                         ("EBIT", "={rev}-{cogs}-{opex}-{da}")],
    "balance_sheet": [("Total Assets", "={cash}+{oca}+{nca}"),
                      ("Total Liabilities & Equity",
                       "={cle}+{std}+{ltd}+{pfd}+{mi}+{te}")],
    "cash_flow": [],
}


_WM_FILL = "FFE8A3"   # amber watermark
WATERMARK = ("SAMPLE DATA (illustrative, in thousands) — replace on EVERY sheet "
             "with your own figures")


def _watermark(cell):
    cell.value = WATERMARK
    cell.font = Font(bold=True, italic=True, color="8A5A00")
    cell.fill = PatternFill("solid", fgColor=_WM_FILL)
    cell.protection = Protection(locked=True)


def _r50(x):
    return int(round(x / 50.0)) * 50


def build_sample_data(ownership: str, standard: str, frequency: str) -> dict:
    """A complete, internally-consistent fictional company: round driver
    figures, a balance sheet that balances every period, cash that reconciles
    to the cash-flow items, and all validator-required company fields. Articulated
    by roll-forward so assets == liabilities+equity holds by construction."""
    if frequency == "quarterly":
        periods = [y * 10 + q for y in (2020, 2021, 2022) for q in (1, 2, 3, 4)]
        # mild seasonality (Q1 low, Q4 high), growing year over year
        revenue = [2250, 2400, 2500, 2850, 2550, 2700, 2850, 3200,
                   2850, 3050, 3200, 3600]
    else:
        periods = [2020, 2021, 2022, 2023, 2024]
        revenue = [10000, 11500, 13200, 15200, 17500]
    n = len(periods)
    cogs = [_r50(0.55 * r) for r in revenue]
    opex = [_r50(0.18 * r) for r in revenue]
    da = [_r50(0.05 * r) for r in revenue]
    interest = [_r50(0.03 * r) for r in revenue]
    capex = [_r50(0.07 * r) for r in revenue]
    net_borrow = [_r50(0.02 * r) for r in revenue]
    dividends = [_r50(0.02 * r) for r in revenue]
    oca = [_r50(0.30 * r) for r in revenue]
    cle = [_r50(0.15 * r) for r in revenue]
    std = [_r50(0.10 * revenue[0])] * n            # constant short-term debt
    pfd = [0] * n
    mi = [0] * n
    # opening balance sheet (period 0): cash is the plug so assets == L+E
    nca = [_r50(0.80 * revenue[0])]
    ltd = [_r50(0.30 * revenue[0])]
    te = [_r50(0.80 * revenue[0])]
    cash = [(cle[0] + std[0] + ltd[0] + te[0]) - oca[0] - nca[0]]
    for t in range(1, n):
        pretax = revenue[t] - cogs[t] - opex[t] - da[t] - interest[t]
        ni = int(round(pretax * 0.75))             # integer NI keeps cells clean
        nca.append(nca[t - 1] + capex[t] - da[t])  # PP&E roll-forward
        ltd.append(ltd[t - 1] + net_borrow[t])
        te.append(te[t - 1] + ni - dividends[t])   # retained-earnings roll
        dnwc = (oca[t] - cle[t]) - (oca[t - 1] - cle[t - 1])
        cash.append(cash[t - 1] + ni + da[t] - dnwc - capex[t]
                    + net_borrow[t] - dividends[t])
    def ser(vals):
        return {str(periods[t]): float(vals[t]) for t in range(n)}
    company = {"tax_rate": 0.25, "risk_free_rate": 0.04,
               "market_risk_premium": 0.055, "cost_of_debt": 0.06}
    if ownership == "public":
        company.update({"shares_outstanding": 1000.0, "share_price": 25.0,
                        "beta": 1.20})
    else:
        company.update({"unlevered_industry_beta": 1.10, "target_debt_to_equity": 0.50,
                        "size_premium": 0.03, "specific_risk_premium": 0.02, "dlom": 0.20})
    return {
        "periods": periods, "company": company,
        "income_statement": {"revenue": ser(revenue), "cogs": ser(cogs),
                             "opex": ser(opex), "depreciation_amortization": ser(da),
                             "interest_expense": ser(interest)},
        "balance_sheet": {"cash": ser(cash), "other_current_assets": ser(oca),
                          "noncurrent_assets": ser(nca), "current_liabilities_ex_debt": ser(cle),
                          "short_term_debt": ser(std), "long_term_debt": ser(ltd),
                          "preferred_equity": ser(pfd), "minority_interest": ser(mi),
                          "total_equity": ser(te)},
        "cash_flow": {"capex": ser(capex), "net_borrowing": ser(net_borrow),
                      "dividends": ser(dividends)}}


def _hdr(cell, text, bg=_INK, fg="FFFFFF", size=11):
    cell.value = text
    cell.font = Font(bold=True, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _input(cell):
    cell.protection = Protection(locked=False)
    cell.fill = PatternFill("solid", fgColor=_INPUT)
    cell.number_format = "#,##0.00"
    cell.border = _border


def _locked_formula(cell, formula):
    cell.value = formula
    cell.protection = Protection(locked=True)
    cell.fill = PatternFill("solid", fgColor=_SUBTOTAL)
    cell.number_format = "#,##0.00"
    cell.font = Font(italic=True)
    cell.border = _border


def build_company_template(*, company_id: int, company_name: str, currency: str,
                           statement_units: str, ownership: str,
                           standard: str = "us_gaap",
                           frequency: str = "annual") -> bytes:
    """Themed, pre-filled input workbook for one company."""
    frequency = (frequency or "annual").lower()
    if frequency not in ("annual", "quarterly"):
        raise ValueError("frequency must be 'annual' or 'quarterly'")
    if standard not in LABELS:
        standard = "us_gaap"
    lab = LABELS[standard]
    ncols = ANNUAL_COLS if frequency == "annual" else QUARTERLY_COLS
    unit_label = {"actual": "actual amounts", "thousands": "in thousands",
                  "millions": "in millions"}.get(statement_units, statement_units)

    sample = build_sample_data(ownership, standard, frequency)
    sample_periods = sample["periods"]

    wb = Workbook()

    # ---- Instructions ----
    ws = wb.active
    ws.title = "Instructions"
    _hdr(ws["A1"], f"AXIOM — {company_name}", bg=_INK, fg=_ACCENT, size=16)
    _watermark(ws["A2"])
    ws["A3"] = f"Reporting currency: {currency}   ·   Units: {unit_label}   ·   {frequency.title()}"
    ws["A3"].font = Font(color="446655")
    for r, line in enumerate([
        "How to complete this workbook:",
        "1. The green cells hold SAMPLE figures for a fictional company (illustrative,",
        "   in thousands). Replace them on EVERY sheet with your own figures in the",
        f"   units shown ({unit_label}). Leaving any sheet's sample data in is rejected on upload.",
        f"2. Enter {ncols} {'years' if frequency=='annual' else 'quarters'} across the statement sheets.",
        "   Row 4 = the period label; row 3 marks Historical or Forecast.",
        "   At least one historical period is required.",
        "3. Only the highlighted (green-tinted) cells accept input; labels and",
        "   subtotal rows are locked. The lock is a guardrail — AXIOM re-validates",
        "   every cell on upload; the server-side validator is the guarantee.",
        "4. Enter rates as decimals (7% = 0.07). Amounts consistently in the",
        f"   stated units ({unit_label}).",
        "5. Upload at POST /companies/{id}/data-upload — the file self-identifies.",
    ], start=5):
        ws[f"A{r}"] = line
    ws.column_dimensions["A"].width = 84
    ws.protection.sheet = True
    ws.protection.password = _LOCK_PWD

    # ---- Company profile (pre-filled) ----
    ws = wb.create_sheet("Company")
    _hdr(ws["A1"], "Company Profile", bg=_GREEN)
    _hdr(ws["B1"], "Value", bg=_GREEN)
    prefill = {"name": company_name, "currency": currency, "ownership": ownership}
    last_company_row = 1
    for r, (field, label, applies) in enumerate(COMPANY_ROWS, start=2):
        last_company_row = r
        ws[f"A{r}"] = label
        c = ws[f"B{r}"]
        _input(c)
        if field in ("name", "ownership", "currency"):
            c.number_format = "General"
        if field in prefill:
            c.value = prefill[field]
        elif field in sample["company"]:      # sample rates/params for this ownership
            c.value = sample["company"][field]
    _watermark(ws[f"A{last_company_row + 2}"])   # banner below the input range
    dv = DataValidation(type="list", formula1='"public,private"', allow_blank=False)
    ws.add_data_validation(dv); dv.add(ws["B3"])
    ws.column_dimensions["A"].width = 54
    ws.column_dimensions["B"].width = 24
    ws.protection.sheet = True
    ws.protection.password = _LOCK_PWD

    # ---- Statement sheets ----
    def colref(letter, row):
        return f"{letter}{row}"

    for block, keys in BLOCK_KEYS.items():
        ws = wb.create_sheet(lab["sheets"][block])
        _hdr(ws["A1"], lab["sheets"][block], bg=_INK, fg=_ACCENT, size=13)
        _watermark(ws["A2"])          # header-area banner, outside the input range
        ws["A3"] = "Period Type (Historical / Forecast)"
        ws["A4"] = "Period (year)"
        ws["A3"].font = ws["A4"].font = Font(bold=True, color="204534")
        dv = DataValidation(type="list", formula1='"Historical,Forecast"', allow_blank=True)
        ws.add_data_validation(dv)
        letters = []
        for i in range(ncols):
            c = FIRST_COL + i
            letter = get_column_letter(c); letters.append(letter)
            _input(ws[f"{letter}3"]); ws[f"{letter}3"].number_format = "General"
            _input(ws[f"{letter}4"]); ws[f"{letter}4"].number_format = "0"
            ws[f"{letter}3"] = "Historical"
            ws[f"{letter}4"] = sample_periods[i]            # sample period label
            dv.add(ws[f"{letter}3"])
            ws.column_dimensions[letter].width = 14
        rowmap = {}
        for r, key in enumerate(keys, start=5):
            ws[f"A{r}"] = lab["lines"][key]
            rowmap[key] = r
            for i, letter in enumerate(letters):
                c = ws[colref(letter, r)]
                _input(c)
                c.value = sample[block][key][str(sample_periods[i])]   # sample figure
        # locked subtotal formulas
        sub_start = 5 + len(keys) + 1
        alias = {"rev": rowmap.get("revenue"), "cogs": rowmap.get("cogs"),
                 "opex": rowmap.get("opex"), "da": rowmap.get("depreciation_amortization"),
                 "cash": rowmap.get("cash"), "oca": rowmap.get("other_current_assets"),
                 "nca": rowmap.get("noncurrent_assets"),
                 "cle": rowmap.get("current_liabilities_ex_debt"),
                 "std": rowmap.get("short_term_debt"), "ltd": rowmap.get("long_term_debt"),
                 "pfd": rowmap.get("preferred_equity"), "mi": rowmap.get("minority_interest"),
                 "te": rowmap.get("total_equity")}
        for j, (label, ftmpl) in enumerate(SUBTOTALS.get(block, [])):
            rr = sub_start + j
            ws[f"A{rr}"] = label; ws[f"A{rr}"].font = Font(italic=True, bold=True)
            for letter in letters:
                try:
                    f = ftmpl.format(**{k: f"{letter}{v}" for k, v in alias.items() if v})
                    _locked_formula(ws[colref(letter, rr)], f)
                except Exception:
                    pass
        ws.column_dimensions["A"].width = 56
        ws.protection.sheet = True
        ws.protection.password = _LOCK_PWD

    # v5.1: every strategy sheet carries ROW_CAPACITY (200) pre-formatted input
    # rows; dropdowns are attached to the FULL column RANGE (not per seeded cell)
    # so a pasted block of 100+ rows stays constrained by validation.
    obj_ids = [f"O{i}" for i in range(1, OBJ_ID_MAX + 1)]
    obj_last = OBJ_DATA_START + ROW_CAPACITY - 1

    # ---- (§9 OKR) Objectives ----
    #   A Objective · B Owner (CXO) · C Priority · D Horizon · E Status · F Objective ID
    ws = wb.create_sheet(OBJECTIVES_SHEET)
    ws["A1"] = ("Your objectives (the O in OKR). Objective, Owner and Priority are recommended; "
                "Horizon Short ≤12m · Medium ≤36m · Long >36m; Status is current health R/A/G. "
                "The Objective ID links each row to its Key Results on the next sheet. "
                "Paste as many rows as you need — up to 200.")
    ws["A1"].font = Font(italic=True, color="446655")
    obj_hdrs = ["Objective", "Owner (CXO)", "Priority", "Horizon", "Status (optional)", "Objective ID"]
    for i, h in enumerate(obj_hdrs):
        _hdr(ws.cell(row=OBJ_HEADER_ROW, column=1 + i), h, bg=_INK, fg=_ACCENT)
    dv_prio = DataValidation(type="list", formula1='"%s"' % ",".join(OBJ_PRIORITIES), allow_blank=True)
    dv_hor = DataValidation(type="list", formula1='"%s"' % ",".join(OBJ_HORIZONS), allow_blank=True)
    dv_stat = DataValidation(type="list", formula1='"%s"' % ",".join(OBJ_STATUSES), allow_blank=True)
    ws.add_data_validation(dv_prio); ws.add_data_validation(dv_hor); ws.add_data_validation(dv_stat)
    dv_prio.add(f"C{OBJ_DATA_START}:C{obj_last}")               # paste-safe: whole-column ranges
    dv_hor.add(f"D{OBJ_DATA_START}:D{obj_last}")
    dv_stat.add(f"E{OBJ_DATA_START}:E{obj_last}")
    for idx, r in enumerate(range(OBJ_DATA_START, obj_last + 1)):
        for c in range(1, 7):
            _input(ws.cell(row=r, column=c)); ws.cell(row=r, column=c).number_format = "General"
        ws.cell(row=r, column=6).value = obj_ids[idx]           # pre-seeded stable short code O1…O200
    for col, w in zip("ABCDEF", (50, 20, 12, 12, 14, 14)):
        ws.column_dimensions[col].width = w
    ws.protection.sheet = True; ws.protection.password = _LOCK_PWD

    # ---- (§9 OKR) Key Results ----
    #   A Objective ID · B Key Result · C Unit · D Baseline · E Target · F Current · G Due date
    ws = wb.create_sheet(KR_SHEET)
    ws["A1"] = ("Key Results (the KR in OKR) — measurable outcomes for each objective. Aim for "
                "At least 1 Key Result per objective (more is better). Objective ID must match a row on the Objectives sheet. "
                "Progress = (Current − Baseline) ÷ (Target − Baseline). Up to 200 rows.")
    ws["A1"].font = Font(italic=True, color="446655")
    kr_hdrs = ["Objective ID", "Key Result", "Unit", "Baseline", "Target", "Current", "Due date"]
    for i, h in enumerate(kr_hdrs):
        _hdr(ws.cell(row=KR_HEADER_ROW, column=1 + i), h, bg=_INK, fg=_ACCENT)
    kr_last = KR_DATA_START + ROW_CAPACITY - 1
    dv_objid = DataValidation(type="list", formula1='"%s"' % ",".join(obj_ids), allow_blank=True)
    ws.add_data_validation(dv_objid)
    dv_objid.add(f"A{KR_DATA_START}:A{kr_last}")                # O1…O200 dropdown over the whole column
    for r in range(KR_DATA_START, kr_last + 1):
        for c in range(1, 8):
            _input(ws.cell(row=r, column=c))
        ws.cell(row=r, column=1).number_format = "General"
        for c in (4, 5, 6):
            ws.cell(row=r, column=c).number_format = "0.00"
        ws.cell(row=r, column=7).number_format = "General"
    for col, w in zip("ABCDEFG", (13, 44, 12, 12, 12, 12, 14)):
        ws.column_dimensions[col].width = w
    ws.protection.sheet = True; ws.protection.password = _LOCK_PWD

    # ---- (§4o) KPI Plan vs Actual ----
    ws = wb.create_sheet(KPI_SHEET)
    ws["A1"] = ("Plan vs actual for your headline KPIs. The four standard rows are seeded — "
                "fill YTD Plan / YTD Actual / Full-year Target; add your own KPIs below. Up to 200 rows.")
    ws["A1"].font = Font(italic=True, color="446655")
    kpi_hdrs = ["KPI name", "Unit", "YTD Plan", "YTD Actual", "Full-year Target"]
    for i, h in enumerate(kpi_hdrs):
        _hdr(ws.cell(row=KPI_HEADER_ROW, column=1 + i), h, bg=_INK, fg=_ACCENT)
    kpi_last = KPI_DATA_START + ROW_CAPACITY - 1
    for i, r in enumerate(range(KPI_DATA_START, kpi_last + 1)):
        for c in range(1, 3):
            _input(ws.cell(row=r, column=c)); ws.cell(row=r, column=c).number_format = "General"
        for c in (3, 4, 5):
            _input(ws.cell(row=r, column=c)); ws.cell(row=r, column=c).number_format = "0.00"
        if i < len(KPI_SEED_ROWS):                              # seed the four standard KPIs
            ws.cell(row=r, column=1).value = KPI_SEED_ROWS[i]
            ws.cell(row=r, column=2).value = "%"
    for col, w in zip("ABCDE", (34, 12, 16, 16, 18)):
        ws.column_dimensions[col].width = w
    ws.protection.sheet = True; ws.protection.password = _LOCK_PWD

    # ---- hidden metadata sheet (self-identifying upload) ----
    ws = wb.create_sheet("_AXIOM")
    ws["A1"] = META_SIG
    ws["A2"] = "company_id"; ws["B2"] = company_id
    ws["A3"] = "frequency"; ws["B3"] = frequency
    ws["A4"] = "template_version"; ws["B4"] = TEMPLATE_VERSION
    ws["A5"] = "standard"; ws["B5"] = standard
    ws["A6"] = "statement_units"; ws["B6"] = statement_units
    ws.sheet_state = "hidden"
    ws.protection.sheet = True
    ws.protection.password = _LOCK_PWD

    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def read_upload_metadata(content: bytes) -> dict | None:
    """Return {company_id, frequency, template_version, standard} from the
    hidden _AXIOM sheet, or None if absent/unreadable."""
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        return None
    if "_AXIOM" not in wb.sheetnames:
        return None
    ws = wb["_AXIOM"]
    if ws["A1"].value != META_SIG:
        return None
    try:
        units = ws["B6"].value
        return {"company_id": int(ws["B2"].value),
                "frequency": str(ws["B3"].value or "annual"),
                "template_version": str(ws["B4"].value or ""),
                "standard": str(ws["B5"].value or "us_gaap"),
                "statement_units": (str(units).strip().lower() if units else None)}
    except (TypeError, ValueError):
        return None


def _sheet_for(engine_error: str, lab: dict) -> str | None:
    for block in ("income_statement", "balance_sheet", "cash_flow"):
        if engine_error.startswith(block):
            return lab["sheets"][block]
    if engine_error.startswith("company."):
        return "Company"
    if engine_error.startswith("periods"):
        return lab["sheets"]["income_statement"]
    return None


def parse_and_validate(content: bytes, expected_company_id: int,
                       statement_units: str | None = None):
    """Parse + validate an uploaded company workbook into the canonical dataset.

    Returns (data|None, errors, meta, warnings). `errors` is a structured list
    of {sheet, cell, message}; when non-empty NOTHING should be written. On
    success `data` is the engine-ready canonical dataset with statement figures
    normalized to the canonical MILLIONS scale (honoring statement_units).
    `statement_units` is the caller's fallback when the template metadata omits
    it (older templates)."""
    errors = []
    meta = read_upload_metadata(content)
    if meta is None:
        return None, [{"sheet": "_AXIOM", "cell": None,
                       "message": "Not an AXIOM company template (metadata sheet "
                                  "missing or altered). Download a fresh template."}], None, []
    if meta["company_id"] != expected_company_id:
        return None, [{"sheet": "_AXIOM", "cell": "B2",
                       "message": f"This template was generated for company "
                                  f"{meta['company_id']}, not {expected_company_id}."}], meta, []
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        return None, [{"sheet": None, "cell": None,
                       "message": f"not a readable .xlsx file: {e}"}], meta, []

    standard = meta.get("standard", "us_gaap")
    if standard not in LABELS:
        standard = "us_gaap"
    lab = LABELS[standard]

    # ---- Company profile ----
    if "Company" not in wb.sheetnames:
        return None, [{"sheet": "Company", "cell": None,
                       "message": "missing Company sheet"}], meta, []
    ws = wb["Company"]
    company = {"standard": standard}
    for r, (field, label, applies) in enumerate(COMPANY_ROWS, start=2):
        v = ws[f"B{r}"].value
        if isinstance(v, str):
            v = v.strip()
        company[field] = v if v not in ("", None) else None
    if isinstance(company.get("ownership"), str):
        company["ownership"] = company["ownership"].lower()
    for r, (field, label, applies) in enumerate(COMPANY_ROWS, start=2):
        if field in ("name", "ownership", "currency"):
            continue
        val = company.get(field)
        if val is None:
            continue
        try:
            company[field] = float(val)
        except (TypeError, ValueError):
            errors.append({"sheet": "Company", "cell": f"B{r}",
                           "message": f"'{label}' must be numeric"})

    # ---- Statement sheets ----
    def read_cols(ws):
        cols = []
        for i in range(30):
            letter = get_column_letter(FIRST_COL + i)
            y, k = ws[f"{letter}4"].value, ws[f"{letter}3"].value
            if y in (None, ""):
                continue
            try:
                y = int(y)
            except (TypeError, ValueError):
                errors.append({"sheet": ws.title, "cell": f"{letter}4",
                               "message": "period must be an integer"})
                continue
            kind = str(k or "").strip().lower()
            if kind not in ("historical", "forecast"):
                errors.append({"sheet": ws.title, "cell": f"{letter}3",
                               "message": "mark this column Historical or Forecast"})
                continue
            cols.append((letter, y, kind))
        return cols

    blocks, ref_cols = {}, None
    for block, keys in BLOCK_KEYS.items():
        name = lab["sheets"][block]
        if name not in wb.sheetnames:
            errors.append({"sheet": name, "cell": None,
                           "message": f"missing sheet '{name}'"})
            continue
        ws = wb[name]
        cols = read_cols(ws)
        if ref_cols is None:
            ref_cols = cols
        elif [(y, k) for _, y, k in cols] != [(y, k) for _, y, k in ref_cols]:
            errors.append({"sheet": name, "cell": None,
                           "message": "period columns must match the "
                                      f"'{lab['sheets']['income_statement']}' sheet"})
        bd = {}
        for r, key in enumerate(keys, start=5):
            row = {}
            for letter, y, kind in cols:
                v = ws[f"{letter}{r}"].value
                if v in (None, ""):
                    errors.append({"sheet": name, "cell": f"{letter}{r}",
                                   "message": f"'{lab['lines'][key]}' — value required for period {y}"})
                    continue
                try:
                    row[str(y)] = float(v)
                except (TypeError, ValueError):
                    errors.append({"sheet": name, "cell": f"{letter}{r}",
                                   "message": f"'{lab['lines'][key]}' — must be numeric"})
            bd[key] = row
        blocks[block] = bd

    ref_cols = ref_cols or []
    hist = [y for _, y, k in ref_cols if k == "historical"]
    fcst = [y for _, y, k in ref_cols if k == "forecast"]
    data = {"company": company,
            "periods": {"historical": hist, "forecast": fcst},
            "income_statement": blocks.get("income_statement", {}),
            "balance_sheet": blocks.get("balance_sheet", {}),
            "cash_flow": blocks.get("cash_flow", {})}

    # cross-field checks (required company fields, increasing periods)
    v = engines.validate_dataset(data)
    for e in v["errors"]:
        errors.append({"sheet": _sheet_for(e, lab), "cell": None, "message": e})
    # BS balance is a HARD error on upload (the engine treats it as a warning);
    # drop the duplicate warning.
    warnings = [w for w in v["warnings"] if "does not balance" not in w]
    bs = data["balance_sheet"]
    bs_sheet = lab["sheets"]["balance_sheet"]
    for y in hist + fcst:
        ys = str(y)
        try:
            assets = (bs["cash"][ys] + bs["other_current_assets"][ys]
                      + bs["noncurrent_assets"][ys])
            le = (bs["current_liabilities_ex_debt"][ys] + bs["short_term_debt"][ys]
                  + bs["long_term_debt"][ys] + bs["preferred_equity"][ys]
                  + bs["minority_interest"][ys] + bs["total_equity"][ys])
        except KeyError:
            continue                       # missing cells already flagged above
        if assets and abs(assets - le) > 0.005 * abs(assets):
            errors.append({"sheet": bs_sheet, "cell": None,
                           "message": f"balance sheet does not balance in {y}: "
                                      f"assets {assets:,.2f} vs liabilities+equity "
                                      f"{le:,.2f} (must match within 0.5%)"})

    # resolve the declared unit (template metadata wins; caller is the fallback)
    units = ((meta or {}).get("statement_units")
             or (statement_units or "").strip().lower() or "millions")
    if units not in UNIT_SCALE:
        units = "millions"
    unit_label = {"actual": "actual amounts", "thousands": "thousands",
                  "millions": "millions"}[units]

    # ---- (b) sentinel: reject a workbook still holding the template SAMPLE data ----
    try:
        sample = build_sample_data(company.get("ownership") or "private", standard,
                                   (meta or {}).get("frequency", "annual"))
        for block in ("income_statement", "balance_sheet", "cash_flow"):
            up_vals = sorted(v for row in data.get(block, {}).values() for v in row.values())
            sm_vals = sorted(v for row in sample.get(block, {}).values() for v in row.values())
            if up_vals and up_vals == sm_vals:
                errors.append({"sheet": lab["sheets"][block], "cell": None,
                               "message": f"The '{lab['sheets'][block]}' sheet still contains the "
                                          "template's sample figures. Replace every sheet with your "
                                          "company's own numbers before uploading."})
    except Exception:
        pass

    # ---- (b) cross-sheet magnitude sanity (catches mixed-scale uploads) ----
    try:
        yref = str((hist or fcst)[-1])
        rev = data["income_statement"]["revenue"][yref]
        ta = (bs["cash"][yref] + bs["other_current_assets"][yref] + bs["noncurrent_assets"][yref])
        if rev and ta:
            turn = rev / ta
            if turn > 50 or turn < 0.01:
                errors.append({"sheet": None, "cell": None,
                               "message": f"Income-statement and balance-sheet figures look like they "
                                          f"are on different scales: revenue is {turn:,.0f}x total assets "
                                          f"in {yref}, an asset turnover of {turn:,.1f}x that is not "
                                          f"plausible. Check that every sheet uses the same units "
                                          f"({unit_label})."})
            elif turn > 20 or turn < 0.04:
                warnings.append(f"Unusual asset turnover in {yref} ({turn:,.1f}x) — please verify every "
                                f"sheet uses the same units ({unit_label}).")
    except (KeyError, ZeroDivisionError, IndexError, TypeError):
        pass

    if errors:
        return None, errors, meta, warnings

    # ---- (a) normalize statement figures to the canonical MILLIONS scale ----
    factor = UNIT_SCALE.get(units, 1.0)
    if factor != 1.0:
        for block in ("income_statement", "balance_sheet", "cash_flow"):
            for row in data[block].values():
                for ys in list(row):
                    row[ys] = row[ys] * factor
    return data, [], meta, warnings


def _cell_str(v):
    if v is None:
        return ""
    return str(v).strip()


def _norm_priority(v):
    s = _cell_str(v).lower()
    for p in GOAL_PRIORITIES:
        if s == p.lower():
            return p
    return None


def _norm_horizon(v):
    """Accept the canonical word ('Short') or a legacy label ('Short (≤12m)')."""
    s = _cell_str(v).lower()
    first = s.split("(")[0].split()[0] if s else ""
    for h in GOAL_HORIZONS:
        if first == h.lower():
            return h
    return None


def _norm_status(v):
    """Optional RAG health — blank is allowed (returns None, no error)."""
    s = _cell_str(v).lower()
    if not s:
        return None
    for st in GOAL_STATUSES:
        if s == st.lower():
            return st
    return "__invalid__"    # sentinel so the caller can flag a bad non-blank value


def parse_okr_and_kpis(content: bytes):
    """Parse the §9 OKR sheets (Objectives + Key Results) + the KPI sheet. Returns
    (objectives, key_results, kpis, errors, warnings, flags). Malformed enums /
    non-numeric figures BLOCK (errors, sentinel pattern); the ≥3-KR guidance and
    orphan-KR references only WARN. A legacy 'Organizational Goals' sheet is read
    as objectives with no KRs, so v4 uploads migrate honestly. Blank key columns
    (Objective / Key Result / KPI name) are skipped, so notes/blanks are tolerated."""
    errors, warnings = [], []
    objectives, key_results, kpis = [], [], []
    flags = {"has_objectives": False, "has_krs": False, "has_kpis": False, "legacy": False}
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        return [], [], [], [{"sheet": None, "cell": None,
                             "message": f"not a readable .xlsx file: {e}"}], [], flags

    has_obj = OBJECTIVES_SHEET in wb.sheetnames
    has_legacy = (not has_obj) and (GOALS_SHEET in wb.sheetnames)
    has_krs = KR_SHEET in wb.sheetnames
    has_kpis = KPI_SHEET in wb.sheetnames

    # ---- Objectives (new sheet: A obj · B owner · C priority · D horizon · E status · F id) ----
    if has_obj:
        ws = wb[OBJECTIVES_SHEET]
        seen, auto = set(), 0
        for r in range(OBJ_DATA_START, (ws.max_row or OBJ_DATA_START) + 1):
            obj = _cell_str(ws.cell(row=r, column=1).value)
            if not obj:
                continue
            priority = _norm_priority(ws.cell(row=r, column=3).value)
            horizon = _norm_horizon(ws.cell(row=r, column=4).value)
            status = _norm_status(ws.cell(row=r, column=5).value)
            oid = _cell_str(ws.cell(row=r, column=6).value)
            if priority is None:
                errors.append({"sheet": OBJECTIVES_SHEET, "cell": f"C{r}",
                               "message": f"'{obj[:40]}' — Priority must be one of {', '.join(OBJ_PRIORITIES)}."})
            if horizon is None:
                errors.append({"sheet": OBJECTIVES_SHEET, "cell": f"D{r}",
                               "message": f"'{obj[:40]}' — Horizon must be one of {', '.join(OBJ_HORIZONS)}."})
            if status == "__invalid__":
                errors.append({"sheet": OBJECTIVES_SHEET, "cell": f"E{r}",
                               "message": f"'{obj[:40]}' — Status must be one of {', '.join(OBJ_STATUSES)} or blank."})
                status = None
            if not oid:
                auto += 1; oid = f"O{auto}"
            if oid in seen:
                errors.append({"sheet": OBJECTIVES_SHEET, "cell": f"F{r}",
                               "message": f"Objective ID '{oid}' is used more than once — IDs must be unique."})
            seen.add(oid)
            objectives.append({"row_index": r, "objective": obj,
                               "owner": _cell_str(ws.cell(row=r, column=2).value) or None,
                               "priority": priority, "horizon": horizon, "status": status,
                               "objective_id": oid})
    elif has_legacy:
        # v4 'Organizational Goals' (A goal · B prio · C horizon · D status · E owner) → objectives, no KRs
        ws = wb[GOALS_SHEET]
        auto = 0
        for r in range(OBJ_DATA_START, (ws.max_row or OBJ_DATA_START) + 1):
            goal = _cell_str(ws.cell(row=r, column=1).value)
            if not goal:
                continue
            auto += 1
            priority = _norm_priority(ws.cell(row=r, column=2).value)
            horizon = _norm_horizon(ws.cell(row=r, column=3).value)
            status = _norm_status(ws.cell(row=r, column=4).value)
            if status == "__invalid__":
                status = None
            if priority is None:
                errors.append({"sheet": GOALS_SHEET, "cell": f"B{r}",
                               "message": f"'{goal[:40]}' — Priority must be one of {', '.join(OBJ_PRIORITIES)}."})
            if horizon is None:
                errors.append({"sheet": GOALS_SHEET, "cell": f"C{r}",
                               "message": f"'{goal[:40]}' — Horizon must be one of {', '.join(OBJ_HORIZONS)}."})
            objectives.append({"row_index": r, "objective": goal,
                               "owner": _cell_str(ws.cell(row=r, column=5).value) or None,
                               "priority": priority, "horizon": horizon, "status": status,
                               "objective_id": f"O{auto}"})

    obj_ids = {o["objective_id"] for o in objectives}

    # ---- Key Results (A id · B kr · C unit · D baseline · E target · F current · G due) ----
    if has_krs:
        ws = wb[KR_SHEET]
        def _num(v, col, r, label):
            if v in (None, ""):
                return None
            try:
                return float(v)
            except (TypeError, ValueError):
                errors.append({"sheet": KR_SHEET, "cell": f"{col}{r}",
                               "message": f"'{label}' — must be a number (got '{v}')."})
                return None
        for r in range(KR_DATA_START, (ws.max_row or KR_DATA_START) + 1):
            kr = _cell_str(ws.cell(row=r, column=2).value)
            if not kr:
                continue
            oid = _cell_str(ws.cell(row=r, column=1).value)
            if not oid:
                errors.append({"sheet": KR_SHEET, "cell": f"A{r}",
                               "message": f"Key Result '{kr[:40]}' needs an Objective ID."})
            elif obj_ids and oid not in obj_ids:
                warnings.append(f"Key Result '{kr[:40]}' references Objective ID '{oid}', "
                                "which is not on the Objectives sheet.")
            key_results.append({"row_index": r, "objective_id": oid or None, "key_result": kr,
                                "unit": _cell_str(ws.cell(row=r, column=3).value) or None,
                                "baseline": _num(ws.cell(row=r, column=4).value, "D", r, kr + " Baseline"),
                                "target": _num(ws.cell(row=r, column=5).value, "E", r, kr + " Target"),
                                "current": _num(ws.cell(row=r, column=6).value, "F", r, kr + " Current"),
                                "due_date": _cell_str(ws.cell(row=r, column=7).value) or None})

    # ≥1 KR per objective (hard rule) — WARN only, never blocks
    if objectives:
        cnt = {}
        for kr in key_results:
            if kr["objective_id"]:
                cnt[kr["objective_id"]] = cnt.get(kr["objective_id"], 0) + 1
        for o in objectives:
            n = cnt.get(o["objective_id"], 0)
            if n < MIN_KRS_PER_OBJECTIVE:
                warnings.append(f"Objective '{o['objective'][:40]}' has {n} key result"
                                f"{'' if n == 1 else 's'} (aim for ≥{MIN_KRS_PER_OBJECTIVE}).")

    # ---- KPIs (unchanged) ----
    if has_kpis:
        ws = wb[KPI_SHEET]
        def _knum(v, col, r, label):
            if v in (None, ""):
                return None
            try:
                return float(v)
            except (TypeError, ValueError):
                errors.append({"sheet": KPI_SHEET, "cell": f"{col}{r}",
                               "message": f"'{label}' — must be a number (got '{v}')."})
                return None
        for r in range(KPI_DATA_START, (ws.max_row or KPI_DATA_START) + 1):
            name = _cell_str(ws.cell(row=r, column=1).value)
            if not name:
                continue
            kpis.append({"row_index": r, "kpi_name": name,
                         "unit": _cell_str(ws.cell(row=r, column=2).value) or None,
                         "ytd_plan": _knum(ws.cell(row=r, column=3).value, "C", r, name + " YTD Plan"),
                         "ytd_actual": _knum(ws.cell(row=r, column=4).value, "D", r, name + " YTD Actual"),
                         "full_year_target": _knum(ws.cell(row=r, column=5).value, "E", r, name + " Full-year Target")})

    flags = {"has_objectives": has_obj or has_legacy, "has_krs": has_krs,
             "has_kpis": has_kpis, "legacy": has_legacy}
    return objectives, key_results, kpis, errors, warnings, flags
