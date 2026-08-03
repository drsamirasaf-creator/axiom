"""Locked GAAP/IFRS input templates + upload parser (ADR-005 §2).

The workbook lock (Product §7.8 guided editing) is UX guidance only —
spreadsheet protection is advisory and trivially removable — so the parser
below re-validates every label and every cell server-side; the validator,
not the lock, is the integrity guarantee. Templates are the deterministic
v0 of the spec's Intelligent Financial Mapping (Product §7.9/§7.10): free-
form import with AI account mapping is the roadmap successor, not v0.
"""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from . import engines
from . import template_policy as policy

# ⭐ THE FAMILY IDENTIFIES; THE VERSION IS FORENSIC METADATA. CORE §7.37 (user
# ruling, 28 Jul): "AXIOM does not track or control template versions as a
# precondition for upload. Any template that parses is accepted. Version is
# never a gate — on either path."
#
# That lane removed ACCEPTED_TEMPLATE_VERSIONS from ingest.py. THIS SIBLING GATE
# SURVIVED IT: the check below was `sig.startswith(TEMPLATE_SIG)` with the
# version baked into the constant, so bumping the constant to v8 would have made
# every workbook a customer downloaded before the bump fail to parse — with the
# error "not an AXIOM financial template", which is false. Same
# fix-one-site-and-stop pattern that cost three sites on the None-arithmetic
# class this week.
#
# The parser keys on sheet and row labels and never needed the version to work.
# So the gate is the FAMILY; the version is read out and returned as metadata.
# Kept as names because other modules and tests read them; they are now VIEWS
# onto template_policy, not second definitions.
TEMPLATE_FAMILY = policy.GENERIC_FAMILY
TEMPLATE_VERSION = policy.version("generic")
TEMPLATE_SIG = f"{TEMPLATE_FAMILY} {TEMPLATE_VERSION}"
# ⭐ 20 COULD NOT EXPRESS THE PLAN THE ENGINE ACCEPTS. engines.MAX_FORECAST_PERIODS
# is {"annual": 15, "quarterly": 40} and ingest.FORECAST_QUARTERLY is 40, while
# this download offered ten forecast columns — so a customer who took the generic
# template literally could not supply a quarterly plan the backend would have
# read. 56 = 1 opening + 15 historical + 40 forecast, which covers the widest
# case with no second limit to keep in sync.
MAX_YEAR_COLS = policy.max_year_cols()
FIRST_YEAR_COL = 2          # column B

LABELS = {
    "us_gaap": {
        "sheets": {"income_statement": "Income Statement",
                   "balance_sheet": "Balance Sheet",
                   "cash_flow": "Cash Flow Data"},
        "lines": {
            "revenue": "Revenue", "cogs": "Cost of Goods Sold",
            "opex": "Operating Expenses (excl. D&A)",
            "depreciation_amortization": "Depreciation & Amortization",
            "interest_expense": "Interest Expense",
            "cash": "Cash & Equivalents",
            "other_current_assets": "Other Current Assets (Receivables, Inventory, etc.)",
            "receivables": "of which: Accounts Receivable",
            "inventory": "of which: Inventory",
            "noncurrent_assets": "Total Non-Current Assets",
            "property_plant_equipment_net": "Property, Plant & Equipment, net",
            "goodwill": "Goodwill",
            "intangible_assets_net": "Intangible Assets, net (excl. Goodwill)",
            "long_term_investments": "Long-Term Investments",
            "other_noncurrent_assets": "Other Non-Current Assets",
            "current_liabilities_ex_debt": "Current Liabilities (excl. Debt)",
            "payables": "of which: Accounts Payable",
            "other_noncurrent_liabilities": "Other Non-Current Liabilities",
            "short_term_debt": "Short-Term Debt",
            "long_term_debt": "Long-Term Debt",
            "preferred_equity": "Preferred Equity",
            "minority_interest": "Noncontrolling (Minority) Interest",
            "total_equity": "Total Stockholders' Equity",
            "capex": "Capital Expenditures",
            "net_borrowing": "Net Borrowing (Issuance - Repayment)",
            "dividends": "Dividends Paid"}},
    "ifrs": {
        "sheets": {"income_statement": "Statement of Profit or Loss",
                   "balance_sheet": "Statement of Financial Position",
                   "cash_flow": "Cash Flow Data"},
        "lines": {
            "revenue": "Revenue", "cogs": "Cost of Sales",
            "opex": "Operating Expenses (excl. D&A)",
            "depreciation_amortization": "Depreciation & Amortisation",
            "interest_expense": "Finance Costs",
            "cash": "Cash & Cash Equivalents",
            "other_current_assets": "Other Current Assets (Trade Receivables, Inventories, etc.)",
            "receivables": "of which: Trade & Other Receivables",
            "inventory": "of which: Inventories",
            "noncurrent_assets": "Total Non-Current Assets",
            "property_plant_equipment_net": "Property, Plant and Equipment (net)",
            "goodwill": "Goodwill",
            "intangible_assets_net": "Intangible Assets (excl. Goodwill)",
            "long_term_investments": "Non-Current Financial Assets",
            "other_noncurrent_assets": "Other Non-Current Assets",
            "current_liabilities_ex_debt": "Current Liabilities (excl. Borrowings)",
            "payables": "of which: Trade & Other Payables",
            "other_noncurrent_liabilities": "Other Non-Current Liabilities",
            "short_term_debt": "Current Borrowings",
            "long_term_debt": "Non-Current Borrowings",
            "preferred_equity": "Preference Shares",
            "minority_interest": "Non-Controlling Interests",
            "total_equity": "Total Equity Attributable to Owners",
            "capex": "Purchases of Property, Plant & Equipment (CapEx)",
            "net_borrowing": "Net Borrowing (Proceeds - Repayments)",
            "dividends": "Dividends Paid"}},
}

COMPANY_ROWS = [  # (field, label, applies)
    ("name", "Company Name", "all"),
    ("ownership", "Ownership (public / private)", "all"),
    ("currency", "Reporting Currency", "all"),
    ("tax_rate", "Effective Tax Rate (decimal, e.g. 0.25)", "all"),
    # ⭐ RULE 1 NEEDS TWO CELLS, NOT ONE. The precedence is
    # admin override > template POLICY rate > IMPLIED EFFECTIVE rate. With a
    # single cell the middle and bottom sources are the same number, the
    # precedence cannot be expressed, and the explainer's provenance field would
    # have nothing truthful to stamp. Neither is required: absent policy rate
    # falls through to the effective rate, and the explainer says which was used.
    ("tax_rate_policy", "Policy / Statutory Tax Rate (decimal)", "all"),
    ("risk_free_rate", "Risk-Free Rate (decimal)", "all"),
    ("market_risk_premium", "Market Risk Premium (decimal)", "all"),
    ("cost_of_debt", "Pre-Tax Cost of Debt (decimal)", "all"),
    # ⭐⭐ THE UNIT IS STATED. Left unstated, a client typed a raw count into a
    # field the engine read as millions and the per-share figure came out a
    # million times small (§7w). The label is the cheapest place to end that.
    ("shares_outstanding",
     "Shares Outstanding (actual number of shares, not millions)", "all"),
    ("share_price", "Share Price (public only)", "public"),
    ("beta", "Equity Beta (public only)", "public"),
    ("unlevered_industry_beta", "Unlevered Industry Beta (private only)", "private"),
    ("target_debt_to_equity", "Target Debt/Equity (private only)", "private"),
    ("size_premium", "Size Premium (decimal, private only)", "private"),
    ("specific_risk_premium", "Company-Specific Risk Premium (private only)", "private"),
    ("dlom", "Discount for Lack of Marketability (decimal, private only)", "private"),
]

BLOCK_KEYS = {"income_statement": engines.IS_KEYS,
              "balance_sheet": engines.BS_KEYS,
              "cash_flow": engines.CF_KEYS}

_HDR = PatternFill("solid", fgColor="1F3B57")
_IN = PatternFill("solid", fgColor="FFF7E0")
_LOCK_PWD = "AXIOM"


def _style_header(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = _HDR
    cell.alignment = Alignment(horizontal="left")


def _input_cell(cell, numeric=True):
    cell.protection = Protection(locked=False)
    cell.fill = _IN
    if numeric:
        cell.number_format = "#,##0.00"


def build_template(standard: str) -> bytes:
    """Build the locked input workbook for a standard, in memory."""
    if standard not in LABELS:
        raise KeyError(standard)
    lab = LABELS[standard]
    wb = Workbook()

    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = f"{TEMPLATE_SIG} {standard}"
    ws["A1"].font = Font(bold=True, size=9, color="888888")
    ws["A3"] = "AXIOM Financial Input Template — " + \
        ("US GAAP" if standard == "us_gaap" else "IFRS")
    ws["A3"].font = Font(bold=True, size=14)
    for r, line in enumerate([
        "1. Fill the Company sheet, then the three statement sheets.",
        "2. Enter years across row 4 of each statement sheet and mark each",
        "   column Historical or Forecast in row 3. At least one historical",
        "   year is required; forecast years are optional (up to 40 quarters",
        "   or 15 years).",
        "2b. BALANCE SHEET ONLY: one column may be marked Opening — the balance",
        "   carried into your first historical period. It lets AXIOM average",
        "   opening and closing capital for that period instead of falling back",
        "   to the opening balance alone. Optional; leave it out and AXIOM says",
        "   on the surface which basis it used.",
        "3. Only the highlighted cells accept input; all labels are locked.",
        "4. Enter rates as decimals (7% = 0.07). Amounts in one currency unit",
        "   (e.g. millions) used consistently throughout.",
        "5. Upload the completed file at POST /api/v1/financials/datasets/upload.",
        "   AXIOM re-validates every cell on upload; the workbook lock is a",
        "   guide, the server-side validator is the guarantee.",
    ], start=5):
        ws[f"A{r}"] = line
    ws.column_dimensions["A"].width = 78
    ws.protection.sheet = True
    ws.protection.password = _LOCK_PWD

    # ⭐ DROPDOWNS COME FROM DEFINED NAMES, NOT INLINE STRING LISTS. Standing
    # Excel rule. An inline `formula1='"Historical,Forecast"'` is capped near 255
    # characters and is invisible to anyone auditing the workbook; a named range
    # is inspectable and editable in one place.
    #
    # ⭐ AND THE LIST SHEET IS VISIBLE, DELIBERATELY. CORE §7.36/§7.37: the
    # participant template stamped its version as a workbook-global defined name
    # pointing at a HIDDEN sheet, and Excel and Google Sheets both re-scope or
    # drop those — the stamp vanished in the customer's editor and legitimate
    # files were rejected. Defined names are used here for the dropdown lists
    # only, never as a stamp, and the sheet they point at is not hidden.
    lw = wb.create_sheet("Lists")
    _style_header(lw["A1"], "Ownership")
    _style_header(lw["B1"], "Period Type")
    _style_header(lw["C1"], "Period Type (Balance Sheet)")
    for i, v in enumerate(("public", "private"), start=2):
        lw[f"A{i}"] = v
    for i, v in enumerate(("Historical", "Forecast"), start=2):
        lw[f"B{i}"] = v
    for i, v in enumerate(("Opening", "Historical", "Forecast"), start=2):
        lw[f"C{i}"] = v
    for col, w in (("A", 18), ("B", 18), ("C", 28)):
        lw.column_dimensions[col].width = w
    wb.defined_names.add(DefinedName("OWNERSHIP", attr_text="Lists!$A$2:$A$3"))
    wb.defined_names.add(DefinedName("PERIODTYPE", attr_text="Lists!$B$2:$B$3"))
    wb.defined_names.add(DefinedName("PERIODTYPE_BS", attr_text="Lists!$C$2:$C$4"))
    lw.protection.sheet = True
    lw.protection.password = _LOCK_PWD

    ws = wb.create_sheet("Company")
    _style_header(ws["A1"], "Company Profile")
    _style_header(ws["B1"], "Value")
    for r, (field, label, applies) in enumerate(COMPANY_ROWS, start=2):
        ws[f"A{r}"] = label
        _input_cell(ws[f"B{r}"], numeric=(field not in
                                          ("name", "ownership", "currency")))
    dv = DataValidation(type="list", formula1="=OWNERSHIP", allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["B3"])   # ownership row
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 22
    ws.protection.sheet = True
    ws.protection.password = _LOCK_PWD

    for block, keys in BLOCK_KEYS.items():
        ws = wb.create_sheet(lab["sheets"][block])
        _style_header(ws["A1"], lab["sheets"][block])
        # ⭐ THE OPENING COLUMN IS BALANCE-SHEET ONLY. Income statement and cash
        # flow are FLOW statements — an opening column there is meaningless and
        # would invite garbage. On the balance sheet it is what makes rule 3's
        # average basis (opening + closing)/2 computable for the earliest
        # period; without it year one falls back to BOP on every dataset
        # forever, and the "computed on BOP" label becomes permanent furniture
        # rather than a signal.
        is_bs = block == "balance_sheet"
        ws["A3"] = ("Period Type (Opening / Historical / Forecast)" if is_bs
                    else "Period Type (Historical / Forecast)")
        ws["A4"] = "Year"
        ws["A3"].font = ws["A4"].font = Font(bold=True)
        dv = DataValidation(type="list",
                            formula1="=PERIODTYPE_BS" if is_bs else "=PERIODTYPE",
                            allow_blank=True)
        ws.add_data_validation(dv)
        for c in range(FIRST_YEAR_COL, FIRST_YEAR_COL + MAX_YEAR_COLS):
            col = get_column_letter(c)
            _input_cell(ws[f"{col}3"], numeric=False)
            _input_cell(ws[f"{col}4"], numeric=False)
            ws[f"{col}4"].number_format = "0"
            dv.add(ws[f"{col}3"])
            if c - FIRST_YEAR_COL < 10:
                ws[f"{col}3"] = "Historical"
        for r, key in enumerate(keys, start=5):
            ws[f"A{r}"] = lab["lines"][key]
            for c in range(FIRST_YEAR_COL, FIRST_YEAR_COL + MAX_YEAR_COLS):
                _input_cell(ws[f"{get_column_letter(c)}{r}"])
        ws.column_dimensions["A"].width = 56
        ws.protection.sheet = True
        ws.protection.password = _LOCK_PWD

    _dimensional_sheets(wb)
    _cost_behaviour_sheets(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ⭐ NAMES, NOT LITERALS. Anything that must NAME one of these sheets — a
# decline sentence, a test, a parser — reads it from here, so the sheet and the
# sentence cannot drift apart.
COST_BEHAVIOUR_SHEET = policy.COST_BEHAVIOUR_SHEET_NAME
CAPACITY_SHEET = policy.CAPACITY_SHEET_NAME


# ═══════════════════════════════════════════════════════════════════════════
# ⭐⭐ THE DIMENSIONAL TAB — LONG FORM, AND THAT IS THE RULING
# ═══════════════════════════════════════════════════════════════════════════

DIMENSION_COLUMNS = [
    ("Period", "The period this row belongs to. Same labels as the statement sheets."),
    ("Frequency", "annual, quarterly or monthly. Must match the statements."),
    ("Dimension Type", "segment, product, customer, channel or geography."),
    ("Code", "A stable code for this line. Reused every period — never renumbered."),
    ("Name", "Display name. Safe to change; the Code is what identifies the line."),
    ("Parent Code", "Optional. Only where this line NESTS INSIDE another of the SAME type."),
    ("Measure", "What this row states. See the Data Dictionary sheet."),
    ("Value", "The amount, in the same units as your statements."),
    ("Currency", "Optional. Defaults to the company currency."),
    ("Unit of Measure", "Optional. Only for unit measures (each, kg, hours)."),
    ("Actual / Plan", "actual or plan. Defaults to actual."),
    ("Notes", "Optional. Never imported as data."),
]


def _dimensional_sheets(wb):
    """One long-form tab, plus the Data Dictionary that makes it self-teaching.

    ⭐⭐ LONG FORM IS WHAT MAKES "PARTIAL DATA IS NEVER AN ERROR" STRUCTURAL.
    With 30% of the data a client supplies 30% of the ROWS — they never see a
    column they cannot fill. A wide sheet with a column per measure presents
    every gap as a blank to be explained, and "partial completion is permitted"
    degrades into a validation exemption rather than the shape of the thing.

    ⭐ `Dimension Type` IS A COLUMN, NOT A SHEET NAME. Segment and product rows
    sit in one table and are never adjacent in a way that invites a sum — the
    anti-double-counting rule made structural, the same way `ax_dimension_map`'s
    existence is what licenses a combination.

    ⭐ AND IT IS THE SHAPE AN ERP EXPORT ALREADY LANDS IN, which is why the
    deferred ERP lane is a column-mapping exercise rather than a second parser.
    """
    from .dimensions import DIMENSION_TYPES, MEASURES

    ws = wb.create_sheet("Segments & Products")
    _style_header(ws["A1"], "Dimensional detail — optional, and partial is fine")
    ws["A2"] = ("Supply only the rows you have. Every row you add unlocks more "
                "analysis; rows you omit are reported as not supplied, never as "
                "zero. Nothing here is required to upload.")
    ws["A2"].font = Font(size=9, italic=True, color="666666")

    for c, (label, _hint) in enumerate(DIMENSION_COLUMNS, start=1):
        _style_header(ws.cell(row=4, column=c), label)
        ws.column_dimensions[get_column_letter(c)].width = max(12, len(label) + 4)
    for r in range(5, 205):
        for c in range(1, len(DIMENSION_COLUMNS) + 1):
            _input_cell(ws.cell(row=r, column=c),
                        numeric=DIMENSION_COLUMNS[c - 1][0] == "Value")
    dv = DataValidation(type="list", formula1="=DIMTYPES", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"C5:C204")
    dvm = DataValidation(type="list", formula1="=MEASURES", allow_blank=True)
    ws.add_data_validation(dvm)
    dvm.add(f"G5:G204")
    ws.protection.sheet = True
    ws.protection.password = _LOCK_PWD

    # ── the Data Dictionary: what each measure unlocks ──────────────────────
    dd = wb.create_sheet("Data Dictionary")
    _style_header(dd["A1"], "Field")
    _style_header(dd["B1"], "What it means")
    for r, (label, hint) in enumerate(DIMENSION_COLUMNS, start=2):
        dd[f"A{r}"] = label
        dd[f"B{r}"] = hint
    r = len(DIMENSION_COLUMNS) + 3
    _style_header(dd[f"A{r}"], "Measure")
    _style_header(dd[f"B{r}"], "What supplying it unlocks")
    # ⭐ DERIVED FROM `MEASURES`, never a second list. A measure added to the
    # vocabulary appears here automatically; a hand list would go stale the
    # first time a tier landed and would then teach the client the wrong thing.
    unlocks = {
        "revenue": "Revenue mix, share, concentration and contribution to growth "
                   "by this dimension — and reconciliation to your income statement.",
        "direct_cost": "Gross profit and gross margin by line; profit pools.",
        "direct_opex": "Direct operating profit by line.",
        "units": "Price-volume-mix analysis and the margin bridge.",
        "list_price": "List-to-net waterfall and discount leakage.",
        "realised_price": "Realised-price trend and pricing variance.",
        "discount": "Discount analysis and leakage.",
    }
    for i, (name, spec) in enumerate(sorted(MEASURES.items()), start=r + 1):
        dd[f"A{i}"] = name
        dd[f"B{i}"] = unlocks.get(name, "")
        dd[f"C{i}"] = f"tier {spec['tier']}"
    dd.column_dimensions["A"].width = 20
    dd.column_dimensions["B"].width = 78
    dd.column_dimensions["C"].width = 10
    dd.protection.sheet = True
    dd.protection.password = _LOCK_PWD

    # dropdown sources, on the existing Lists sheet
    lw = wb["Lists"]
    col_t, col_m = "H", "I"
    for i, t in enumerate(DIMENSION_TYPES, start=1):
        lw[f"{col_t}{i}"] = t
    for i, m in enumerate(sorted(MEASURES), start=1):
        lw[f"{col_m}{i}"] = m
    wb.defined_names["DIMTYPES"] = DefinedName(
        "DIMTYPES", attr_text=f"'Lists'!${col_t}$1:${col_t}${len(DIMENSION_TYPES)}")
    wb.defined_names["MEASURES"] = DefinedName(
        "MEASURES", attr_text=f"'Lists'!${col_m}$1:${col_m}${len(MEASURES)}")


def parse_workbook(content: bytes) -> tuple[dict | None, list]:
    """Parse an uploaded template into the canonical dataset.
    Returns (dataset, errors); errors carry cell-level locations
    (Product §7.14 interactive validation)."""
    errors = []
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        return None, [{"cell": None, "error": f"not a readable .xlsx file: {e}"}]
    sig = wb["Instructions"]["A1"].value if "Instructions" in wb.sheetnames else None
    # ⭐ FAMILY, NOT VERSION — see TEMPLATE_FAMILY above and CORE §7.37. A file
    # stamped v1, v8 or anything else parses identically; what is rejected is a
    # workbook that is not an AXIOM financial template at all.
    if not policy.identifies(sig):
        return None, [{"cell": "Instructions!A1",
                       "error": "not an AXIOM financial template; download one "
                                "from GET /api/v1/financials/templates"}]
    standard = sig.split()[-1]
    if standard not in LABELS:
        return None, [{"cell": "Instructions!A1",
                       "error": f"unknown template standard '{standard}'"}]
    lab = LABELS[standard]

    company = {"standard": standard}
    ws = wb["Company"]
    for r, (field, label, applies) in enumerate(COMPANY_ROWS, start=2):
        if (ws[f"A{r}"].value or "").strip() != label:
            errors.append({"cell": f"Company!A{r}",
                           "error": f"label altered; expected '{label}'"})
        v = ws[f"B{r}"].value
        if isinstance(v, str):
            v = v.strip()
        company[field] = v if v not in ("", None) else None
    if isinstance(company.get("ownership"), str):
        company["ownership"] = company["ownership"].lower()

    # Read year headers/types from the income statement sheet, then require
    # the other sheets to match them exactly.
    def read_columns(ws, allow_opening=False):
        cols = []
        for c in range(FIRST_YEAR_COL, FIRST_YEAR_COL + MAX_YEAR_COLS):
            col = get_column_letter(c)
            y, k = ws[f"{col}4"].value, ws[f"{col}3"].value
            if y in (None, ""):
                continue
            try:
                y = int(y)
            except (TypeError, ValueError):
                errors.append({"cell": f"{ws.title}!{col}4",
                               "error": "year must be an integer"})
                continue
            kind = (str(k or "")).strip().lower()
            allowed = (("opening", "historical", "forecast") if allow_opening
                       else ("historical", "forecast"))
            if kind not in allowed:
                errors.append({"cell": f"{ws.title}!{col}3",
                               "error": ("mark the column Opening, Historical or "
                                         "Forecast" if allow_opening else
                                         "mark the column Historical or Forecast")})
                continue
            cols.append((col, y, kind))
        return cols

    blocks, ref_cols, bs_cols = {}, None, None
    for block, keys in BLOCK_KEYS.items():
        name = lab["sheets"][block]
        if name not in wb.sheetnames:
            errors.append({"cell": None, "error": f"missing sheet '{name}'"})
            continue
        ws = wb[name]
        cols = read_columns(ws, allow_opening=(block == "balance_sheet"))
        # ⭐ THE OPENING COLUMN IS EXCLUDED FROM THE CROSS-SHEET MATCH. Every
        # sheet must agree on the reporting periods; the balance sheet carries
        # one extra column that the flow statements cannot have, and comparing
        # it against them would reject every correctly-filled v8 workbook.
        cmp_cols = [(y, k) for _, y, k in cols if k != "opening"]
        if block == "balance_sheet":
            bs_cols = cols
        if ref_cols is None:
            ref_cols = cols
        elif cmp_cols != [(y, k) for _, y, k in ref_cols if k != "opening"]:
            errors.append({"cell": f"{name}!B3",
                           "error": "year columns must match the "
                                    f"'{lab['sheets']['income_statement']}' sheet"})
        block_data = {}
        for r, key in enumerate(keys, start=5):
            if (ws[f"A{r}"].value or "").strip() != lab["lines"][key]:
                errors.append({"cell": f"{name}!A{r}",
                               "error": f"label altered; expected "
                                        f"'{lab['lines'][key]}'"})
            row = {}
            for col, y, kind in cols:
                v = ws[f"{col}{r}"].value
                if v in (None, ""):
                    # ⭐ THE v8 ROWS MAY BE BLANK, AND THE PARSER HAS ITS OWN
                    # GATE. validate_dataset was taught that these are optional;
                    # this check is a SECOND required-ness rule in a different
                    # file, and leaving it strict rejected the whole upload at
                    # 422 while the validator would have accepted it. Two owners
                    # for one policy — the migration only works if both agree.
                    #
                    # A customer on the previous template uploads with these
                    # rows empty. That is absence, not an error: the
                    # operating-side capital build renders an em dash and says
                    # which rows it lacked.
                    if not policy.required(block, key):
                        continue
                    errors.append({"cell": f"{name}!{col}{r}",
                                   "error": "value required"})
                    continue
                if not isinstance(v, (int, float)):
                    errors.append({"cell": f"{name}!{col}{r}",
                                   "error": "numeric value required"})
                    continue
                row[str(y)] = float(v)
            block_data[key] = row
        blocks[block] = block_data

    if errors:
        return None, errors
    hist = sorted(y for _, y, k in ref_cols if k == "historical")
    fcst = sorted(y for _, y, k in ref_cols if k == "forecast")

    # ⭐ THE OPENING PERIOD IS NOT A REPORTING PERIOD. It goes in its own slot,
    # never into periods.historical — it is the balance carried INTO the first
    # period, not a period the company reported. Folding it into the historical
    # list would put a year in the series that has no income statement and no
    # cash flow, and every consumer that zips the three blocks would silently
    # misalign by one.
    opening_year = None
    for _c, y, k in (bs_cols or []):
        if k == "opening":
            opening_year = y
    periods = {"historical": hist, "forecast": fcst}
    if opening_year is not None:
        periods["opening"] = opening_year

    # ⭐ noncurrent_assets IS DERIVED FROM ITS COMPONENTS, AND ABSENCE
    # PROPAGATES. v8 splits the aggregate; the total stays in the data model
    # because eight consumers read it. If any component is absent the total is
    # absent — a partial sum presented as a total is a fabricated balance sheet,
    # and it would foot against nothing.
    #
    # A pre-v8 workbook supplies no components and its own entered total is left
    # exactly as it is.
    bs = blocks.get("balance_sheet") or {}
    comps = [bs.get(k) or {} for k in engines.BS_NONCURRENT_COMPONENTS]
    if any(comps):
        derived_nca = {}
        for y in {yy for c in comps for yy in c}:
            vals = [c.get(y) for c in comps]
            derived_nca[y] = (None if any(v is None for v in vals)
                              else round(sum(vals), 6))
        bs["noncurrent_assets"] = {**(bs.get("noncurrent_assets") or {}),
                                   **{y: v for y, v in derived_nca.items()
                                      if v is not None}}

    dataset = {"company": company,
               "periods": periods,
               **blocks}
    v = engines.validate_dataset(dataset)
    errors = [{"cell": None, "error": e} for e in v["errors"]]
    return (dataset if not errors else None,
            errors or [{"cell": None, "warning": w} for w in v["warnings"]])


# ═══════════════════════════════════════════════════════════════════════════
# ⭐⭐ T4.1 — COST BEHAVIOUR AND CAPACITY (v13)
# ═══════════════════════════════════════════════════════════════════════════

def _long_form_sheet(wb, name, columns, header_row, intro, numeric_cols,
                     rows=200):
    """One long-form tab, built from the policy's column list.

    ⭐ LONG FORM IS WHAT MAKES "PARTIAL SUPPLY IS NEVER AN ERROR" STRUCTURAL —
    with a third of the data a client supplies a third of the ROWS and never
    sees a column they cannot fill. The dimensional tab established the shape;
    these two follow it rather than inventing a second one.
    """
    ws = wb.create_sheet(name)
    _style_header(ws["A1"], intro[0])
    ws["A2"] = intro[1]
    ws["A2"].font = Font(size=9, italic=True, color="666666")
    for c, (label, _hint) in enumerate(columns, start=1):
        _style_header(ws.cell(row=header_row, column=c), label)
        ws.column_dimensions[get_column_letter(c)].width = max(12, len(label) + 4)
    for r in range(header_row + 1, header_row + 1 + rows):
        for c in range(1, len(columns) + 1):
            _input_cell(ws.cell(row=r, column=c),
                        numeric=columns[c - 1][0] in numeric_cols)
    ws.protection.sheet = True
    ws.protection.password = _LOCK_PWD
    return ws


def _cost_behaviour_sheets(wb):
    """Cost behaviour per pool per period, and capacity as a declared ceiling.

    ⭐⭐ NEITHER SHEET IS REQUIRED AND BOTH SAY SO ON THEIR FIRST LINE. A client
    who fills neither uploads exactly as cleanly as one who never saw them.
    """
    from openpyxl.workbook.defined_name import DefinedName

    # ── the behaviour vocabulary, as a named range so the dropdown is the
    # vocabulary rather than free text collecting "mostly fixed" and "depends"
    lw = wb["Lists"] if "Lists" in wb.sheetnames else wb.create_sheet("Lists")
    start = (lw.max_row or 0) + 2
    for i, cls in enumerate(policy.COST_BEHAVIOUR_CLASSES):
        lw.cell(row=start + i, column=8, value=cls)
    ref = (f"'Lists'!$H${start}:$H${start + len(policy.COST_BEHAVIOUR_CLASSES) - 1}")
    wb.defined_names.add(DefinedName("BEHAVIOUR", attr_text=ref))
    mstart = start + len(policy.COST_BEHAVIOUR_CLASSES) + 1
    for i, m in enumerate(policy.CAPACITY_MEASURES):
        lw.cell(row=mstart + i, column=8, value=m)
    mref = f"'Lists'!$H${mstart}:$H${mstart + len(policy.CAPACITY_MEASURES) - 1}"
    wb.defined_names.add(DefinedName("CAPMEASURES", attr_text=mref))

    ws = _long_form_sheet(
        wb, COST_BEHAVIOUR_SHEET, policy.COST_BEHAVIOUR_COLUMNS,
        policy.COST_BEHAVIOUR_HEADER_ROW,
        ("Cost behaviour — optional, and partial is fine",
         "One row per cost pool per period. You do not need a row for every "
         "pool, and you do not need to split costs by product line — that is "
         "what AXIOM does with the drivers. Pools you omit are reported as not "
         "supplied, never as zero."),
        numeric_cols={"Amount", "Fixed Portion", "Variable Portion",
                      "Step Threshold", "Step Size", "Driver Value"})
    behaviour_col = policy.cost_behaviour_labels().index("Cost Behaviour") + 1
    last = policy.COST_BEHAVIOUR_HEADER_ROW + 200
    dv = DataValidation(type="list", formula1="=BEHAVIOUR", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(behaviour_col)}"
           f"{policy.COST_BEHAVIOUR_HEADER_ROW + 1}:"
           f"{get_column_letter(behaviour_col)}{last}")

    cap = _long_form_sheet(
        wb, CAPACITY_SHEET, policy.CAPACITY_COLUMNS,
        policy.CAPACITY_HEADER_ROW,
        ("Capacity and constraints — optional, and partial is fine",
         "Tell AXIOM what limits you: how much of a resource a period has, how "
         "much of it one unit of a line consumes, and the most of a line you "
         "could sell if capacity allowed. That last figure is your CEILING, "
         "not a forecast — AXIOM never estimates it for you."),
        numeric_cols={"Value"})
    mcol = [l for l, _h in policy.CAPACITY_COLUMNS].index("Measure") + 1
    clast = policy.CAPACITY_HEADER_ROW + 200
    dvc = DataValidation(type="list", formula1="=CAPMEASURES", allow_blank=True)
    cap.add_data_validation(dvc)
    dvc.add(f"{get_column_letter(mcol)}{policy.CAPACITY_HEADER_ROW + 1}:"
            f"{get_column_letter(mcol)}{clast}")

    # ── the Data Dictionary gains both sheets, derived from the same lists ──
    # ⭐ A column added without an explanation is impossible rather than merely
    # discouraged: the sheet and the dictionary are built from one list.
    dd = wb["Data Dictionary"]
    r = (dd.max_row or 1) + 2
    for title, cols in ((COST_BEHAVIOUR_SHEET, policy.COST_BEHAVIOUR_COLUMNS),
                        (CAPACITY_SHEET, policy.CAPACITY_COLUMNS)):
        _style_header(dd[f"A{r}"], title)
        _style_header(dd[f"B{r}"], "What it means")
        r += 1
        for label, hint in cols:
            dd[f"A{r}"] = label
            dd[f"B{r}"] = hint
            r += 1
        r += 1
    _style_header(dd[f"A{r}"], "Capacity measure")
    _style_header(dd[f"B{r}"], "What supplying it unlocks")
    r += 1
    for m in policy.CAPACITY_MEASURES:
        dd[f"A{r}"] = m
        dd[f"B{r}"] = policy.CAPACITY_MEASURE_HELP[m]
        r += 1
