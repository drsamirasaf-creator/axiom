"""Participant List bulk upload (new lane) — Excel template generation + parser.

Pure functions (no DB): the accounts.py endpoints own persistence, reconciliation,
seat-count and the invitation flow. Template is version-stamped (v1) in a named cell
`PLU_VERSION`; the parser reads and validates the stamp. Three tabs — Assessors,
Viewers, Decision Makers — with defined-name dropdowns sourced from a hidden, protected
Lists sheet (no inline validation, no volatile formulas: corruption-prevention
discipline shared with the financial template)."""
import io
import re

VERSION = "v1"
TABS = ["Assessors", "Viewers", "Decision Makers"]
SENIORITY_BANDS = ["Executive", "Senior management", "Mid-level", "Junior", "External partner"]
YESNO = ["Yes", "No"]
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# column layout per tab (header text; order matters — parser reads by position)
COLUMNS = {
    "Assessors":       ["Full Name", "Email", "Department", "Seniority Band", "Title/Role"],
    "Viewers":         ["Full Name", "Email", "Department", "Title/Role"],
    "Decision Makers": ["Full Name", "Email", "Department", "Title/Role", "Is CEO"],
}
ROLE_OF_TAB = {"Assessors": "assessor", "Viewers": "viewer", "Decision Makers": "decision_maker"}
_HEADER_ROW = 3
_DATA_START = 4
_CAPACITY = 300


def build_participant_template(departments: list[str]) -> bytes:
    """Themed, version-stamped participant template. `departments` populates the
    DEPARTMENTS dropdown (align admin names with the org chart — the parser matches
    case-insensitively and never auto-creates departments)."""
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    INK, ACCENT = "17231F", "C8A24B"
    hdr_fill = PatternFill("solid", fgColor=INK)
    hdr_font = Font(bold=True, color=ACCENT)

    wb = Workbook()
    wb.remove(wb.active)

    # hidden Lists sheet — named ranges + version stamp
    lists = wb.create_sheet("Lists")
    lists["E1"] = VERSION
    lists["G1"] = ("AXIOM Participant List — do not edit this sheet. "
                   "The dropdowns and version stamp live here.")
    for i, d in enumerate(departments or [], start=2):
        lists.cell(row=i, column=1, value=d)
    for i, s in enumerate(SENIORITY_BANDS, start=2):
        lists.cell(row=i, column=2, value=s)
    for i, yn in enumerate(YESNO, start=2):
        lists.cell(row=i, column=3, value=yn)
    n_dep = max(2, 1 + len(departments or []))
    wb.defined_names.add(DefinedName("PLU_VERSION", attr_text="Lists!$E$1"))
    wb.defined_names.add(DefinedName("DEPARTMENTS", attr_text=f"Lists!$A$2:$A${n_dep}"))
    wb.defined_names.add(DefinedName("SENIORITY", attr_text=f"Lists!$B$2:$B${1 + len(SENIORITY_BANDS)}"))
    wb.defined_names.add(DefinedName("YESNO", attr_text=f"Lists!$C$2:$C${1 + len(YESNO)}"))
    lists.sheet_state = "hidden"
    lists.protection.sheet = True
    lists.protection.password = "axiom-lists"

    intro = {
        "Assessors": ("People who COMPLETE the assessment instrument. Department and "
                      "Seniority Band are required. Email is the identity key."),
        "Viewers": "View-only access (dashboards, cockpit). Email is the identity key.",
        "Decision Makers": ("Disposition rights on Recommendations & Proposals. Title/Role "
                            "required. Mark exactly one 'Is CEO' = Yes."),
    }
    for tab in TABS:
        ws = wb.create_sheet(tab)
        ws["A1"] = intro[tab]
        ws["A1"].font = Font(italic=True, color="446655")
        ws["A1"].alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(COLUMNS[tab]))
        cols = COLUMNS[tab]
        for c, h in enumerate(cols, start=1):
            cell = ws.cell(row=_HEADER_ROW, column=c, value=h)
            cell.fill = hdr_fill; cell.font = hdr_font
        last = _DATA_START + _CAPACITY - 1
        # dropdowns via defined names (no inline lists)
        def _add_dv(name, col_letter):
            dv = DataValidation(type="list", formula1=f"={name}", allow_blank=True)
            dv.showErrorMessage = (name in ("SENIORITY", "YESNO"))   # enum-strict; DEPARTMENTS free-typed→collision
            ws.add_data_validation(dv); dv.add(f"{col_letter}{_DATA_START}:{col_letter}{last}")
        dep_col = get_column_letter(cols.index("Department") + 1)
        _add_dv("DEPARTMENTS", dep_col)
        if "Seniority Band" in cols:
            _add_dv("SENIORITY", get_column_letter(cols.index("Seniority Band") + 1))
        if "Is CEO" in cols:
            _add_dv("YESNO", get_column_letter(cols.index("Is CEO") + 1))
        for c in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 26 if cols[c - 1] in ("Email", "Full Name") else 20
        ws.freeze_panes = f"A{_DATA_START}"

    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def _cell(v):
    if v is None:
        return ""
    return str(v).strip()


def parse_participant_workbook(content: bytes, valid_departments: list[str]) -> dict:
    """Parse + WITHIN-FILE validation. `valid_departments` = the company's org-chart
    department names (case-insensitive match; unmatched → collision, never auto-create).
    Returns per-tab valid rows, the unioned participant map (email → roles/attrs), and
    errors/warnings/collisions with row numbers. DB reconciliation is the caller's job."""
    from openpyxl import load_workbook
    out = {"version": None, "version_ok": False, "tabs": {}, "participants": {},
           "errors": [], "warnings": [], "collisions": [], "counts": {}}
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        out["errors"].append({"tab": None, "row": None, "message": f"Not a readable .xlsx file ({e})."})
        return out

    # version stamp (named cell)
    ver = None
    try:
        dn = wb.defined_names["PLU_VERSION"]
        for title, coord in dn.destinations:
            ver = _cell(wb[title][coord].value); break
    except Exception:
        ver = None
    out["version"] = ver
    out["version_ok"] = (ver == VERSION)
    if not out["version_ok"]:
        out["errors"].append({"tab": None, "row": None,
                              "message": f"Template version stamp is '{ver or 'missing'}', expected '{VERSION}'. "
                                         f"Download a fresh template."})
        return out

    dep_lookup = {d.strip().lower(): d for d in (valid_departments or [])}
    parts = {}                       # email -> unioned participant
    ceo_yes = []

    def add_role(email, name, role, department, seniority, title, is_ceo):
        p = parts.setdefault(email, {"email": email, "name": name, "roles": [],
                                     "department": None, "seniority": None,
                                     "title": None, "is_ceo": False})
        if role not in p["roles"]:
            p["roles"].append(role)
        if name and not p["name"]:
            p["name"] = name
        if department:
            p["department"] = department
        if seniority:
            p["seniority"] = seniority
        if title:
            p["title"] = title
        if is_ceo:
            p["is_ceo"] = True

    for tab in TABS:
        cols = COLUMNS[tab]
        idx = {h: i for i, h in enumerate(cols)}
        rows_ok, seen_emails = [], {}
        ws = wb[tab] if tab in wb.sheetnames else None
        n = 0
        if ws is not None:
            for r in range(_DATA_START, (ws.max_row or _DATA_START) + 1):
                vals = [_cell(ws.cell(row=r, column=c + 1).value) for c in range(len(cols))]
                if not any(vals):
                    continue
                name = vals[idx["Full Name"]]
                email = vals[idx["Email"]].lower()
                dept_raw = vals[idx["Department"]] if "Department" in idx else ""
                band = vals[idx["Seniority Band"]] if "Seniority Band" in idx else ""
                title = vals[idx["Title/Role"]] if "Title/Role" in idx else ""
                is_ceo = (vals[idx["Is CEO"]].lower() == "yes") if "Is CEO" in idx else False
                rowerrs = []
                if not name:
                    rowerrs.append("Full Name is required")
                if not email:
                    rowerrs.append("Email is required")
                elif not _EMAIL_RE.match(email):
                    rowerrs.append(f"Email '{email}' is not a valid address")
                dept_match = None
                if dept_raw:
                    dept_match = dep_lookup.get(dept_raw.lower())
                    if dept_match is None:
                        out["collisions"].append({"tab": tab, "row": r, "email": email,
                                                  "field": "Department",
                                                  "message": f"Department '{dept_raw}' does not match any org-chart "
                                                             f"department (never auto-created)"})
                        rowerrs.append(f"unknown department '{dept_raw}'")
                if tab == "Assessors":
                    if not dept_raw:
                        rowerrs.append("Department is required for assessors")
                    if not band:
                        rowerrs.append("Seniority Band is required for assessors")
                    elif band not in SENIORITY_BANDS:
                        rowerrs.append(f"Seniority Band '{band}' is not one of the five allowed bands")
                if tab == "Decision Makers" and not title:
                    rowerrs.append("Title/Role is required for decision makers")
                if is_ceo:
                    ceo_yes.append({"email": email, "row": r})
                # duplicate on SAME tab
                if email and email in seen_emails:
                    rowerrs.append(f"duplicate of row {seen_emails[email]} (same email twice on this tab)")
                elif email:
                    seen_emails[email] = r

                if rowerrs:
                    for m in rowerrs:
                        out["errors"].append({"tab": tab, "row": r, "email": email, "message": m})
                else:
                    rows_ok.append({"row": r, "name": name, "email": email,
                                    "department": dept_match, "seniority": band or None,
                                    "title": title or None, "is_ceo": is_ceo})
                    add_role(email, name, ROLE_OF_TAB[tab], dept_match, band or None,
                             title or None, is_ceo)
                n += 1
        out["tabs"][ROLE_OF_TAB[tab]] = {"rows": rows_ok, "count": len(rows_ok), "seen": n}

    if len(ceo_yes) == 0:
        out["warnings"].append({"message": "No 'Is CEO = Yes' marked — expected exactly one."})
    elif len(ceo_yes) > 1:
        out["warnings"].append({"message": f"{len(ceo_yes)} rows marked 'Is CEO = Yes' — expected exactly one "
                                           f"({', '.join(c['email'] for c in ceo_yes)})."})
    out["participants"] = parts
    out["counts"] = {"assessors": out["tabs"]["assessor"]["count"],
                     "viewers": out["tabs"]["viewer"]["count"],
                     "decision_makers": out["tabs"]["decision_maker"]["count"],
                     "valid_participants": len(parts),
                     "errors": len(out["errors"]), "collisions": len(out["collisions"])}
    return out
