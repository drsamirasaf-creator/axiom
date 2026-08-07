"""The founder's survey library, read as instruments.

⭐⭐ THE WORKBOOK IS THE AUTHORING SURFACE AND STAYS THAT WAY. This module reads
`docs/specs/AXIOM_Survey_Library_*.xlsx` and composes
`ax_assessment_instruments` + `ax_assessment_instrument_items` from it. **No
question text is stored here** — an instrument is a NAME plus a membership over
`ax_assessment_items`, so a question is authored in one place and worded in one
place.

## ⛔ WHAT IT REFUSES TO DO

- **It never deletes.** A row the template stops mentioning is `flagged_absent`,
  and a removal is a revoke with an actor (§4v.1 ruling 1). *Without `source`, a
  re-upload whose template omits a link DELETES a link a human created in the
  app; without `flagged_absent`, an omission is indistinguishable from a
  deletion* — `GoalInitiativeLink`'s own docstring, and the reason both columns
  exist here.
- ⛔ **It never touches `ax_assessment_responses`.** A participant's instrument
  DERIVES from their audience, so 15,371 fielded responses stay valid and no
  second owner of "which questionnaire" appears.
- ⛔ **It does not re-decide Compliance.** §16.4 ruled it into L1 axis 11 and
  this module asserts that mapping rather than restating it.

## ⭐ THE SCALE

The store is **1-10** and the workbook was corrected to match (founder ruling,
8 Aug). ⛔ **No conversion happens anywhere** — 15,371 responses were fielded on
a real scale against a dropdown default that never was, and *a 7 of 10 is not a
7 of 7*: any conversion invents a respondent's answer.
"""
import os
import re
from datetime import datetime

SPECS = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(
    os.path.abspath(__file__)))), "docs", "specs")

DEPARTMENTS_WB = "AXIOM_Survey_Library_Departments_Completed.xlsx"
STAKEHOLDERS_WB = "AXIOM_Survey_Library_Stakeholders_Completed.xlsx"

# ⭐⭐ MERIDIAN'S DEMO SCOPE — EIGHT DEPARTMENTS, FOUNDER RULING 8 Aug.
# ⛔ Sales and Marketing are SEPARATE, matching the template's own sheets: the
# template collects them separately, so a customer filling it in produces eight
# instruments. A merged demo would show a structure the product does not collect.
DEMO_DEPARTMENTS = (
    "Executive Management",
    "Finance and Accounting",
    "Sales",
    "Marketing",
    "Operations",
    "Information Technology",
    "Human Resources",
    "Internal Audit",
)

# ⛔ §16.4 — ASSERTED, NOT RE-DECIDED. Compliance maps into L1 #11 and the axis
# set stays at 13. Two questions do not justify renormalising every weight.
COMPLIANCE_AXIS = 11
CATEGORY_AXIS_OVERRIDES = {"Compliance": COMPLIANCE_AXIS}

_SHARED = re.compile(r"^\d+$")
_UNIQUE = re.compile(r"^[Uu]\d+$")


def _key(*parts):
    slug = "-".join(re.sub(r"[^a-z0-9]+", "-", str(p).strip().lower()).strip("-")
                    for p in parts if p)
    return slug[:64]


def read_workbook(path):
    """-> {sheet_name: [ {ref, block, category, title, guidance, readiness, scale} ]}

    ⛔ THE HEADER IS ON ROW 4, under a title and a section banner. A first census
    scanned for the first header containing 'id' and returned 15 questions for a
    703-question workbook — absurd, and therefore caught. A plausible wrong
    number would not have been (§III.18).
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for sheet in wb.sheetnames:
        if sheet == "README":
            continue
        rows = []
        for r in wb[sheet].iter_rows(values_only=True):
            if not r or r[0] is None:
                continue
            ref = str(r[0]).strip()
            if _SHARED.match(ref):
                block = "shared"
            elif _UNIQUE.match(ref):
                block = "unique"
            else:
                continue                      # a banner or the header row
            g = lambda i: (str(r[i]).strip() if len(r) > i and r[i] is not None else None)
            rows.append({"ref": ref, "block": block, "category": g(1),
                         "title": g(2), "guidance": g(3),
                         "readiness": (g(4) or "").lower() == "yes",
                         "scale": g(5)})
        out[sheet] = rows
    return out


def library(kind="departments"):
    name = DEPARTMENTS_WB if kind == "departments" else STAKEHOLDERS_WB
    return read_workbook(os.path.join(SPECS, name))


def scale_declared(kind="departments"):
    """Every distinct declared scale. ⭐ Asserted rather than assumed — the
    workbook and the store disagreed once and it cost a ruling."""
    seen = set()
    for rows in library(kind).values():
        seen |= {r["scale"] for r in rows if r["scale"]}
    return seen


def compose(db, company_id, framework_id, sheets, *, kind="departments",
            audience_kind="department", audience_ref_for=None, now=None,
            actor=None):
    """Create or reconcile instruments for `sheets`. -> a summary dict.

    ⛔ RECONCILE, NEVER REPLACE. A second run over a changed library flags what
    the template dropped and revives what it restored; it deletes nothing.
    """
    from .accounts import (AssessmentInstrument, AssessmentInstrumentItem,
                           AssessmentItem)
    now = now or datetime.utcnow()
    lib = library(kind)
    items = {i.title.strip().lower(): i for i in
             db.query(AssessmentItem).filter_by(framework_id=framework_id).all()}
    summary = {"instruments": [], "unmatched": [], "flagged": 0, "revived": 0}

    for sheet in sheets:
        rows = lib.get(sheet)
        if rows is None:
            summary["unmatched"].append({"sheet": sheet, "why": "no such sheet"})
            continue
        key = _key(audience_kind, sheet)
        inst = (db.query(AssessmentInstrument)
                  .filter_by(company_id=company_id, framework_id=framework_id,
                             key=key).first())
        if inst is None:
            inst = AssessmentInstrument(
                company_id=company_id, framework_id=framework_id, key=key,
                name=sheet, audience_kind=audience_kind,
                audience_ref=(audience_ref_for or {}).get(sheet),
                orientation=("internal" if any(r["block"] == "shared" for r in rows)
                             else "external"),
                created_at=now, created_by=actor)
            db.add(inst)
            db.flush()

        seen_ids = set()
        for pos, row in enumerate(rows):
            it = items.get((row["title"] or "").strip().lower())
            if it is None:
                summary["unmatched"].append(
                    {"sheet": sheet, "ref": row["ref"],
                     "title": (row["title"] or "")[:60],
                     "why": "no item in this framework carries this title"})
                continue
            seen_ids.add(it.id)
            link = (db.query(AssessmentInstrumentItem)
                      .filter_by(instrument_id=inst.id, item_id=it.id).first())
            if link is None:
                db.add(AssessmentInstrumentItem(
                    instrument_id=inst.id, item_id=it.id, position=pos,
                    block=row["block"], source="template", created_at=now,
                    created_by=actor))
            elif link.flagged_absent:
                # ⭐ THE TEMPLATE BROUGHT IT BACK. An absence that returns is a
                # revival, not a new row — the history stays continuous.
                link.flagged_absent = False
                summary["revived"] += 1

        # ⛔ WHAT THE TEMPLATE NO LONGER MENTIONS IS FLAGGED, NEVER DELETED —
        # and an IN-APP link is never flagged by a template's silence, which is
        # the whole reason `source` exists.
        for link in db.query(AssessmentInstrumentItem).filter_by(
                instrument_id=inst.id).all():
            if (link.item_id not in seen_ids and link.source == "template"
                    and not link.flagged_absent and link.revoked_at is None):
                link.flagged_absent = True
                summary["flagged"] += 1

        summary["instruments"].append(
            {"key": key, "name": sheet, "id": inst.id,
             "shared": sum(1 for r in rows if r["block"] == "shared"),
             "unique": sum(1 for r in rows if r["block"] == "unique"),
             "authored": len(rows)})
    return summary


def revoke(db, instrument, actor=None, now=None):
    """⛔ REMOVAL IS A REVOKE, NEVER A DELETE (§4v.1 ruling 1)."""
    instrument.revoked_at = now or datetime.utcnow()
    instrument.revoked_by = actor
    return instrument
