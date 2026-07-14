"""API contract tests over the full vertical slice. REQ-TEST-002."""
import os, tempfile
os.environ["DATABASE_URL"] = "sqlite:///" + tempfile.mktemp(suffix=".db")
import pytest
from fastapi.testclient import TestClient
from services.api.main import app

@pytest.fixture(scope="module")
def client():
    with TestClient(app) as c:
        yield c

def test_health(client):
    r = client.get("/health")
    assert r.status_code == 200 and r.json()["status"] == "ok"

def test_enterprise_state_lifecycle(client):
    r = client.post("/api/v1/enterprises", json={"name": "Meridian Corp", "sector": "Industrials"})
    assert r.status_code == 201
    eid = r.json()["id"]
    r = client.post(f"/api/v1/enterprises/{eid}/state",
                    json={"payload": {"capital": 4.0, "horizon": 6}, "note": "opening state"})
    assert r.status_code == 201
    r = client.get(f"/api/v1/enterprises/{eid}/state")
    assert r.json()["payload"]["capital"] == 4.0

def test_tenant_isolation(client):
    r = client.post("/api/v1/enterprises", json={"name": "Hidden"},
                    headers={"X-AXIOM-Tenant": "other"})
    other_id = r.json()["id"]
    assert client.get(f"/api/v1/enterprises/{other_id}").status_code == 404

def test_reo_solve_and_provenance(client):
    r = client.get("/api/v1/reo/problems")
    names = {p["problem"] for p in r.json()}
    assert names == {"allocation_sqrt", "quadratic_form", "duality_demo",
                     "switch_family", "dp_switch", "value_iteration",
                     "pareto_frontier", "kkt_circle"}
    r = client.post("/api/v1/reo/solve", json={"problem": "switch_family", "params": {}})
    assert r.status_code == 201
    body = r.json()["result"]
    assert body["all_checkpoints_pass"] is True and abs(body["value"] - 39.6863) < 5e-4
    runs = client.get("/api/v1/reo/runs").json()
    assert runs and runs[0]["problem"] == "switch_family"

def test_reo_unknown_problem_404(client):
    assert client.post("/api/v1/reo/solve", json={"problem": "nope"}).status_code == 404

def test_reo_bad_params_422(client):
    r = client.post("/api/v1/reo/solve",
                    json={"problem": "allocation_sqrt", "params": {"a": -1}})
    assert r.status_code == 422

def test_education_registry(client):
    mods = client.get("/api/v1/education/modules").json()
    assert len(mods) == 32
    assert sum(1 for m in mods if m["status"] == "live") == 14
    assert mods[16]["seed"] == 26201 and mods[16]["volume"] == "II"

def test_simulation_run_and_provenance(client):
    r = client.get("/api/v1/simulation/scenarios")
    assert {s["scenario"] for s in r.json()} == {"trajectory", "twin_sync",
                                                 "stability_dial", "twin_decision"}
    r = client.post("/api/v1/simulation/run",
                    json={"scenario": "twin_decision", "params": {}})
    assert r.status_code == 201
    body = r.json()["result"]
    assert body["all_checkpoints_pass"] is True
    assert abs(body["solution"]["regret_open_twin"] - 1.3605) < 5e-4
    assert body["solution"]["chart_data"][1]["sync_pick"] is True
    runs = client.get("/api/v1/simulation/runs").json()
    assert runs and runs[0]["scenario"] == "twin_decision"

def test_simulation_unknown_scenario_404(client):
    assert client.post("/api/v1/simulation/run", json={"scenario": "nope"}).status_code == 404

def test_simulation_bad_params_422(client):
    r = client.post("/api/v1/simulation/run",
                    json={"scenario": "twin_sync", "params": {"gains": [2.0]}})
    assert r.status_code == 422


def test_phase2_problems_solve_via_api(client):
    for problem, key, expected in (("dp_switch", "V0", 39.6863),
                                   ("value_iteration", "V_G", 70.0),
                                   ("kkt_circle", "lambda_star", 0.5)):
        r = client.post("/api/v1/reo/solve", json={"problem": problem, "params": {}})
        assert r.status_code == 201
        body = r.json()["result"]
        assert body["all_checkpoints_pass"] is True
        assert abs(body["solution"][key] - expected) < 5e-4


def test_risk_analyses_and_provenance(client):
    r = client.get("/api/v1/risk/analyses")
    assert {a["analysis"] for a in r.json()} == {"chance_constraint", "dro_flip",
                                                 "robust_radius", "gbm_valuation"}
    r = client.post("/api/v1/risk/run", json={"analysis": "dro_flip", "params": {}})
    assert r.status_code == 201
    body = r.json()["result"]
    assert body["all_checkpoints_pass"] is True
    assert abs(body["solution"]["flip_radius"] - 0.125) < 5e-4
    runs = client.get("/api/v1/risk/runs").json()
    assert runs and runs[0]["analysis"] == "dro_flip"

def test_risk_unknown_analysis_404(client):
    assert client.post("/api/v1/risk/run", json={"analysis": "nope"}).status_code == 404

def test_risk_bad_params_422(client):
    r = client.post("/api/v1/risk/run",
                    json={"analysis": "chance_constraint", "params": {"mu": -1}})
    assert r.status_code == 422


def test_learning_experiments_and_provenance(client):
    r = client.get("/api/v1/learning/experiments")
    assert {e["experiment"] for e in r.json()} == {
        "generalization_duel", "kmeans_clustering", "prediction_regret",
        "q_learning", "knowledge_augmented", "anfis_sugeno"}
    r = client.post("/api/v1/learning/run", json={"experiment": "q_learning", "params": {}})
    assert r.status_code == 201
    body = r.json()["result"]
    assert body["all_checkpoints_pass"] is True
    assert body["solution"]["sweep_policy_correct"] == 5
    assert body["solution"]["sweeps_to_tol"] == 173
    runs = client.get("/api/v1/learning/runs").json()
    assert runs and runs[0]["experiment"] == "q_learning"

def test_learning_unknown_404_and_bad_params_422(client):
    assert client.post("/api/v1/learning/run", json={"experiment": "nope"}).status_code == 404
    r = client.post("/api/v1/learning/run",
                    json={"experiment": "anfis_sugeno", "params": {"mode": "psychic"}})
    assert r.status_code == 422


def test_education_module_detail_deep_link(client):
    r = client.get("/api/v1/education/modules/axiom-07")
    assert r.status_code == 200
    body = r.json()
    assert body["any_live"] is True and len(body["volumes"]) == 2
    vol2 = next(m for m in body["volumes"] if m["volume"] == "II")
    keys = {e["key"] for e in vol2["experiences"]}
    assert keys == {"dp_switch", "value_iteration"}
    assert vol2["course_links"]["chapter"].endswith("/chapters/v2ch07.html")

def test_education_unknown_slug_404(client):
    assert client.get("/api/v1/education/modules/axiom-99").status_code == 404

def test_education_summary(client):
    s = client.get("/api/v1/education/summary").json()
    assert s["modules_total"] == 32 and s["modules_live"] == 14
    assert s["experiences_total"] == 22

def test_every_experience_key_resolves_to_a_real_engine(client):
    reg = {
        "reo": {p["problem"] for p in client.get("/api/v1/reo/problems").json()},
        "simulation": {s["scenario"] for s in client.get("/api/v1/simulation/scenarios").json()},
        "risk": {a["analysis"] for a in client.get("/api/v1/risk/analyses").json()},
        "learning": {e["experiment"] for e in client.get("/api/v1/learning/experiments").json()},
    }
    mods = client.get("/api/v1/education/modules").json()
    for m in mods:
        for e in m["experiences"]:
            assert e["key"] in reg[e["kind"]], f'{m["slug"]}: {e["kind"]}/{e["key"]} missing'


# --------------------------- Phase 6: Financial Core ------------------------

def _meridian():
    from tests.fixtures.refcases import meridian
    return meridian()


def test_financial_templates_download(client):
    r = client.get("/api/v1/financials/templates")
    assert {t["standard"] for t in r.json()} == {"us_gaap", "ifrs"}
    for std in ("us_gaap", "ifrs"):
        r = client.get(f"/api/v1/financials/templates/{std}")
        assert r.status_code == 200
        assert r.content[:2] == b"PK"           # xlsx zip magic
        assert "spreadsheetml" in r.headers["content-type"]
    assert client.get("/api/v1/financials/templates/frs102").status_code == 404


def test_dataset_direct_input_and_derived(client):
    r = client.post("/api/v1/financials/datasets",
                    json={"name": "Meridian FY25", "data": _meridian()})
    assert r.status_code == 201
    ds = r.json()
    assert ds["standard"] == "us_gaap" and ds["ownership"] == "public"
    r = client.get(f"/api/v1/financials/datasets/{ds['id']}/derived")
    d = r.json()
    i = d["years"].index(2025)
    assert abs(d["fcff"][i] - 124.95) < 5e-4
    assert abs(d["fcfe"][i] - 126.95) < 5e-4


def test_dataset_direct_input_validation_422(client):
    bad = _meridian()
    del bad["income_statement"]["revenue"]["2025"]
    bad["company"].pop("beta")
    r = client.post("/api/v1/financials/datasets",
                    json={"name": "broken", "data": bad})
    assert r.status_code == 422
    detail = " ".join(r.json()["detail"])
    assert "revenue[2025]" in detail and "company.beta" in detail


def test_template_fill_and_upload_roundtrip(client):
    """The full client journey: download template, fill it, upload it."""
    import io
    from openpyxl import load_workbook
    from services.api.modules.financials import templates as tpl
    m = _meridian()
    content = client.get("/api/v1/financials/templates/us_gaap").content
    wb = load_workbook(io.BytesIO(content))
    ws = wb["Company"]
    values = {"name": m["company"]["name"], "ownership": "public",
              "currency": "USD"}
    for r_i, (field, label, applies) in enumerate(tpl.COMPANY_ROWS, start=2):
        ws[f"B{r_i}"] = values.get(field, m["company"].get(field))
    years = m["periods"]["historical"] + m["periods"]["forecast"]
    kinds = ["Historical"] * 5 + ["Forecast"] * 5
    for block, keys in tpl.BLOCK_KEYS.items():
        ws = wb[tpl.LABELS["us_gaap"]["sheets"][block]]
        for c, (y, k) in enumerate(zip(years, kinds), start=tpl.FIRST_YEAR_COL):
            ws.cell(row=3, column=c, value=k)
            ws.cell(row=4, column=c, value=y)
        for r_i, key in enumerate(keys, start=5):
            for c, y in enumerate(years, start=tpl.FIRST_YEAR_COL):
                ws.cell(row=r_i, column=c, value=m[block][key][str(y)])
    buf = io.BytesIO(); wb.save(buf)
    r = client.post("/api/v1/financials/datasets/upload",
                    files={"file": ("meridian.xlsx", buf.getvalue(),
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 201, r.text
    ds = r.json()
    assert ds["source"] == "upload"
    d = client.get(f"/api/v1/financials/datasets/{ds['id']}/derived").json()
    assert abs(d["fcff"][d["years"].index(2025)] - 124.95) < 5e-4


def test_upload_rejects_foreign_workbook(client):
    import io
    from openpyxl import Workbook
    wb = Workbook(); buf = io.BytesIO(); wb.save(buf)
    r = client.post("/api/v1/financials/datasets/upload",
                    files={"file": ("random.xlsx", buf.getvalue(), "application/octet-stream")})
    assert r.status_code == 422
    assert "AXIOM financial template" in str(r.json()["detail"])


def test_forecast_endpoint_and_persist(client):
    from tests.fixtures.refcases import halcyon
    r = client.post("/api/v1/financials/datasets",
                    json={"name": "Halcyon FY25", "data": halcyon()})
    hid = r.json()["id"]
    r = client.post(f"/api/v1/financials/datasets/{hid}/forecast",
                    json={"assumptions": {"horizon": 5}, "persist": True})
    assert r.status_code == 200
    body = r.json()
    assert abs(body["provenance"]["revenue_growth"] - 0.05948) < 5e-5
    assert body["derived"]["n_forecast"] == 5
    assert "dataset_id" in body


def test_valuation_three_modes_and_runs(client):
    r = client.get("/api/v1/valuation/modes")
    assert {m["mode"] for m in r.json()} == {"proforma", "auto_forecast"}
    r = client.post("/api/v1/financials/datasets",
                    json={"name": "Meridian for valuation", "data": _meridian()})
    mid = r.json()["id"]
    r = client.post("/api/v1/valuation/run",
                    json={"dataset_id": mid, "mode": "proforma"})
    assert r.status_code == 201
    res = r.json()["result"]
    assert res["all_checkpoints_pass"] is True
    assert abs(res["deterministic"]["enterprise_value"] - 2481.3499) < 5e-2
    assert abs(res["risk_adjusted"]["raev"] - 2313.27) < 0.05
    # wrong mode for this dataset -> 422; unknown dataset -> 404
    assert client.post("/api/v1/valuation/run",
                       json={"dataset_id": mid, "mode": "auto_forecast"}).status_code == 422
    assert client.post("/api/v1/valuation/run",
                       json={"dataset_id": 99999, "mode": "proforma"}).status_code == 404
    runs = client.get("/api/v1/valuation/runs").json()
    assert runs and runs[0]["dataset_id"] == mid


def test_dashboard_metrics_endpoint(client):
    r = client.post("/api/v1/financials/datasets",
                    json={"name": "Meridian dash", "data": _meridian()})
    mid = r.json()["id"]
    client.post("/api/v1/valuation/run",
                json={"dataset_id": mid, "mode": "proforma"})
    r = client.get(f"/api/v1/metrics/dashboard/{mid}")
    assert r.status_code == 200
    dash = r.json()
    strip = {k["kpi"]: k["current"] for k in dash["kpi_strip"]}
    assert abs(strip["ROE"] - 0.240046) < 5e-4
    assert abs(strip["EVA (Economic Profit)"] - 86.7075) < 5e-3
    assert "Risk-Adjusted Enterprise Value" in strip     # valuation attached
    assert abs(dash["health"]["health_index"] - 96.36) < 0.05


def test_document_plumbing_honest_status(client):
    r = client.post("/api/v1/financials/documents",
                    files={"file": ("strategy.txt", b"Five-year strategic plan.",
                                    "text/plain")},
                    data={"note": "board strategy memo"})
    assert r.status_code == 201
    doc = r.json()
    assert doc["size_bytes"] == 25 and doc["ai_analysis"] is None  # Phase 7
    docs = client.get("/api/v1/financials/documents").json()
    assert docs[0]["filename"] == "strategy.txt"


# ----------------- Phase 6.1: tab split + tooltip glossary ------------------

def test_risk_analyses_carry_tab_category(client):
    r = client.get("/api/v1/risk/analyses")
    cats = {a["analysis"]: a["category"] for a in r.json()}
    assert cats["gbm_valuation"] == "valuation"          # Valuation tab
    assert {cats["chance_constraint"], cats["dro_flip"],
            cats["robust_radius"]} == {"risk"}           # Risk Analysis tab


def test_glossary_covers_tabs_and_headline_terms(client):
    g = client.get("/api/v1/metrics/glossary").json()
    for term in ("Dashboard", "Data Input", "Valuation", "Risk Analysis",
                 "FCFF", "FCFE", "WACC", "EVA (Economic Profit)",
                 "Enterprise Health Index", "DLOM", "CVaR95",
                 "Risk-Adjusted Enterprise Value", "GBM Valuation Fan",
                 "DRO Flip Map (TV Ambiguity Ball)", "Volatility Drag"):
        assert term in g and len(g[term]) > 20, term


def test_kpi_strip_carries_definitions(client):
    r = client.post("/api/v1/financials/datasets",
                    json={"name": "Meridian tooltips", "data": _meridian()})
    mid = r.json()["id"]
    dash = client.get(f"/api/v1/metrics/dashboard/{mid}").json()
    for card in dash["kpi_strip"]:
        assert card["definition"], f"missing definition for {card['kpi']}"


# --------------------------- Phase 7: Intelligence Layer --------------------

def _upload_doc(client, text, content_type="text/plain", dataset_id=None):
    data = {"note": "test doc"}
    if dataset_id:
        data["dataset_id"] = str(dataset_id)
    r = client.post("/api/v1/financials/documents",
                    files={"file": ("plan.txt", text.encode(), content_type)},
                    data=data)
    assert r.status_code == 201
    return r.json()["id"]


def test_analyze_503_when_ai_unconfigured(client, monkeypatch):
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    did = _upload_doc(client, "Revenue growth of 5% (0.05) is planned.")
    r = client.post(f"/api/v1/intelligence/documents/{did}/analyze")
    assert r.status_code == 503
    assert "ANTHROPIC_API_KEY" in r.json()["detail"]


def test_analyze_415_for_binary_document(client):
    r = client.post("/api/v1/financials/documents",
                    files={"file": ("plan.pdf", b"%PDF-1.4 ...",
                                    "application/pdf")})
    did = r.json()["id"]
    r = client.post(f"/api/v1/intelligence/documents/{did}/analyze")
    assert r.status_code == 415


def test_analyze_decide_flow_with_mocked_ai(client, monkeypatch):
    """Full journey with the AI seam mocked: analyze -> gates -> decisions
    -> valuation-ready assumptions -> valuation run."""
    import json as _json
    from services.api.modules.intelligence import ai_client
    doc_text = ("Strategic plan: we target revenue growth of 8% (0.08) and "
                "an EBIT margin of 15% (0.15). Long-run inflation-like "
                "terminal growth of 2% (0.02) is assumed.")
    model_reply = _json.dumps([
        {"field": "revenue_growth", "value": 0.08, "rationale": "stated target",
         "source_quote": "revenue growth of 8% (0.08)"},
        {"field": "ebit_margin", "value": 0.15, "rationale": "stated margin",
         "source_quote": "an EBIT margin of 15% (0.15)"},
        {"field": "terminal_growth", "value": 0.02, "rationale": "stated",
         "source_quote": "terminal growth of 2% (0.02)"},
        {"field": "revenue_growth", "value": 0.30, "rationale": "hallucinated",
         "source_quote": "we will triple revenue overnight"}])
    monkeypatch.setattr(ai_client, "complete",
                        lambda system, user_text, max_tokens=2000: model_reply)
    from tests.fixtures.refcases import halcyon
    ds = client.post("/api/v1/financials/datasets",
                     json={"name": "Halcyon AI", "data": halcyon()}).json()
    did = _upload_doc(client, doc_text, dataset_id=ds["id"])
    r = client.post(f"/api/v1/intelligence/documents/{did}/analyze")
    assert r.status_code == 200
    a = r.json()
    assert a["status"] == "proposed" and len(a["suggestions"]) == 3
    assert len(a["rejected"]) == 1                     # hallucinated quote gated
    assert all(s["verified_quote"] for s in a["suggestions"])
    # persisted on the document
    docs = client.get("/api/v1/financials/documents").json()
    assert any(d["id"] == did and d["ai_analysis"]["status"] == "proposed"
               for d in docs)
    # decide: accept growth + terminal, reject margin
    r = client.post(f"/api/v1/intelligence/documents/{did}/decisions",
                    json={"decisions": {0: "accept", 1: "reject", 2: "accept"}})
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "decided"
    va = body["valuation_assumptions"]
    assert va["terminal_growth"] == 0.02
    assert va["forecast"] == {"revenue_growth": 0.08}
    # assumptions run through the certified engine
    r = client.post("/api/v1/valuation/run",
                    json={"dataset_id": ds["id"], "mode": "auto_forecast",
                          "assumptions": va})
    assert r.status_code == 201
    assert r.json()["result"]["provenance"]["revenue_growth"] == 0.08


def test_decisions_409_without_analysis(client):
    did = _upload_doc(client, "no analysis yet")
    r = client.post(f"/api/v1/intelligence/documents/{did}/decisions",
                    json={"decisions": {0: "accept"}})
    assert r.status_code == 409


def test_reo_health_endpoint(client):
    r = client.post("/api/v1/financials/datasets",
                    json={"name": "Meridian health", "data": _meridian()})
    mid = r.json()["id"]
    h = client.get(f"/api/v1/intelligence/health/{mid}").json()
    assert abs(h["health_index"] - 95.5) < 0.05
    assert h["version"] == "reo_distance_v1"
    assert len(h["wacc_curve"]) >= 10


def test_recommendations_endpoint(client):
    from tests.fixtures.refcases import halcyon
    r = client.post("/api/v1/financials/datasets",
                    json={"name": "Halcyon recs", "data": halcyon()})
    hid = r.json()["id"]
    recs = client.get(f"/api/v1/intelligence/recommendations/{hid}").json()
    assert recs["recommendations"][0]["move"] == "optimal_capital_structure"
    assert recs["all_checkpoints_pass"] is True
    assert all("expected_ev_impact" in m for m in recs["recommendations"])


def test_stress_endpoint_persists_run(client):
    r = client.post("/api/v1/financials/datasets",
                    json={"name": "Meridian stress", "data": _meridian()})
    mid = r.json()["id"]
    r = client.post("/api/v1/valuation/stress",
                    json={"dataset_id": mid, "mode": "proforma"})
    assert r.status_code == 201
    res = r.json()["result"]
    assert res["resilient_beyond"] == 0.5
    assert res["all_checkpoints_pass"] is True
    runs = client.get("/api/v1/valuation/runs").json()
    assert any(x["mode"] == "dro_stress" for x in runs)


def test_glossary_covers_phase7_terms(client):
    g = client.get("/api/v1/metrics/glossary").json()
    for term in ("AI Document Analysis", "Source Quote", "Approval Gate",
                 "Enterprise Health Index (REO)", "Optimal Capital Structure",
                 "Transformation Recommendations", "DRO Stress Test",
                 "Breakeven Ambiguity Radius"):
        assert term in g and len(g[term]) > 20, term


# ------------------------- Phase 7.5: Benchmarking --------------------------

def test_benchmark_sectors_endpoint(client):
    r = client.get("/api/v1/benchmarks/sectors")
    body = r.json()
    assert "Industrials" in body["sectors"] and len(body["sectors"]) >= 10
    assert body["source"]["kind"] == "curated"
    dirs = {k["kpi"]: k["direction"] for k in body["kpis"]}
    assert dirs["debt_to_equity"] == "lower" and dirs["capex_pct_revenue"] == "context"


def test_benchmark_compare_endpoint(client):
    r = client.post("/api/v1/financials/datasets",
                    json={"name": "Meridian bench", "data": _meridian()})
    mid = r.json()["id"]
    r = client.post("/api/v1/benchmarks/compare",
                    json={"dataset_id": mid, "sector": "Industrials"})
    assert r.status_code == 200
    body = r.json()
    assert abs(body["benchmark_performance_index"] - 142.62) < 0.05
    assert body["narrative"] and body["all_checkpoints_pass"] is True
    # unknown sector -> 404 with pointer; no sector anywhere -> 422
    assert client.post("/api/v1/benchmarks/compare",
                       json={"dataset_id": mid, "sector": "Nope"}).status_code == 404
    assert client.post("/api/v1/benchmarks/compare",
                       json={"dataset_id": mid}).status_code == 422


def test_benchmark_glossary_terms(client):
    g = client.get("/api/v1/metrics/glossary").json()
    for term in ("Benchmarking", "Benchmark Performance Index",
                 "Implied Value", "Custom Peer Set", "Traffic Light",
                 "Context KPI"):
        assert term in g and len(g[term]) > 20, term


def test_sector_warning_on_direct_input(client):
    ds = _meridian()
    ds["company"]["sector"] = "Underwater Basket Weaving"
    r = client.post("/api/v1/financials/datasets",
                    json={"name": "odd sector", "data": ds})
    assert r.status_code == 201
    assert any("no curated benchmark" in w
               for w in r.json()["validation"]["warnings"])
