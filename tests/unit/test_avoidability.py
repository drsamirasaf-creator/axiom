"""T5.1 — avoidability: what actually leaves if you discontinue a line.

⭐⭐ THE §22 CORRECTIVE'S SECOND HALF. T4.2 answers whether a line covers its own
variable cost. Avoidability answers what leaves if you stop selling it — and
until it is declared, the corrective's conclusion rests on a premise nobody
supplied. The same line, the same figures, can produce opposite advice depending
on a number only the client holds.

⛔ RULING 2: blank is not a declaration; an explicit zero is.
⛔ RULING 3: no horizon — a standing annual amount, because discounting is
   `prescience_decision`'s.
"""
import os
import tempfile

os.environ.setdefault("DATABASE_URL",
                      "sqlite:///" + tempfile.mktemp(prefix="avoid-", suffix=".db"))

import pytest

from services.api.modules.financials import managerial as M


DECLARED = [
    {"period": 2025, "line_code": "PL-X", "pool": "Customer Support",
     "avoidable_amount": 40.0, "notice_period_months": 3.0},
    {"period": 2025, "line_code": "PL-X", "pool": "Logistics",
     "avoidable_amount": 0.0, "notice_period_months": 0.0},
]


# ── 1 · stranded is the complement of a declaration ────────────────────────

def test_stranded_is_allocated_less_avoidable():
    """⭐ THE COMPLEMENT OF A DECLARATION IS NOT AN INFERENCE. The client said
    40 disappears; the remaining 60 is what stays, and saying so invents
    nothing."""
    a = M.avoidability(DECLARED, allocated_charge=100.0, period=2025,
                       line_code="PL-X")
    assert a["available"]
    assert a["value"]["avoidable"] == pytest.approx(40.0)
    assert a["value"]["stranded"] == pytest.approx(60.0)


def test_an_explicit_zero_is_a_declaration_and_computes():
    """⭐⭐ RULING 2. A client entering 0 has TOLD you nothing disappears."""
    zero = [{"period": 2025, "line_code": "PL-X", "pool": "Corporate",
             "avoidable_amount": 0.0}]
    a = M.avoidability(zero, allocated_charge=100.0, period=2025,
                       line_code="PL-X")
    assert a["available"]
    assert a["value"]["avoidable"] == 0.0
    assert a["value"]["stranded"] == pytest.approx(100.0)


def test_a_blank_declines_and_names_the_column():
    """⭐⭐ RULING 2, THE OTHER HALF. Blank-as-zero is `or 0` on the input that
    decides whether a line should be exited, and it would make every line's
    stranded cost 100% of its allocated share — reproducing the corrective's
    unstated premise, this time computed and therefore invisible."""
    blank = [{"period": 2025, "line_code": "PL-X", "pool": "Corporate",
              "avoidable_amount": None}]
    a = M.avoidability(blank, allocated_charge=100.0, period=2025,
                       line_code="PL-X")
    assert a["available"] is False
    assert a["value"] is None
    assert "Avoidable Amount" in a["unlocks"]
    assert "Cost Avoidability" in a["unlocks"]
    for token in ("avoidable_amount", "allocated_charge"):
        assert token not in a["unlocks"], f"engine token {token!r} in a sentence"


def test_no_declaration_at_all_declines():
    a = M.avoidability([], allocated_charge=100.0, period=2025, line_code="PL-X")
    assert a["available"] is False
    assert a["value"] is None


def test_avoidable_above_the_allocated_charge_declines_rather_than_going_negative():
    """⭐ A declaration larger than the charge is a data error the client can
    fix. Trusting it would produce NEGATIVE stranded cost — a line whose exit
    saves more shared cost than it was ever charged."""
    too_much = [{"period": 2025, "line_code": "PL-X", "pool": "Corporate",
                 "avoidable_amount": 140.0}]
    a = M.avoidability(too_much, allocated_charge=100.0, period=2025,
                       line_code="PL-X")
    assert a["available"] is False
    assert "140" in a["unlocks"] and "100" in a["unlocks"]


# ── 2 · exit economics ─────────────────────────────────────────────────────

def test_exit_economics_nets_contribution_lost_against_cost_saved():
    e = M.exit_economics(contribution=53.0, avoidable=240.0, stranded=610.0)
    assert e["available"]
    assert e["value"]["net"] == pytest.approx(240.0 - 53.0)
    assert e["value"]["better_off"] is True


def test_exit_economics_can_conclude_the_company_is_worse_off():
    """⭐⭐ THE CONCLUSION FLIPS ON THE DECLARATION. Same line, same
    contribution — and the advice reverses on a number only the client holds.
    That is the entire reason ruling 1 exists."""
    e = M.exit_economics(contribution=53.0, avoidable=10.0, stranded=840.0)
    assert e["value"]["net"] == pytest.approx(10.0 - 53.0)
    assert e["value"]["better_off"] is False


def test_exit_economics_declines_without_the_declaration():
    e = M.exit_economics(contribution=53.0, avoidable=None, stranded=None)
    assert e["available"] is False
    assert "Avoidable Amount" in e["unlocks"]


# ── 3 · ⛔ ruling 3 — no horizon, and no discounting ───────────────────────

def test_the_notice_period_phases_the_saving_without_discounting_it():
    """⭐⭐ RULING 3. The notice period says WHEN the saving starts, not what it
    is worth today. Discounting it would cross into `prescience_decision`'s
    territory — the boundary the mix optimiser held by reporting contribution
    and never enterprise value."""
    p = M.avoidable_this_year(avoidable=120.0, notice_period_months=3.0)
    assert p["available"]
    # nine of twelve months of saving in the first year, undiscounted
    assert p["value"] == pytest.approx(120.0 * 9.0 / 12.0)
    assert p["basis"] == "undiscounted"


def test_a_notice_period_beyond_the_year_saves_nothing_in_it():
    p = M.avoidable_this_year(avoidable=120.0, notice_period_months=18.0)
    assert p["value"] == pytest.approx(0.0)


def test_nothing_in_this_module_discounts_or_values():
    """⛔ RULING 3, ASSERTED. AST body, docstrings excluded — §III.9, which fired
    three times in T4.5."""
    import ast
    import inspect
    tree = ast.parse(inspect.getsource(M))
    docs = set()
    for node in ast.walk(tree):
        if isinstance(node, (ast.Module, ast.FunctionDef, ast.ClassDef)):
            d = ast.get_docstring(node, clean=False)
            if d:
                docs.add(d)
    names = {n.id.lower() for n in ast.walk(tree) if isinstance(n, ast.Name)}
    names |= {n.attr.lower() for n in ast.walk(tree) if isinstance(n, ast.Attribute)}
    strings = " ".join(n.value.lower() for n in ast.walk(tree)
                       if isinstance(n, ast.Constant)
                       and isinstance(n.value, str) and n.value not in docs)
    for banned in ("npv", "discount_rate", "present_value", "enterprise_value"):
        assert banned not in names, f"{banned!r} used in the module"
        assert banned not in strings, f"{banned!r} in a rendered string"


# ── 4 · the corrective, in both states ─────────────────────────────────────

def test_without_a_declaration_the_corrective_states_its_own_premise():
    """⭐⭐ RULING 1. Leaving the assumption unstated is not a position; it is
    the defect continuing."""
    c = M.covers_variable_cost(contribution=53.0, allocated_ebit=-13.6,
                               line="Control Electronics")
    s = c["statement"]
    assert "covers its own variable cost" in s
    assert "assumes none of that cost disappears" in s.lower()
    assert "Avoidable Amount" in s
    assert c["assumes_nothing_avoidable"] is True


def test_with_a_declaration_the_corrective_quantifies_and_may_flip():
    c = M.covers_variable_cost(
        contribution=53.0, allocated_ebit=-13.6, line="Control Electronics",
        avoidable=240.0, stranded=610.0)
    s = c["statement"]
    assert "240" in s and "610" in s
    assert "better off" in s.lower()
    assert c["assumes_nothing_avoidable"] is False

    worse = M.covers_variable_cost(
        contribution=53.0, allocated_ebit=-13.6, line="Control Electronics",
        avoidable=10.0, stranded=840.0)
    assert "worse off" in worse["statement"].lower()


def test_the_corrective_never_asserts_a_conclusion_it_has_not_earned():
    """⭐ The undeclared form must not claim the company would be worse off — it
    is precisely that claim the premise was smuggling in."""
    c = M.covers_variable_cost(contribution=53.0, allocated_ebit=-13.6)
    assert "worse off, not better" not in c["statement"]


# ── 5 · the template extension ─────────────────────────────────────────────

def test_the_sheet_is_built_with_its_client_facing_labels():
    from services.api.modules.financials import templates
    from services.api.modules.financials import template_policy as policy
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(templates.build_template("us_gaap")))
    assert templates.AVOIDABILITY_SHEET in wb.sheetnames
    ws = wb[templates.AVOIDABILITY_SHEET]
    header = [ws.cell(row=policy.AVOIDABILITY_HEADER_ROW, column=c).value
              for c in range(1, len(policy.AVOIDABILITY_COLUMNS) + 1)]
    assert header == [l for l, _h in policy.AVOIDABILITY_COLUMNS]
    for needed in ("Avoidable Amount", "Notice Period (months)",
                   "Capacity Released", "Capacity Re-usable?"):
        assert needed in header


def test_prior_versions_parse_identically():
    """⭐ Fifth time on this discipline: the sheet is optional and additive, and
    a workbook without it must parse exactly as one with it does."""
    import io
    from openpyxl import load_workbook
    from services.api.modules.financials import templates
    full = templates.build_template("us_gaap")
    wb = load_workbook(io.BytesIO(full))
    del wb[templates.AVOIDABILITY_SHEET]
    buf = io.BytesIO(); wb.save(buf)
    _a, errs_with = templates.parse_workbook(full)
    _b, errs_without = templates.parse_workbook(buf.getvalue())
    assert errs_with == errs_without
