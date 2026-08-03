"""T1 — the foundation: reconciliation, the refusal, and the two composition rules.

⭐ NO ANALYTICS ARE TESTED HERE because none were built. This asserts the
properties everything above the foundation will depend on:

  · a dimensional total plus Unallocated equals the statement line EXACTLY;
  · two parallel dimensions are REFUSED, and the refusal names why;
  · a derived result takes the WEAKEST status of its inputs;
  · a confidence factor AXIOM cannot measure is EXCLUDED and SAID to be.

And it asserts the four forbidden items stay forbidden — a test is the only form
of "we decided not to" that survives a future contributor who never read CORE.
"""
import os
import tempfile

os.environ.setdefault("DATABASE_URL",
                      "sqlite:///" + tempfile.mktemp(prefix="dim-", suffix=".db"))

import pytest

from services.api.modules.financials import dimensions as D


# ── reconciliation ─────────────────────────────────────────────────────────

def test_detail_plus_unallocated_equals_the_statement_line_exactly():
    """⭐⭐ THE PROPERTY THE WHOLE MODULE EXISTS FOR. Not 'close to' — exactly,
    because the Unallocated member is stored and every chart sums it."""
    detail = {"seg-a": 620.0, "seg-b": 310.0}
    r = D.reconcile(detail, company_total=1000.0)
    assert r["detail_total"] + r["unallocated"] == pytest.approx(1000.0, abs=1e-12)
    assert r["unallocated"] == pytest.approx(70.0)
    assert r["status"] == D.UNDERALLOCATED


def test_a_complete_decomposition_reconciles_with_a_zero_residual():
    r = D.reconcile({"a": 400.0, "b": 600.0}, company_total=1000.0)
    assert r["status"] == D.RECONCILED
    assert r["unallocated"] == pytest.approx(0.0, abs=1e-12)


def test_detail_exceeding_the_company_total_is_a_suspected_overlap():
    """The shape that means a subtotal row was supplied beside its components.
    ⭐ It must never be silently dropped or clamped."""
    r = D.reconcile({"a": 700.0, "total-row": 1000.0}, company_total=1000.0)
    assert r["status"] == D.SUSPECTED_OVERLAP
    assert r["unallocated"] < 0, "a negative residual is kept as measured"


def test_no_detail_is_insufficient_detail_not_a_zero_reconciliation():
    r = D.reconcile({}, company_total=1000.0)
    assert r["status"] == D.INSUFFICIENT_DETAIL
    assert r["unallocated"] is None, "absent detail must not read as a full residual"


def test_a_measure_with_no_statement_line_says_so():
    """⭐ `units` has nothing to reconcile against. NOT_RECONCILABLE is a stated
    fact; silently exempting it is how an unreconciled figure looks reconciled."""
    r = D.reconcile({"p1": 12.0}, company_total=None)
    assert r["status"] == D.NOT_RECONCILABLE
    assert r["reason"]


def test_every_measure_declares_its_reconciliation_target():
    """⭐ Derived from MEASURES, so a measure added without a target fails here
    rather than reconciling against nothing by accident."""
    for name, spec in D.MEASURES.items():
        assert "reconciles_to" in spec, f"{name} declares no reconciliation target"
        tgt = spec["reconciles_to"]
        assert tgt is None or (isinstance(tgt, tuple) and len(tgt) == 2)


# ── the refusal ────────────────────────────────────────────────────────────

def test_parallel_dimensions_are_refused_without_a_mapping():
    """⭐⭐ `Company = Segments + Products` is the most consequential arithmetic
    error available in this module, and the licence to avoid it is a TABLE ROW
    rather than a reviewer's memory."""
    assert D.may_combine("segment", "product", mapping_exists=False) is False
    r = D.reconcile_across("segment", "product", mapping_exists=False,
                           detail_by_member={"a": 1.0}, company_total=1.0)
    assert r["status"] == D.REFUSED_PARALLEL
    assert "parallel" in r["reason"]
    assert "separately" in r["reason"], "the refusal must say what to do instead"


def test_a_supplied_mapping_licenses_the_combination():
    assert D.may_combine("segment", "product", mapping_exists=True) is True
    r = D.reconcile_across("segment", "product", mapping_exists=True,
                           detail_by_member={"a": 400.0, "b": 600.0},
                           company_total=1000.0)
    assert r["status"] == D.RECONCILED


def test_one_dimension_against_itself_never_needs_a_mapping():
    assert D.may_combine("product", "product", mapping_exists=False) is True


# ── composition rule 1: weakest status ─────────────────────────────────────

def test_a_derived_result_takes_the_weakest_status_of_its_inputs():
    assert D.weakest_status(D.OBSERVED, D.OBSERVED) == D.OBSERVED
    assert D.weakest_status(D.OBSERVED, D.ALLOCATED) == D.ALLOCATED
    assert D.weakest_status(D.OBSERVED, D.ALLOCATED, D.ESTIMATED) == D.ESTIMATED
    assert D.weakest_status(D.OBSERVED, D.UNAVAILABLE) == D.UNAVAILABLE


def test_one_unavailable_operand_beats_any_number_of_observed_ones():
    assert D.weakest_status(*([D.OBSERVED] * 50), D.UNAVAILABLE) == D.UNAVAILABLE


def test_an_unknown_status_degrades_rather_than_being_ignored():
    """A caller that invents a status must not produce a STRONGER result than one
    that admits it does not know."""
    assert D.weakest_status(D.OBSERVED, "definitely_fine") == D.UNAVAILABLE


def test_no_inputs_is_unavailable_not_observed():
    assert D.weakest_status() == D.UNAVAILABLE


# ── composition rule 2: confidence excludes what it cannot measure ─────────

def test_an_unmeasurable_factor_is_excluded_and_named():
    """⭐⭐ §III.4's coverage floor, applied to a score. Defaulting an unmeasured
    component to 1.0 would RAISE confidence for a company AXIOM knows less about."""
    out = D.score_confidence({"direct_observation_ratio": 1.0})
    assert "forecast_backtest_error" in out["excluded"]
    assert "structural_instability" in out["excluded"]
    for reason in out["excluded"].values():
        assert reason, "an exclusion without a reason is a silent one"


def test_the_band_is_derived_from_the_score_and_cannot_be_passed_in():
    hi = D.score_confidence({k: 1.0 for k in D.CONFIDENCE_FACTORS})
    lo = D.score_confidence({k: 0.1 for k in D.CONFIDENCE_FACTORS})
    assert hi["band"] == "high" and lo["band"] == "insufficient_basis"
    assert hi["score"] > lo["score"]


def test_no_measurable_factor_is_insufficient_basis_not_a_zero_score():
    """⭐ An absence of evidence is not evidence of low confidence, and the two
    must not render alike."""
    out = D.score_confidence({})
    assert out["score"] is None
    assert out["band"] == "insufficient_basis"


def test_unmeasurable_factors_never_contribute_to_the_score():
    a = D.score_confidence({"direct_observation_ratio": 0.5})
    b = D.score_confidence({"direct_observation_ratio": 0.5,
                            "forecast_backtest_error": 1.0,
                            "structural_instability": 1.0})
    assert a["score"] == b["score"], "an unmeasurable factor moved the score"


# ── the forbidden four ─────────────────────────────────────────────────────

def test_imputed_is_not_a_permitted_data_status():
    """⛔ The source document defines it; AXIOM forbids it. Filling a missing
    observation is exactly what absence propagation forbids (CORE §8a)."""
    assert "imputed" not in D.DATA_STATUSES
    assert "imputed_status" in D.FORBIDDEN


def test_the_reconciler_offers_no_gross_up_path():
    """⛔ The document permits a proportional gross-up 'unless explicitly
    approved'. There is no approval that reaches a fabricated number.

    ⭐ THE DOCSTRING IS EXCLUDED, per §III.9. The first draft matched raw source
    and failed on `reconcile`'s own docstring, which names `approve_gross_up` in
    order to forbid it — a guard keyed on text punishing the file that states its
    own rule. The assertion reads the EXECUTABLE BODY.
    """
    import ast
    import inspect
    fn = ast.parse(inspect.getsource(D.reconcile).lstrip()).body[0]
    body = fn.body[1:] if (isinstance(fn.body[0], ast.Expr)
                           and isinstance(fn.body[0].value, ast.Constant)) else fn.body
    code = " ".join(ast.unparse(n) for n in body)
    assert "gross" not in code.lower(), "the reconciler has a gross-up path"
    params = inspect.signature(D.reconcile).parameters
    assert not any("approve" in p or "gross" in p for p in params)


def test_all_four_forbidden_items_are_recorded_with_their_ruling():
    for key in ("imputed_status", "proportional_gross_up",
                "probability_across_allocation_methods",
                "multiplicative_priority_score"):
        assert key in D.FORBIDDEN
        assert "§8a" in D.FORBIDDEN[key], f"{key} does not cite its ruling"


# ── the period model is inherited, not reinvented ──────────────────────────

def test_a_dimensional_row_gets_its_period_from_the_statements_module():
    """⭐ A second period representation is how a quarterly client's dimension
    rows stop lining up with their own statements."""
    import inspect
    assert "parse_period" in inspect.getsource(D.period_of)
    assert D.period_of("2024", "annual")[0] == 2024


# ── partial data is never a blocker ────────────────────────────────────────

def test_a_workbook_with_no_dimensional_tab_still_parses():
    """⭐⭐ THE RULING, AS BEHAVIOUR. A v11 workbook has no Segments & Products
    sheet at all. That must read as "no dimensional detail supplied" — never as
    zeroes, and never as an upload error. 30% of the rows yields 30% of the
    output; the rest states why it could not be produced."""
    from services.api.modules.financials import engines as fin
    from tests.fixtures.refcases import meridian
    v = fin.validate_dataset(meridian())
    assert v["errors"] == [], (
        f"a dataset with no dimensional rows was rejected: {v['errors']}")


def test_the_template_carries_the_tab_and_the_dictionary():
    import io
    from openpyxl import load_workbook
    from services.api.modules.financials.templates import build_template
    wb = load_workbook(io.BytesIO(build_template("us_gaap")))
    assert "Segments & Products" in wb.sheetnames
    assert "Data Dictionary" in wb.sheetnames


def test_the_tab_is_long_form_so_a_client_never_sees_an_unfillable_column():
    """⭐ Long form is what makes partial completion structural: a client
    supplies ROWS, not columns. A wide sheet would present a column per measure
    and every gap as a blank to be explained."""
    from services.api.modules.financials.templates import DIMENSION_COLUMNS
    names = [c[0] for c in DIMENSION_COLUMNS]
    assert "Measure" in names and "Value" in names, "not a long-form layout"
    from services.api.modules.financials.dimensions import MEASURES
    for m in MEASURES:
        assert m not in names, f"{m} is a COLUMN — that is the wide layout"


def test_the_data_dictionary_is_derived_from_the_measure_vocabulary():
    """⭐ A hand list would go stale the first time a tier landed, and would then
    teach the client the wrong thing."""
    import io
    from openpyxl import load_workbook
    from services.api.modules.financials.templates import build_template
    from services.api.modules.financials.dimensions import MEASURES
    dd = load_workbook(io.BytesIO(build_template("us_gaap")))["Data Dictionary"]
    seen = {c.value for row in dd.iter_rows() for c in row if c.value}
    for m in MEASURES:
        assert m in seen, f"{m} is missing from the Data Dictionary"


# ── the tables actually get made ───────────────────────────────────────────

def test_the_models_are_registered_before_create_all():
    """⭐⭐ THE §4u-c DEFECT, THIRD INSTANCE, CAUGHT BY A TEST THIS TIME.

    I wrote migration 0027 first and it created nothing: `changeset.py` records
    that new ax_ tables come from `Base.metadata.create_all` at boot, and the
    Procfile starts gunicorn with NO migration step. A model imported after
    `include_accounts` is a table that is never made, and the failure surfaces
    far away as a missing relation at query time.

    This asserts the mechanism that actually runs, not the one that looks
    official.
    """
    # ⭐ IMPORT THE APP, not the model. Importing `dimensional` directly would
    # register the tables and pass regardless of whether main.py ever does —
    # the test would then assert its own import rather than the boot path.
    import services.api.main  # noqa: F401
    from services.api.accounts import Base
    for t in ("ax_dimension_member", "ax_dimension_map",
              "ax_dimension_observation"):
        assert t in Base.metadata.tables, (
            f"{t} is not registered on Base — create_all will not make it, "
            f"however complete the migration is")


def test_main_imports_the_models_before_include_accounts():
    """The ordering, asserted on the source: `create_all` runs inside
    `include_accounts`, so an import after it is too late."""
    import inspect
    from services.api import main
    src = inspect.getsource(main)
    assert "from . import dimensional" in src
    assert src.index("from . import dimensional") < src.index("include_accounts(app)")


def test_the_observation_table_is_unique_per_dataset_version():
    """⭐ A re-upload creates a new dataset version; the same member/period/
    measure may exist once per version and is never overwritten in place."""
    from services.api.dimensional import DimensionObservation
    cols = {c.name for c in DimensionObservation.__table__.columns}
    assert {"dataset_id", "member_id", "period", "measure", "basis"} <= cols
    uq = [c for c in DimensionObservation.__table__.constraints
          if c.__class__.__name__ == "UniqueConstraint"]
    assert uq and "dataset_id" in {c.name for c in list(uq)[0].columns}


def test_no_sole_owned_quantity_is_restated_here():
    """⭐ The five guarded quantities (net_debt, roic, eva, wacc, total_debt,
    invested_capital) must not acquire a second definition in this module."""
    import inspect
    from services.api.modules.financials import dimensions as DM
    src = inspect.getsource(DM)
    body = "\n".join(l for l in src.splitlines()
                     if not l.strip().startswith("#"))
    for q in ("def net_debt", "def roic", "def eva", "def wacc",
              "def total_debt", "def invested_capital"):
        assert q not in body, f"{q} is restated in dimensions.py"
