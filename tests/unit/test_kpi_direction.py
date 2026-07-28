"""KPI direction — stated, not guessed.

The B2 seed dry-run showed the name-keyword heuristic misclassifying 4 of 8
realistic KPI names: "Audit findings open: 5 against a target of 2" rendered
GREEN. These pin the replacement.
"""
import pytest
from fastapi.testclient import TestClient
from services.api.main import app
from services.api.accounts import _kpi_variance, _norm_kpi_direction, KPI_DIRECTIONS
from services.api.modules.financials import ingest


@pytest.fixture(scope="module")
def _app():
    with TestClient(app) as c:
        yield c


def test_lower_better_above_plan_is_now_unfavorable(_app):
    """THE PRODUCT-WIDE DEFECT. Meridian's Unplanned downtime at 101 against a
    95 plan came back FAVORABLE, and that verdict reached the dashboard, the
    reports and the department page alike. More downtime is not favourable."""
    old = _kpi_variance(101, 95)                       # default higher_better
    assert old["status"] == "favorable", "the old, direction-blind reading"

    fixed = _kpi_variance(101, 95, "lower_better")
    assert fixed["status"] == "unfavorable"
    assert fixed["abs"] == 6.0, "magnitude is direction-independent; only the verdict flips"
    assert fixed["direction"] == "lower_better"


def test_lower_better_below_plan_is_favorable(_app):
    v = _kpi_variance(88, 95, "lower_better")
    assert v["status"] == "favorable" and v["abs"] == -7.0


def test_higher_better_is_unchanged(_app):
    """No regression: the direction that was always assumed still behaves."""
    for actual, plan, want in ((96, 94, "favorable"), (91, 94, "unfavorable"), (94, 94, "favorable")):
        assert _kpi_variance(actual, plan, "higher_better")["status"] == want
        assert _kpi_variance(actual, plan)["status"] == want, "default matches"


def test_equality_is_favorable_in_both_directions(_app):
    """Hitting plan exactly is meeting it, whichever way is good."""
    assert _kpi_variance(50, 50, "higher_better")["status"] == "favorable"
    assert _kpi_variance(50, 50, "lower_better")["status"] == "favorable"


def test_missing_values_still_return_a_null_verdict(_app):
    for a, p in ((None, 5), (5, None), (None, None)):
        v = _kpi_variance(a, p, "lower_better")
        assert v["status"] is None and v["abs"] is None


# ── the normaliser ───────────────────────────────────────────────────────────
def test_direction_accepts_what_people_actually_type(_app):
    for v in ("lower", "Lower", "lower_better", "lower-better", "lower is better",
              "LOWER IS BETTER", "min", "minimise", "down"):
        assert _norm_kpi_direction(v) == "lower_better", v
    for v in ("higher", "higher_better", "up", "max", "maximize", "", None):
        assert _norm_kpi_direction(v) == "higher_better", v
    assert set(KPI_DIRECTIONS) == {"higher_better", "lower_better"}


def test_unrecognised_direction_warns_and_falls_back(_app):
    """Warn-never-block: a direction typo must not cost someone their KPI."""
    warnings = []
    got = _norm_kpi_direction("sideways", warnings, "Scrap rate %")
    assert got == "higher_better"
    assert len(warnings) == 1
    assert "Scrap rate %" in warnings[0]["message"] and "sideways" in warnings[0]["message"]


def test_blank_direction_records_no_warning(_app):
    """A pre-v7.5 workbook has no column I at all. Silence is not an error."""
    warnings = []
    assert _norm_kpi_direction(None, warnings, "X") == "higher_better"
    assert warnings == []


# ── template ─────────────────────────────────────────────────────────────────
def test_template_carries_column_I_and_still_accepts_older_versions(_app):
    assert ingest.TEMPLATE_VERSION == "7M-v7.6"
    # ⭐ NO VERSION GATE EXISTS. The old assertion checked membership of
    # ACCEPTED_TEMPLATE_VERSIONS, which nothing ever read. The guarantee it was
    # reaching for — an older workbook still uploads — is now absolute rather
    # than list-based, so assert the ABSENCE of any gate (CORE §7.37).
    assert not hasattr(ingest, "ACCEPTED_TEMPLATE_VERSIONS"), \
        "a version allow-list is a gate waiting to be wired up"


def test_generated_workbook_has_the_direction_header(_app):
    import io
    from openpyxl import load_workbook
    data = ingest.build_company_template(
        company_id=1, company_name="T", currency="USD", statement_units="actual",
        ownership="private", standard="us_gaap", frequency="annual")
    ws = load_workbook(io.BytesIO(data))[ingest.KPI_SHEET]
    assert ws.cell(row=ingest.KPI_HEADER_ROW, column=9).value.startswith("Direction")


def test_the_names_the_heuristic_got_wrong_are_now_stated(_app):
    """The four failures from the B2 dry-run, each now correct because the
    direction is declared rather than read out of the name."""
    for name, direction, actual, plan in (
            ("Budget variance %", "lower_better", 2.4, 2.0),
            ("Audit findings open", "lower_better", 5, 4),
            ("Change failure rate %", "lower_better", 9, 12),
            ("CAC payback months", "lower_better", 17, 19)):
        v = _kpi_variance(actual, plan, direction)
        expect = "unfavorable" if actual > plan else "favorable"
        assert v["status"] == expect, (name, v)
