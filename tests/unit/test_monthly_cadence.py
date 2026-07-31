"""Monthly periods and cadence granularity.

⭐ THE ACCEPTANCE IS `test_three_way_equivalence` — the same underlying economics
expressed monthly, quarterly and annually must produce the same enterprise value
within tolerance. Not a code read: a quarterly dataset was once discounted at
annual rates, understating EV by 4.66x, and monthly is that defect class at a
divisor three times larger. It fails silently, reading as a poor business rather
than a broken number.
"""
import os
import tempfile

os.environ.setdefault("DATABASE_URL", "sqlite:///" + tempfile.mktemp(suffix=".db"))

import pytest

from services.api.modules.financials import periods as PR
from services.api.modules.valuation import engines as VE


# ═══════════════════════════════════════════════════════════════════════════
# 1 · THE PERIOD LATTICE
# ═══════════════════════════════════════════════════════════════════════════

@pytest.mark.parametrize("value,freq,expect", [
    (202412, "monthly", 202501),        # ⭐ the carry, at twelve
    (202401, "monthly", 202402),
    (20244, "quarterly", 20251),
    (2024, "annual", 2025),
])
def test_succession_carries(value, freq, expect):
    assert PR.next_period(value, freq) == expect


@pytest.mark.parametrize("value,freq,ok", [
    (202413, "monthly", False),         # there is no thirteenth month
    (202400, "monthly", False),
    (202412, "monthly", True),
    (20245, "quarterly", False),
])
def test_validity(value, freq, ok):
    assert PR.period_is_valid(value, freq) is ok


def test_span_walks_rather_than_subtracts():
    """⭐ SUBTRACTION IS NOT A COUNT — the quarterly lesson at a new modulus.
    202501 - 202412 is 89; the true distance is one month."""
    assert PR.period_span(202412, 202501, "monthly") == 1
    assert PR.period_span(202401, 202412, "monthly") == 11
    assert PR.period_span(202401, 202501, "monthly") == 12


def test_forecast_periods_do_not_generate_a_thirteenth_month():
    out = PR.forecast_periods(202411, 4, "monthly")
    assert out == [202412, 202501, 202502, 202503]
    assert all(PR.period_is_valid(p, "monthly") for p in out)


def test_format_is_legible():
    assert PR.format_period(202403, "monthly") == "Mar 2024"
    assert PR.format_period(20243, "quarterly") == "2024Q3"
    assert PR.format_period(2024, "annual") == "2024"


# ═══════════════════════════════════════════════════════════════════════════
# 3 · ⭐ FREQUENCY IS DERIVED, NOT TRUSTED
# ═══════════════════════════════════════════════════════════════════════════

def test_the_three_encodings_cannot_collide():
    """⭐ THE PROPERTY THAT MAKES DERIVATION A READING RATHER THAN A GUESS.
    Monthly is YYYYMM and not YYYYM for exactly this reason: a five-digit 20241
    would mean both 2024 Q1 and 2024 January."""
    assert PR.derive_frequency([2024]) == "annual"
    assert PR.derive_frequency([20241]) == "quarterly"
    assert PR.derive_frequency([202401]) == "monthly"
    widths = {len(str(v)) for v in (2024, 20241, 202401)}
    assert len(widths) == 3, "the encodings must differ in digit count"


def test_a_mixed_set_derives_to_None_rather_than_a_majority():
    """⭐ A dataset whose periods are not all one shape has NO single frequency,
    and picking the majority would silently discard the rest."""
    assert PR.derive_frequency([2024, 202401]) is None
    assert PR.derive_frequency([]) is None


def test_a_lying_label_does_not_change_the_divisor():
    """⭐⭐ THE RULING, ENCODED. The divisor must match the ACTUAL period spacing
    or the discounting is wrong whatever the label says — and being wrong there
    is the 4.66x defect."""
    liar = {"periods": {"frequency": "annual",
                        "historical": [202401, 202402], "forecast": []}}
    assert VE.periods_per_year(liar) == 12
    assert PR.frequency_of(liar) == "monthly"


def test_a_disagreement_is_REPORTED_not_silently_reconciled(  ):
    """⭐ A label disagreeing with its own periods is a live inconsistency."""
    liar = {"periods": {"frequency": "quarterly",
                        "historical": [202401], "forecast": []}}
    chk = PR.frequency_check(liar)
    assert chk["agree"] is False
    assert chk["declared"] == "quarterly" and chk["derived"] == "monthly"
    assert "declared" in chk["reason"] and "202401" not in chk["reason"]
    assert chk["effective"] == "monthly", "compute follows the values"


def test_agreement_is_the_normal_case_and_changes_nothing():
    """Measured before changing: all 36 stored datasets agree across column,
    payload and values, so deriving changes no existing result."""
    for freq, hist in (("annual", [2023, 2024]),
                       ("quarterly", [20241, 20242]),
                       ("monthly", [202401, 202402])):
        d = {"periods": {"frequency": freq, "historical": hist, "forecast": []}}
        chk = PR.frequency_check(d)
        assert chk["agree"] and chk["effective"] == freq


def test_a_dataset_with_no_declaration_still_derives():
    """Datasets written before the key existed read as annual — and now they read
    as annual because their VALUES are annual, which is a stronger claim."""
    d = {"periods": {"historical": [2023, 2024], "forecast": []}}
    assert PR.frequency_of(d) == "annual"
    assert VE.periods_per_year(d) == 1


# ═══════════════════════════════════════════════════════════════════════════
# 2 · ⭐⭐ THE ACCEPTANCE — THREE-WAY EQUIVALENCE
# ═══════════════════════════════════════════════════════════════════════════

def _dataset(freq, n_hist, n_fcst, annual_revenue, first_hist, first_fcst):
    """The SAME economics at three granularities.

    Revenue per period is the annual figure divided by periods-per-year, so the
    business earns the same amount per year in all three. Everything else scales
    with it.
    """
    ppy = {"annual": 1, "quarterly": 4, "monthly": 12}[freq]
    rev = annual_revenue / ppy
    hist = [first_hist]
    for _ in range(n_hist - 1):
        hist.append(PR.next_period(hist[-1], freq))
    fcst = [first_fcst]
    for _ in range(n_fcst - 1):
        fcst.append(PR.next_period(fcst[-1], freq))

    def series(v):
        return {**{str(p): v for p in hist}, **{str(p): v for p in fcst}}

    return {
        "company": {"name": "Equiv", "standard": "us_gaap", "ownership": "private",
                    "tax_rate": 0.25, "risk_free_rate": 0.04,
                    "market_risk_premium": 0.055, "cost_of_debt": 0.06,
                    "beta": 1.0, "unlevered_industry_beta": 1.0,
                    "target_debt_to_equity": 0.4, "dlom": 0.0,
                    "size_premium": 0.0, "specific_risk_premium": 0.0,
                    "shares_outstanding": 1000.0},
        "periods": {"historical": hist, "forecast": fcst, "frequency": freq},
        # ⭐ THE FIXTURE'S OWN KEY NAMES. The first version of this used
        # `cost_of_sales`, `operating_expenses` and `depreciation_amortisation`;
        # the engine reads `cogs`, `opex` and `depreciation_amortization`, so
        # every operand came back absent and FCFF derived to None. Absence
        # propagated correctly — the dataset was wrong, not the engine.
        "income_statement": {
            "revenue": series(rev),
            "cogs": series(rev * 0.55),
            "opex": series(rev * 0.25),
            "depreciation_amortization": series(rev * 0.05),
            "interest_expense": series(rev * 0.01),
        },
        "balance_sheet": {
            "cash": series(annual_revenue * 0.10),
            "other_current_assets": series(annual_revenue * 0.20),
            "noncurrent_assets": series(annual_revenue * 0.60),
            "current_liabilities_ex_debt": series(annual_revenue * 0.15),
            "short_term_debt": series(annual_revenue * 0.05),
            "long_term_debt": series(annual_revenue * 0.25),
            "preferred_equity": series(0.0),
            "minority_interest": series(0.0),
            "total_equity": series(annual_revenue * 0.45),
        },
        "cash_flow": {
            "capex": series(rev * 0.04),
            "net_borrowing": series(0.0),
            "dividends": series(0.0),
        },
    }


def test_three_way_equivalence():
    """⭐⭐ THE ACCEPTANCE FOR THIS LANE. Five years of the same business,
    expressed annually, quarterly and monthly, must value the same.

    ⭐ TOLERANCE IS NOT ZERO AND MUST NOT BE TUNED TO ZERO. Cash arriving monthly
    is genuinely worth slightly more than the same cash arriving in annual lumps;
    the existing quarterly comment measures that residual at ~3% and says it
    "should not be tuned away". A test demanding exact equality would be a test
    demanding the arithmetic be wrong.
    """
    YEARS = 5
    annual = _dataset("annual", 3, YEARS, 1000.0, 2021, 2024)
    quarterly = _dataset("quarterly", 12, YEARS * 4, 1000.0, 20211, 20241)
    monthly = _dataset("monthly", 36, YEARS * 12, 1000.0, 202101, 202401)

    evs = {}
    for name, data in (("annual", annual), ("quarterly", quarterly),
                       ("monthly", monthly)):
        out = VE.run(data, "proforma")
        ev = out["deterministic"]["enterprise_value"]
        assert isinstance(ev, (int, float)), f"{name} produced no EV"
        evs[name] = float(ev)

    base = evs["annual"]
    assert base > 0, f"the annual case must value positively: {evs}"
    for name in ("quarterly", "monthly"):
        ratio = evs[name] / base
        assert 0.85 <= ratio <= 1.15, (
            f"{name} EV {evs[name]:.2f} vs annual {base:.2f} — ratio {ratio:.3f}. "
            f"The 4.66x defect looked exactly like this.")


def test_the_divisor_is_what_makes_them_agree():
    """⭐ THE NEGATIVE CONTROL. Force the monthly case to discount at annual rates
    — the pre-fix behaviour — and require it to DISAGREE. Without this, the
    equivalence test could pass because the three cases were too similar to
    distinguish."""
    monthly = _dataset("monthly", 36, 60, 1000.0, 202101, 202401)
    annual = _dataset("annual", 3, 5, 1000.0, 2021, 2024)
    good = VE.run(monthly, "proforma")["deterministic"]["enterprise_value"]
    base = VE.run(annual, "proforma")["deterministic"]["enterprise_value"]

    orig = VE.PERIODS_PER_YEAR.copy()
    try:
        VE.PERIODS_PER_YEAR["monthly"] = 1          # the defect, reintroduced
        bad = VE.run(monthly, "proforma")["deterministic"]["enterprise_value"]
    finally:
        VE.PERIODS_PER_YEAR.clear(); VE.PERIODS_PER_YEAR.update(orig)

    assert abs(good / base - 1.0) < abs(bad / base - 1.0), \
        "the divisor made no difference — the equivalence proves nothing"
    assert not (0.85 <= bad / base <= 1.15), \
        f"discounting monthly at annual rates should FAIL the tolerance: {bad/base:.3f}"


def test_period_rates_compound_back_to_the_annual_rate():
    for ppy in (1, 4, 12):
        r = VE.to_period_rate(0.10, ppy)
        assert abs((1 + r) ** ppy - 1 - 0.10) < 1e-12


def test_dividing_instead_of_compounding_is_not_what_happens():
    """⭐ Dividing by 12 would overstate the monthly rate and understate every PV."""
    assert VE.to_period_rate(0.10, 12) != pytest.approx(0.10 / 12)
    assert VE.to_period_rate(0.10, 12) < 0.10 / 12


# ═══════════════════════════════════════════════════════════════════════════
# 4-5 · CADENCE FOLLOWS THE DATA, AND THE PACK SAYS SO
# ═══════════════════════════════════════════════════════════════════════════

def _p(hist):
    return {"periods": {"historical": hist, "forecast": []}}


@pytest.mark.parametrize("hist,expect", [
    ([2022, 2023], "annual"),
    ([20241, 20242], "quarterly"),
    ([202401, 202402], "monthly"),
    ([2022, 2023, 202401], "monthly"),      # ⭐ mixed: finest wins
    ([2022, 20241], "quarterly"),
])
def test_cadence_follows_the_data(hist, expect):
    from services.api import pack as P
    assert P.cadence_for(_p(hist))["cadence"] == expect


def test_mixed_granularity_is_the_NORMAL_case_and_is_named():
    """⭐ Annual history with monthly recent periods. A company-level setting
    would be wrong for most of its own series."""
    from services.api import pack as P
    c = P.cadence_for(_p([2021, 2022, 2023, 202401, 202402]))
    assert c["cadence"] == "monthly"
    assert c["mixed"] is True
    assert c["available"] == ["monthly", "annual"]
    assert "also carries" in c["reason"]


def test_cadence_is_not_a_company_level_setting():
    """⭐ It takes the DATA, not a company id — so it cannot consult a per-company
    flag even if one were added."""
    import inspect

    from services.api import pack as P
    sig = inspect.signature(P.cadence_for)
    assert list(sig.parameters)[0] != "cid"
    src = inspect.getsource(P.cadence_for)
    assert "PackSchedule" not in src and "db.query" not in src


def test_a_pack_states_its_cadence_and_the_age_of_its_financials():
    """⭐ A monthly pack carrying quarterly financials is HONEST. A monthly pack
    silently carrying two-month-old financials is not."""
    from services.api import pack as P
    data = _p([2022, 2023, 20241])
    cad = P.cadence_for(data)
    age = P.financial_input_age(data, cad)
    assert age["present"]
    assert age["newest_period"] == 20241
    assert age["pack_cadence"] == "quarterly"
    assert age["note"]


def test_a_coarser_financial_than_the_cadence_is_NAMED():
    from services.api import pack as P
    frozen = {"classes": {"active_financial_dataset": {
        "present": True,
        "payload": {"periods": {"historical": [202401, 202402, 20243],
                                "forecast": []}}}}}
    cad = P.cadence_for(frozen)
    age = P.financial_input_age(frozen, cad)
    assert cad["cadence"] == "monthly"
    assert age["newest_frequency"] == "quarterly"
    assert "coarser than the cadence" in age["note"]


def test_absence_declares_for_cadence_too():
    from services.api import pack as P
    assert P.cadence_for(None)["present"] is False
    assert P.cadence_for(_p([]))["reason"]
    assert P.financial_input_age(_p([]))["present"] is False


# ═══════════════════════════════════════════════════════════════════════════
# 6 · ⭐ A GRANULARITY SWITCH MID-SERIES MUST NOT BREAK THE SERIES
# ═══════════════════════════════════════════════════════════════════════════

def _frozen(payload):
    from services.api.pack import _jsonable, cadence_for, financial_input_age
    f = {"schema": "7s1.1", "classes": {
        "active_financial_dataset": {"present": True, "payload": payload,
                                     "version": 1},
        "forecast_sets": {"present": False, "reason": "n/a"},
        "initiatives": {"present": False, "reason": "n/a"},
        "cfo_overrides": {"present": False, "reason": "n/a"},
        "valuation_runs": {"present": False, "reason": "n/a"},
    }, "versions": {}}
    f["cadence"] = cadence_for(f)
    f["financial_input_age"] = financial_input_age(f, f["cadence"])
    return _jsonable(f)


def test_the_bridge_renders_across_a_granularity_switch():
    """⭐ A prior QUARTERLY pack and a subsequent MONTHLY pack are BOTH VALID.
    The series must not break because the client started reporting monthly."""
    from services.api import value_bridge as VB
    q = _frozen(_dataset("quarterly", 12, 20, 1000.0, 20211, 20241))
    m = _frozen(_dataset("monthly", 36, 60, 1000.0, 202101, 202401))
    br = VB.build(q, m)
    assert br["from"]["equity_value"] is not None
    assert br["to"]["equity_value"] is not None
    assert br["total_movement"] is not None
    # ⭐ the residual is still shown, not suppressed by the frequency change
    assert (br["residual"] is not None) ^ (br["residual_absent"] is not None)


def test_each_side_is_valued_at_ITS_OWN_frequency():
    """⭐ THE FAILURE THIS PREVENTS. Valuing the monthly side with the quarterly
    side's divisor is the 4.66x defect arriving through the bridge instead of
    through the engine."""
    from services.api import value_bridge as VB
    q_payload = _dataset("quarterly", 12, 20, 1000.0, 20211, 20241)
    m_payload = _dataset("monthly", 36, 60, 1000.0, 202101, 202401)
    assert VE.periods_per_year(q_payload) == 4
    assert VE.periods_per_year(m_payload) == 12
    br = VB.build(_frozen(q_payload), _frozen(m_payload))
    ev_q, ev_m = br["from"]["equity_value"], br["to"]["equity_value"]
    # same economics, so the two sides must be close despite different divisors
    assert 0.85 <= ev_m / ev_q <= 1.15, \
        f"a granularity switch moved equity value by {ev_m / ev_q:.3f}"


def test_the_bridge_DECLARES_when_it_cannot_render_across_the_switch():
    """⭐ Or declares that it cannot — never renders an empty bridge."""
    from services.api import value_bridge as VB
    q = _frozen(_dataset("quarterly", 12, 20, 1000.0, 20211, 20241))
    empty = {"schema": "7s1.1", "classes": {
        "active_financial_dataset": {"present": False,
                                     "reason": "no active dataset"}},
        "versions": {}}
    br = VB.build(q, empty)
    assert br["total_movement"] is None
    assert br["residual"] is None
    assert br["residual_absent"], "the bridge must say why it cannot reconcile"
    assert br["to"]["absent"]


def test_the_two_packs_state_different_cadences_and_both_are_valid():
    q = _frozen(_dataset("quarterly", 12, 20, 1000.0, 20211, 20241))
    m = _frozen(_dataset("monthly", 36, 60, 1000.0, 202101, 202401))
    assert q["cadence"]["cadence"] == "quarterly"
    assert m["cadence"]["cadence"] == "monthly"
    assert q["financial_input_age"]["present"]
    assert m["financial_input_age"]["present"]


def test_newest_period_is_correct_across_a_MIXED_series():
    """⭐ THE DEFECT THIS LANE CAUGHT IN ITSELF. `max()` over mixed encodings
    returns the wrong period: quarterly 20243 is numerically SMALLER than monthly
    202401, so a stale quarter would be reported as the newest input."""
    # ⭐ MY FIRST ASSERTION HERE WAS WRONG, NOT THE CODE: 2024 Q3 runs Jul-Sep and
    # IS later than January 2024. The ordinal compares period STARTS, which is
    # the only comparison that means anything across frequencies.
    assert PR.newest_period([20243, 202401]) == 20243      # Q3 (Jul) > Jan
    assert PR.newest_period([20243, 202410]) == 202410     # Oct > Q3 (Jul)
    assert PR.newest_period([2023, 20241, 202402]) == 202402
    assert PR.newest_period([]) is None
    # ⭐ AND max() IS WRONG ON THAT SAME PAIR — it picks the numerically larger
    # encoding, which here is the EARLIER period.
    assert max([20243, 202410]) == 202410
    assert max([20243, 202401]) == 202401, "numerically larger"
    assert PR.newest_period([20243, 202401]) != max([20243, 202401]), \
        "max() and the ordinal must disagree here — that is the whole defect"


# ═══════════════════════════════════════════════════════════════════════════
# ⭐ THE PARSER ACCEPTS v8 AND v9 — CONFIRMED, NOT ASSUMED
# ═══════════════════════════════════════════════════════════════════════════

def test_no_version_gate_exists_on_either_template_path():
    """⭐ CORE §7.37 — version is FORENSIC METADATA, never a precondition for
    upload. Re-asserted here because a bump is exactly when someone adds one."""
    from services.api.modules.financials import ingest
    assert not hasattr(ingest, "ACCEPTED_TEMPLATE_VERSIONS")
    assert not any("ACCEPTED" in n and "VERSION" in n for n in dir(ingest)), \
        "a version allow-list under another name is still a gate"


def test_a_v8_workbook_still_parses_after_the_v9_bump():
    """⭐ v8 FILES ARE IN THE WILD. The bump must not strand a customer holding
    last quarter's workbook."""
    import io

    from openpyxl import load_workbook

    from services.api.modules.financials import templates as T
    wb = load_workbook(io.BytesIO(T.build_template("us_gaap")))
    wb["Instructions"]["A1"] = "AXIOM-FIN-TEMPLATE v8 us_gaap"
    b = io.BytesIO(); wb.save(b)
    _ds, issues = T.parse_workbook(b.getvalue())
    assert not any("not an AXIOM" in (i.get("error") or "") for i in issues), \
        "a v8 stamp was rejected"
    assert not any("version" in (i.get("error") or "").lower() for i in issues)


def test_monthly_parses_as_ABSENT_for_a_prior_version_not_as_an_error():
    """⭐ THE TOLERANCE, STATED. A v8 file has no monthly columns to read — which
    is a FACT ABOUT THE FILE, not a failure of it. It must parse, and its
    frequency must derive to what its periods actually are.
    """
    import io

    from openpyxl import load_workbook

    from services.api.modules.financials import periods as _PR
    from services.api.modules.financials import templates as T
    wb = load_workbook(io.BytesIO(T.build_template("us_gaap")))
    wb["Instructions"]["A1"] = "AXIOM-FIN-TEMPLATE v8 us_gaap"
    b = io.BytesIO(); wb.save(b)
    ds, issues = T.parse_workbook(b.getvalue())
    # ⭐ NARROWED, AND THE FIRST VERSION WAS WRONG. An EMPTY template naturally
    # reports missing company fields — that is content validation doing its job,
    # not a version rejection. Asserting "no errors at all" tested whether the
    # blank workbook was filled in. The claim that matters is that NOTHING
    # rejects it for its VERSION.
    version_errors = [i for i in issues
                      if "version" in (i.get("error") or "").lower()
                      or "not an AXIOM" in (i.get("error") or "")
                      or "template" in (i.get("error") or "").lower()]
    assert version_errors == [], f"a v8 stamp was rejected: {version_errors}"
    if ds and ds.get("periods"):
        hist = ds["periods"].get("historical") or []
        if hist:
            # ⭐ no monthly period appears, and that reads as ABSENCE
            assert _PR.derive_frequency(hist) != "monthly"
            assert all(len(str(p)) != 6 for p in hist)


def test_the_current_template_stamps_v9_and_offers_monthly():
    from services.api.modules.financials import ingest
    from services.api.modules.financials import template_policy as policy
    assert policy.VERSION_MAJOR == 9
    assert ingest.TEMPLATE_VERSION == "7M-v9.0"
    assert "monthly" in ingest.HISTORY_COLS and "monthly" in ingest.FORECAST_COLS


def test_the_three_frequencies_all_build_a_company_workbook():
    """⭐ The COMPANY template is the frequency-aware one — `build_template` is
    the generic download and takes no frequency. Monthly must produce a workbook
    with monthly columns, not silently fall back to quarterly."""
    from services.api.modules.financials import ingest
    for freq in ("annual", "quarterly", "monthly"):
        assert freq in ingest.HISTORY_COLS, f"{freq} has no history budget"
        assert freq in ingest.FORECAST_COLS, f"{freq} has no forecast budget"
    # ⭐ AND THEY DIFFER — a lookup that returned the same number for all three
    # would pass a membership check while giving monthly the quarterly caps.
    assert len({ingest.HISTORY_COLS[f] for f in
                ("annual", "quarterly", "monthly")}) == 3
    assert ingest.FORECAST_COLS["monthly"] == 60
