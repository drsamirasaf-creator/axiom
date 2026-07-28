"""participant_upload.py — 0/164 statements before this file. Nothing executed it.

⭐ WHY A DEFECT HERE LOOKS LIKE A DATA PROBLEM, NOT A BUG.

This module assigns IDENTITY AND ATTRIBUTES TO PEOPLE from a file the customer
authored. That is the whole reason its failures are misattributed: the only
artifact in view is the customer's spreadsheet, so anything wrong is read as
something they typed.

  · The role union is keyed on EMAIL across three tabs. Get it wrong and a
    person's department silently changes — and downstream their responses carry
    the other department's name, so a department shows fewer respondents and
    reads as "they didn't engage". That is exactly the shape that hid 312
    responses on company 39.
  · Department matching is case-insensitive against the org chart, and an
    unmatched name becomes a COLLISION, never an auto-created department. A
    matching defect surfaces as "the customer typed the department wrong".
  · The seniority vocabulary is closed. Drop a band and those participants land
    with seniority=None, so the §4u slice shows them missing — "people didn't
    fill it in".
  · Row numbers are reported against _DATA_START=4. Off by one and the admin is
    sent to the wrong row: "the error message is confusing".
  · The version stamp. Until today this block REJECTED a file whose named cell
    the customer's spreadsheet app had dropped. The customer saw "your template
    is invalid" — a template problem, in their reading. It was ours.

None of these produce a number that looks wrong. They produce people who are
missing, mis-departmented, or a file that is refused.
"""
import io
import pytest
from openpyxl import load_workbook

from services.api.participant_upload import (
    build_participant_template, parse_participant_workbook,
    TABS, COLUMNS, ROLE_OF_TAB, SENIORITY_BANDS, VERSION, _DATA_START,
)

DEPTS = ["Operations", "Finance and Accounting", "Human Resources"]


def _fill(rows_by_tab, departments=DEPTS):
    """Build the real template and fill it, so the parser meets the writer."""
    wb = load_workbook(io.BytesIO(build_participant_template(departments)))
    for tab, rows in rows_by_tab.items():
        ws = wb[tab]
        cols = COLUMNS[tab]
        for i, row in enumerate(rows):
            for c, header in enumerate(cols):
                ws.cell(row=_DATA_START + i, column=c + 1, value=row.get(header))
    out = io.BytesIO(); wb.save(out)
    return out.getvalue()


def _p(**kw):
    return kw


# ── the version stamp: recorded, never gated ────────────────────────────────
def test_version_is_recorded_not_gated():
    blob = _fill({"Assessors": [_p(**{"Full Name": "A One", "Email": "a1@x.com",
                                      "Department": "Operations",
                                      "Seniority Band": "Mid-level"})]})
    out = parse_participant_workbook(blob, DEPTS)
    assert out["version"] == VERSION
    assert out["errors"] == [], out["errors"]
    assert "version_ok" not in out, \
        "version_ok is back — a field that exists is a field something branches on"


def test_a_workbook_whose_stamp_was_LOST_still_parses_completely():
    """⭐ THE CUSTOMER'S ACTUAL FILE. Their spreadsheet application dropped the
    PLU_VERSION defined name; AXIOM rejected a real, complete roster over it.
    Losing the stamp must cost the forensic metadata and nothing else."""
    blob = _fill({"Assessors": [_p(**{"Full Name": "A One", "Email": "a1@x.com",
                                      "Department": "Operations",
                                      "Seniority Band": "Junior"})]})
    wb = load_workbook(io.BytesIO(blob))
    del wb.defined_names["PLU_VERSION"]
    buf = io.BytesIO(); wb.save(buf)

    out = parse_participant_workbook(buf.getvalue(), DEPTS)
    assert out["version"] is None
    assert out["errors"] == [], f"a missing stamp blocked the parse: {out['errors']}"
    assert out["participants"]["a1@x.com"]["roles"] == ["assessor"]


# ── the role union ──────────────────────────────────────────────────────────
def test_one_person_on_two_tabs_unions_their_roles():
    blob = _fill({
        "Assessors": [_p(**{"Full Name": "Dual Role", "Email": "d@x.com",
                            "Department": "Operations", "Seniority Band": "Executive"})],
        "Viewers": [_p(**{"Full Name": "Dual Role", "Email": "d@x.com"})],
    })
    out = parse_participant_workbook(blob, DEPTS)
    assert out["errors"] == [], out["errors"]
    roles = out["participants"]["d@x.com"]["roles"]
    assert set(roles) == {"assessor", "viewer"}, roles


def test_a_later_tab_does_not_erase_the_department_set_earlier():
    """The silent-change case: Viewers carries no Department column, and must
    not blank what Assessors established for the same person."""
    blob = _fill({
        "Assessors": [_p(**{"Full Name": "Keeps Dept", "Email": "k@x.com",
                            "Department": "Finance and Accounting",
                            "Seniority Band": "Senior management"})],
        "Viewers": [_p(**{"Full Name": "Keeps Dept", "Email": "k@x.com"})],
    })
    out = parse_participant_workbook(blob, DEPTS)
    assert out["participants"]["k@x.com"]["department"] == "Finance and Accounting"
    assert out["participants"]["k@x.com"]["seniority"] == "Senior management"


def test_email_is_the_identity_key_and_is_case_folded():
    blob = _fill({
        "Assessors": [_p(**{"Full Name": "Mixed Case", "Email": "MiXeD@X.com",
                            "Department": "Operations", "Seniority Band": "Junior"})],
        "Viewers": [_p(**{"Full Name": "Mixed Case", "Email": "mixed@x.com"})],
    })
    out = parse_participant_workbook(blob, DEPTS)
    assert list(out["participants"]) == ["mixed@x.com"], out["participants"].keys()
    assert set(out["participants"]["mixed@x.com"]["roles"]) == {"assessor", "viewer"}


# ── departments: matched case-insensitively, never invented ─────────────────
def test_department_matches_case_insensitively_and_returns_the_ORG_CHART_spelling():
    # ⭐ UPPERCASE, NOT lowercase. `dep_lookup` is keyed on the lowered org-chart
    # name, so a lowercase input still matches even if the lookup stops lowering
    # what the CUSTOMER typed — the mutation check caught that this test was
    # passing for the wrong reason. Only a case that differs from BOTH the org
    # chart and the lookup key discriminates.
    blob = _fill({"Assessors": [_p(**{"Full Name": "Shouty Dept", "Email": "l@x.com",
                                      "Department": "OPERATIONS",
                                      "Seniority Band": "Mid-level"})]})
    out = parse_participant_workbook(blob, DEPTS)
    assert out["errors"] == [], out["errors"]
    assert out["participants"]["l@x.com"]["department"] == "Operations", \
        "the org-chart spelling must win, or responses carry a name no department has"


def test_an_unknown_department_is_a_collision_and_is_never_auto_created():
    blob = _fill({"Assessors": [_p(**{"Full Name": "Ghost Dept", "Email": "g@x.com",
                                      "Department": "Department That Is Not There",
                                      "Seniority Band": "Junior"})]})
    out = parse_participant_workbook(blob, DEPTS)
    assert out["collisions"], "an unmatched department produced no collision"
    c = out["collisions"][0]
    assert c["field"] == "Department" and c["email"] == "g@x.com"
    assert "g@x.com" not in out["participants"], \
        "a row with an unknown department was accepted anyway"


# ── the closed vocabularies and per-tab requirements ────────────────────────
def test_seniority_outside_the_five_bands_is_rejected():
    blob = _fill({"Assessors": [_p(**{"Full Name": "Bad Band", "Email": "b@x.com",
                                      "Department": "Operations",
                                      "Seniority Band": "Middle Management"})]})
    out = parse_participant_workbook(blob, DEPTS)
    assert any("five allowed bands" in e["message"] for e in out["errors"]), out["errors"]


@pytest.mark.parametrize("band", SENIORITY_BANDS)
def test_every_documented_band_is_accepted(band):
    blob = _fill({"Assessors": [_p(**{"Full Name": "Band OK", "Email": "ok@x.com",
                                      "Department": "Operations",
                                      "Seniority Band": band})]})
    out = parse_participant_workbook(blob, DEPTS)
    assert out["errors"] == [], f"band {band!r} rejected: {out['errors']}"


def test_assessors_require_department_and_band_others_do_not():
    blob = _fill({
        "Assessors": [_p(**{"Full Name": "No Dept", "Email": "nd@x.com"})],
        "Viewers": [_p(**{"Full Name": "Viewer Fine", "Email": "vf@x.com"})],
    })
    out = parse_participant_workbook(blob, DEPTS)
    msgs = [e["message"] for e in out["errors"] if e["email"] == "nd@x.com"]
    assert any("Department is required" in m for m in msgs), msgs
    assert any("Seniority Band is required" in m for m in msgs), msgs
    assert "vf@x.com" in out["participants"], "a viewer was held to assessor rules"


def test_decision_makers_require_a_title():
    blob = _fill({"Decision Makers": [_p(**{"Full Name": "No Title", "Email": "nt@x.com"})]})
    out = parse_participant_workbook(blob, DEPTS)
    assert any("Title/Role is required" in e["message"] for e in out["errors"]), out["errors"]


# ── row-level reporting ─────────────────────────────────────────────────────
def test_error_rows_point_at_the_row_the_admin_will_look_at():
    """Off-by-one here sends someone to the wrong line and reads as a confusing
    message rather than a wrong one."""
    blob = _fill({"Assessors": [
        _p(**{"Full Name": "Good", "Email": "good@x.com",
              "Department": "Operations", "Seniority Band": "Junior"}),
        _p(**{"Full Name": "Bad Email", "Email": "not-an-email",
              "Department": "Operations", "Seniority Band": "Junior"}),
    ]})
    out = parse_participant_workbook(blob, DEPTS)
    bad = [e for e in out["errors"] if "not a valid address" in e["message"]]
    assert bad, out["errors"]
    assert bad[0]["row"] == _DATA_START + 1, f"row reported as {bad[0]['row']}"


def test_a_duplicate_on_one_tab_names_the_earlier_row():
    blob = _fill({"Assessors": [
        _p(**{"Full Name": "First", "Email": "dup@x.com",
              "Department": "Operations", "Seniority Band": "Junior"}),
        _p(**{"Full Name": "Second", "Email": "dup@x.com",
              "Department": "Operations", "Seniority Band": "Junior"}),
    ]})
    out = parse_participant_workbook(blob, DEPTS)
    dups = [e for e in out["errors"] if "duplicate of row" in e["message"]]
    assert dups, out["errors"]
    assert f"row {_DATA_START}" in dups[0]["message"], dups[0]["message"]


def test_is_ceo_survives_into_the_participant_map():
    blob = _fill({"Decision Makers": [_p(**{"Full Name": "The Chief", "Email": "ceo@x.com",
                                            "Title/Role": "CEO", "Is CEO": "Yes"})]})
    out = parse_participant_workbook(blob, DEPTS)
    assert out["errors"] == [], out["errors"]
    assert out["participants"]["ceo@x.com"]["is_ceo"] is True


def test_a_file_that_is_not_a_workbook_is_reported_not_raised():
    out = parse_participant_workbook(b"this is not xlsx", DEPTS)
    assert out["errors"] and "readable .xlsx" in out["errors"][0]["message"]
    assert out["participants"] == {}


def test_the_public_sample_template_carries_no_tenant_departments():
    """The ungated sample must carry FORMAT only — never a real org chart."""
    from services.api.participant_upload import SAMPLE_DEPARTMENTS
    wb = load_workbook(io.BytesIO(build_participant_template(SAMPLE_DEPARTMENTS)))
    text = " ".join(str(c.value) for ws in wb.worksheets
                    for row in ws.iter_rows() for c in row if c.value)
    for real in DEPTS:
        assert real not in text
