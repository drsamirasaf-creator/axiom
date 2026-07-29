"""Period entry format and display labels (Parts B and C).

⭐ THE STORAGE ENCODING DID NOT CHANGE. YYYYQ integers remain canonical for
sorting, joining and comparison — A3 confirmed nothing computes on a label, which
is what made this additive rather than a migration.
"""
import io
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from services.api.modules.financials import ingest
from services.api.modules.financials.periods import (
    parse_period, format_period, entry_label, period_labels, PeriodParseError,
)
from services.api.report_pdf import _fmt_period as pdf_period
from services.api.reporting import _fmt_period as pptx_period


# ── B2: every accepted entry form, and the interpretation ───────────────────
@pytest.mark.parametrize("raw", ["2024Q1", "2024-Q1", "2024 Q1", "Q1 2024", "2024q1", "20241", 20241])
def test_every_accepted_form_normalises_to_the_stored_integer(raw):
    value, how = parse_period(raw, "quarterly")
    assert value == 20241
    assert how == "2024 Q1", how


def test_the_interpretation_is_reported_not_assumed():
    """A parser that silently accepts leaves the customer trusting it read them
    right; the near-ambiguous forms are exactly where that matters.

    Covers BOTH parse routes — the typed forms and the legacy 5-digit branch —
    because a mutation to the legacy branch's interpretation survived a version
    of this test that only exercised the typed ones."""
    assert parse_period("Q4 2023", "quarterly")[1] == "2023 Q4"
    assert parse_period("2023-Q4", "quarterly")[1] == "2023 Q4"
    assert parse_period("20234", "quarterly")[1] == "2023 Q4", "legacy path"
    assert parse_period(20234, "quarterly")[1] == "2023 Q4"


@pytest.mark.parametrize("bad", ["20245", "2024Q5", "next quarter", "", "Q5 2024", "24Q1"])
def test_genuine_ambiguity_is_rejected(bad):
    with pytest.raises(PeriodParseError):
        parse_period(bad, "quarterly")


def test_the_rejection_names_what_is_wrong_and_what_to_type():
    with pytest.raises(PeriodParseError) as e:
        parse_period("20245", "quarterly")
    assert "quarter" in str(e.value) and "2024Q1" in str(e.value)


def test_legacy_yyyyq_is_still_accepted():
    """Files in the wild carry it — the template shipped that form. Rejecting it
    would repeat the version-stamp mistake: refusing a real, complete file over a
    representation AXIOM itself chose."""
    assert parse_period("20234", "quarterly")[0] == 20234


def test_annual_entry_is_a_plain_year():
    assert parse_period(2024, "annual") == (2024, "2024")
    assert parse_period("2024", "annual") == (2024, "2024")
    with pytest.raises(PeriodParseError):
        parse_period("2024Q1", "annual")


# ── B1: the generated template ──────────────────────────────────────────────
def _tpl(freq):
    return ingest.build_company_template(
        company_id=1, company_name="T", currency="USD", statement_units="actual",
        ownership="private", standard="us_gaap", frequency=freq,
        last_historical_year=2025)


def test_quarterly_row4_is_text_in_the_canonical_entry_form():
    """⭐ TEXT FORMAT IS THE POINT. A bare 5-digit number in a period cell is
    what Excel reaches for a date with — the coercion class that produced this."""
    ws = load_workbook(io.BytesIO(_tpl("quarterly")))["Income Statement"]
    assert ws["B4"].value == "2020Q1"
    assert ws["B4"].number_format == "@"
    fc = next(c for c in range(2, ws.max_column + 1)
              if str(ws.cell(row=3, column=c).value or "").lower() == "forecast")
    assert ws.cell(row=4, column=fc).number_format == "@"
    assert ws.cell(row=4, column=fc).value is None, "forecast labels still ship blank"


def test_annual_row4_is_unchanged():
    ws = load_workbook(io.BytesIO(_tpl("annual")))["Income Statement"]
    assert ws["B4"].value == 2020 and isinstance(ws["B4"].value, int)
    assert ws["B4"].number_format == '0000"-A"'


def test_the_tooltip_and_instructions_teach_the_entry_form_not_the_encoding():
    wb = load_workbook(io.BytesIO(_tpl("quarterly")))
    ws = wb["Income Statement"]
    prompt = [d for d in ws.data_validations.dataValidation if d.promptTitle][0].prompt
    assert "2024Q1" in prompt and "YYYYQ" not in prompt
    inst = " ".join(str(wb["Instructions"].cell(row=r, column=1).value or "")
                    for r in range(1, 40))
    assert "2024Q1" in inst


# ── C1: labels, and the deliberate asymmetry ────────────────────────────────
def test_period_labels_is_a_map_keyed_on_the_raw_value():
    got = period_labels([20231, 20232, 20241], "quarterly")
    assert got == {20231: "2023Q1", 20232: "2023Q2", 20241: "2024Q1"}


def test_annual_labels_are_the_year_itself():
    assert period_labels([2024, 2025], "annual") == {2024: "2024", 2025: "2025"}


def test_entry_form_and_display_label_agree_today_but_are_separate_functions():
    """They answer different questions and could reasonably diverge; asserting
    they agree is a fact about today, not a definition."""
    assert entry_label(20241, "quarterly") == "2024Q1"
    assert format_period(20241, "quarterly") == "2024Q1"


# ── V2: screen and board pack cannot drift ──────────────────────────────────
@pytest.mark.parametrize("value,freq", [
    (20231, "quarterly"), (20244, "quarterly"), (2024, "annual"), (2030, "annual"),
])
def test_screen_and_board_pack_render_the_same_label(value, freq):
    """⭐ ASSERTED IDENTICAL, NOT EACH CORRECT. Two renderers each 'correct in
    isolation' is exactly how the PDF and PPTX money formatters diverged."""
    screen = format_period(value, freq)
    assert pdf_period(value, freq) == screen, "PDF disagrees with the screen"
    assert pptx_period(value, freq) == screen, "PPTX disagrees with the screen"


def test_all_three_renderers_are_the_same_function_object():
    assert pdf_period is format_period
    assert pptx_period is format_period


def test_the_pdf_column_headers_use_the_shared_formatter():
    """Asserting the IMPORT is the right function does not assert the header
    CONSTRUCTION uses it — that gap let a `str(y)` mutation survive."""
    from services.api.report_pdf import period_headers
    assert period_headers([20231, 20232], "quarterly") == ["2023Q1", "2023Q2"]
    assert period_headers([2024, 2025], "annual") == ["2024", "2025"]
    assert period_headers(None, "annual") == []
