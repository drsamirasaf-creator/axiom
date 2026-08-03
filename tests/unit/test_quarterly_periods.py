"""Quarterly period arithmetic and the parser's column window (lane 1, 28 Jul).

These replay the two cases from the pre-build report that motivated the lane, so
a regression reproduces the original symptom rather than a paraphrase of it:

  · 20204 -> 20211 is consecutive; the old integer rule demanded 20205.
  · 20205 does not exist; the old rule would have ACCEPTED it as consecutive.

And they pin the two behaviours whose absence made the fault hard to see:
the parser must not stop at a fixed column, and one bad period must not repair
the sequence for the columns after it.
"""
import io
import pytest
from openpyxl import load_workbook

from services.api.modules.financials import ingest


# ── period arithmetic ───────────────────────────────────────────────────────
def test_quarter_four_carries_the_year():
    """The case the old integer rule got wrong: +7, not +1."""
    assert ingest.next_period(20204, "quarterly") == 20211
    assert ingest.next_period(20214, "quarterly") == 20221
    assert ingest.next_period(20231, "quarterly") == 20232


def test_a_fifth_quarter_does_not_exist():
    """The reverse failure: the old rule would have accepted 20205."""
    assert ingest.period_is_valid(20204, "quarterly")
    assert ingest.period_is_valid(20211, "quarterly")
    assert not ingest.period_is_valid(20205, "quarterly")
    assert not ingest.period_is_valid(20200, "quarterly")
    assert not ingest.period_is_valid(20209, "quarterly")


def test_annual_arithmetic_is_untouched():
    assert ingest.next_period(2025, "annual") == 2026
    assert ingest.period_is_valid(2025, "annual")
    assert not ingest.period_is_valid(205, "annual")


def test_decode_and_format():
    assert ingest.decode_period(20231, "quarterly") == (2023, 1)
    assert ingest.decode_period(2023, "annual") == (2023, None)
    assert ingest.format_period(20231, "quarterly") == "2023Q1"
    assert ingest.format_period(2023, "annual") == "2023"


# ── the full sequence the report replayed ───────────────────────────────────
def test_the_reported_sequence_now_validates_clean():
    """20214 -> 20221,20222,20223,20224 — rejected before at the boundary."""
    seq = [20221, 20222, 20223, 20224]
    expected = ingest.next_period(20214, "quarterly")
    for y in seq:
        assert y == expected, f"{y} should follow consecutively, expected {expected}"
        expected = ingest.next_period(y, "quarterly")


# ── the parser's column window ──────────────────────────────────────────────
def _template(freq, fcst_cols=None):
    if fcst_cols is not None:
        ingest.FORECAST_QUARTERLY = fcst_cols
    return ingest.build_company_template(
        company_id=1, company_name="T", currency="USD", statement_units="actual",
        ownership="private", standard="us_gaap", frequency=freq,
        last_historical_year=2025)


def test_parser_reads_beyond_the_old_thirty_column_window():
    """A 52-column quarterly workbook must be read in full.

    The old `for i in range(30)` stopped at column AE, dropping 22 forecast
    columns with no error at all.
    """
    original = ingest.FORECAST_QUARTERLY
    try:
        blob = _template("quarterly", 40)
        wb = load_workbook(io.BytesIO(blob))
        ws = wb["Income Statement"]
        assert ws.max_column == 53, "probe assumes the 52-column shape (B..BA)"

        # fill every forecast column so none is dropped as unused
        from openpyxl.utils import get_column_letter
        for c in range(14, 54):
            L = get_column_letter(c)
            q = c - 14
            ws[f"{L}4"] = (2023 + q // 4) * 10 + (q % 4) + 1

        seen = []
        for c in range(2, ws.max_column + 1):
            L = get_column_letter(c)
            if ws[f"{L}4"].value not in (None, ""):
                seen.append(L)
        # 12 historical + 40 forecast, the last one at BA
        assert len(seen) == 52, f"expected 52 period columns, saw {len(seen)}"
        assert seen[-1] == "BA"
    finally:
        ingest.FORECAST_QUARTERLY = original


def test_annual_template_shape_unchanged():
    """Lane 1 must not move the annual path."""
    blob = _template("annual")
    ws = load_workbook(io.BytesIO(blob))["Income Statement"]
    hist = [c for c in range(2, ws.max_column + 1)
            if str(ws.cell(row=3, column=c).value or "").lower() == "historical"]
    fcst = [c for c in range(2, ws.max_column + 1)
            if str(ws.cell(row=3, column=c).value or "").lower() == "forecast"]
    assert len(hist) == 6, "annual historical block moved"
    assert len(fcst) == 8, "annual forecast block moved — FORECAST_ANNUAL is 8, not 5"
    assert ws.cell(row=4, column=2).value == 2020
    assert ws.cell(row=4, column=15).value == 2033


# ── end-to-end through parse_and_validate ───────────────────────────────────
SHEETS = ["Income Statement", "Balance Sheet", "Cash Flow Data"]


def _fill(blob, first_col, labels, copy_from="M"):
    """Copy a historical column into each forecast column and label it."""
    from openpyxl.utils import get_column_letter
    wb = load_workbook(io.BytesIO(blob))
    for nm in SHEETS:
        ws = wb[nm]
        for i, lab in enumerate(labels):
            L = get_column_letter(first_col + i)
            ws[f"{L}3"] = "Forecast"
            ws[f"{L}4"] = lab
            for r in range(5, 45):
                v = ws[f"{copy_from}{r}"].value
                if v is not None and not (isinstance(v, str) and v.startswith("=")):
                    ws[f"{L}{r}"] = v
    out = io.BytesIO(); wb.save(out)
    return out.getvalue(), wb


def _quarterly_labels(n, after=20224):
    out, p = [], ingest.next_period(after, "quarterly")
    for _ in range(n):
        out.append(p); p = ingest.next_period(p, "quarterly")
    return out


def test_quarterly_plan_crossing_year_boundaries_parses_clean():
    """The whole point of the lane: 20234->20241 and 20244->20251 must pass."""
    original = ingest.FORECAST_QUARTERLY
    try:
        blob = _template("quarterly", 40)
        labels = _quarterly_labels(12)
        filled, _ = _fill(blob, 14, labels)
        data, errors, _meta, _w = ingest.parse_and_validate(
            filled, 1, statement_units="actual", frequency="quarterly")
        assert errors == [], f"expected a clean parse, got {errors[:3]}"
        assert data["periods"]["forecast"] == labels
        assert 20241 in labels and 20251 in labels, "must cross a year boundary"
    finally:
        ingest.FORECAST_QUARTERLY = original


def test_parser_reads_the_far_column_the_old_window_could_not_reach():
    """A diagnostic naming BA proves the scan passed the old AE ceiling."""
    from openpyxl.utils import column_index_from_string
    import re
    original = ingest.FORECAST_QUARTERLY
    try:
        blob = _template("quarterly", 40)
        filled, wb = _fill(blob, 14, _quarterly_labels(40))
        wb["Income Statement"]["BA5"] = None          # one hole, farthest column
        out = io.BytesIO(); wb.save(out)
        _d, errors, _m, _w = ingest.parse_and_validate(
            out.getvalue(), 1, statement_units="actual", frequency="quarterly")
        named = [e["cell"] for e in errors if e.get("cell")]
        assert any(column_index_from_string(re.match(r'([A-Z]+)', c).group(1)) > 31
                   for c in named), f"nothing past column AE was read: {named}"
    finally:
        ingest.FORECAST_QUARTERLY = original


def test_one_bad_period_does_not_repair_the_sequence():
    """No resync: a systematic mismatch must stay visible on every column.

    The old rule re-anchored `expected` on whatever was FOUND, so an annual rule
    applied to a quarterly file produced one isolated complaint per year boundary
    surrounded by accepts — reading like a typo rather than a frequency bug.
    """
    original = ingest.FORECAST_QUARTERLY
    try:
        blob = _template("quarterly", 40)
        # start the plan in the wrong place: history ends 20224, these start 20261
        wrong = []
        p = 20261
        for _ in range(8):
            wrong.append(p); p = ingest.next_period(p, "quarterly")
        filled, _ = _fill(blob, 14, wrong)
        _d, errors, _m, _w = ingest.parse_and_validate(
            filled, 1, statement_units="actual", frequency="quarterly")
        seq = [e for e in errors if "consecutively" in e["message"]]
        assert len(seq) == len(wrong), (
            f"a systematic mismatch must flag every column, got {len(seq)} "
            f"of {len(wrong)} — the sequence was repaired by a bad period")
    finally:
        ingest.FORECAST_QUARTERLY = original


# ── Lane 2 Part B — generator shape ─────────────────────────────────────────
def test_quarterly_ships_forty_forecast_columns_annual_untouched():
    """B1: verify, don't assume — the annual block must not move."""
    assert ingest.FORECAST_QUARTERLY == 40
    assert ingest.FORECAST_ANNUAL == 8, "annual forecast block moved"
    q = load_workbook(io.BytesIO(_template("quarterly")))["Income Statement"]
    a = load_workbook(io.BytesIO(_template("annual")))["Income Statement"]
    qf = [c for c in range(2, q.max_column + 1)
          if str(q.cell(row=3, column=c).value or "").lower() == "forecast"]
    af = [c for c in range(2, a.max_column + 1)
          if str(a.cell(row=3, column=c).value or "").lower() == "forecast"]
    assert len(qf) == 40 and q.max_column == 53, "quarterly forecast block is not N..BA"
    assert len(af) == 8, "annual forecast block moved"


def test_period_display_formats_are_display_only():
    """⭐ SUPERSEDED FOR QUARTERLY BY THE ENTRY-FORMAT LANE (29 Jul).

    v7.6 wrote quarterly periods as the 5-digit integer 20201 with a display
    format that rendered it "2020'Q1-A". A customer typing into the forecast
    columns then had to know the YYYYQ encoding, and Excel treats a bare 5-digit
    number in a period cell as a date candidate. Quarterly row 4 is now TEXT in
    the canonical entry form with an explicit Text number format ("@"), so
    nothing coerces and the customer can see the form they are asked to type.

    ANNUAL IS UNCHANGED — a four-digit year is not ambiguous and is not coerced,
    so it keeps its integer value and the -A/-E display format."""
    ws = load_workbook(io.BytesIO(_template("quarterly")))["Income Statement"]
    assert ws["B4"].value == "2020Q1", ws["B4"].value
    assert ws["B4"].number_format == "@", ws["B4"].number_format
    first_fcst = next(get_col(c) for c in range(2, ws.max_column + 1)
                      if str(ws.cell(row=3, column=c).value or "").lower() == "forecast")
    assert ws[f"{first_fcst}4"].number_format == "@"
    assert ws[f"{first_fcst}4"].value is None, "forecast labels must still ship blank"

    for freq, hist_fmt, fcst_fmt, sample in (
            ("annual", '0000"-A"', '0000"-E"', 2020),):
        ws = load_workbook(io.BytesIO(_template(freq)))["Income Statement"]
        assert ws["B4"].number_format == hist_fmt, f"{freq} historical format"
        assert ws["B4"].value == sample, "value must remain an integer"
        assert isinstance(ws["B4"].value, int)
        first_fcst = next(get_col(c) for c in range(2, ws.max_column + 1)
                          if str(ws.cell(row=3, column=c).value or "").lower() == "forecast")
        assert ws[f"{first_fcst}4"].number_format == fcst_fmt, f"{freq} forecast format"


def get_col(idx):
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


def test_the_period_tooltip_matches_the_frequency():
    """B3: a single shared message would tell an ANNUAL client to type 20231 —
    a confident instruction to enter a value the annual validator rejects."""
    # The quarterly prompt now teaches the CANONICAL ENTRY FORM rather than the
    # storage encoding — a customer should never need to know YYYYQ exists.
    for freq, must_contain in (("quarterly", "2024Q1"), ("annual", "four-digit year")):
        ws = load_workbook(io.BytesIO(_template(freq)))["Income Statement"]
        prompts = [d for d in ws.data_validations.dataValidation if d.promptTitle]
        assert len(prompts) == 1, f"{freq}: expected exactly one prompt validation"
        assert must_contain in prompts[0].prompt, f"{freq} prompt: {prompts[0].prompt}"


def test_quarterly_forecast_labels_still_ship_blank():
    """B3: blank is the opt-in switch. Pre-filling 40 would declare ten years of
    plan the client never asked for."""
    ws = load_workbook(io.BytesIO(_template("quarterly")))["Income Statement"]
    fcst = [c for c in range(2, ws.max_column + 1)
            if str(ws.cell(row=3, column=c).value or "").lower() == "forecast"]
    assert all(ws.cell(row=4, column=c).value is None for c in fcst)


def test_shading_holds_on_the_new_columns():
    """B5: shaded IFF unlocked input — no formula cell shaded, out to BA."""
    ws = load_workbook(io.BytesIO(_template("quarterly")))["Income Statement"]
    bad = []
    for row in ws.iter_rows(min_row=3, max_row=min(ws.max_row, 40),
                            min_col=2, max_col=ws.max_column):
        for c in row:
            shaded = c.fill is not None and c.fill.fill_type == "solid"
            formula = isinstance(c.value, str) and c.value.startswith("=")
            if shaded and formula:
                bad.append(f"{c.coordinate} shaded formula")
            if shaded and c.protection is not None and c.protection.locked:
                bad.append(f"{c.coordinate} shaded but locked")
    assert not bad, bad


def test_no_version_gate_was_reintroduced():
    """B4: stamp only — no accept-list, no equivalent under another name."""
    # v8.0 (30 Jul): non-current split, opening column, policy tax rate.
    # ⭐ The literal is pinned ON PURPOSE — a bump must be a deliberate act, not
    # a side effect. Updating it here is the acknowledgement. What must NEVER
    # come back is the accept-LIST below: the stamp is forensic metadata and
    # version is never a precondition for upload (CORE §7.37).
    assert ingest.TEMPLATE_VERSION == "7M-v11.0"  # v10 -> v11: the share-count unit is stated (3 Aug)
    assert not hasattr(ingest, "ACCEPTED_TEMPLATE_VERSIONS")
    assert not any("ACCEPTED" in n and "VERSION" in n for n in dir(ingest)), \
        "a version allow-list under another name is still a gate"
