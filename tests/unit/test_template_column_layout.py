"""v7.7 — per-column instructions, derived widths, computed row-1 height.

⭐ ASSERTIONS, NOT A VISUAL PASS. Nine headers were silently truncated (KPI
D/G/H/I, Objectives D/E/F, Key Results D/E/F, Organization C/E/F) because widths
were hand-maintained across four builders and the headers grew past them. Nothing
failed; it just looked wrong, and only to someone who opened the file. So the
guard is programmatic and covers every column of every list sheet.
"""
import io
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from services.api.modules.financials import ingest as ing

LIST_SHEETS = [
    (ing.ORG_SHEET, ing.ORG_HEADER_ROW),
    (ing.OBJECTIVES_SHEET, ing.OBJ_HEADER_ROW),
    (ing.KR_SHEET, ing.KR_HEADER_ROW),
    (ing.KPI_SHEET, ing.KPI_HEADER_ROW),
]
STATEMENT_SHEETS = ["Income Statement", "Balance Sheet", "Cash Flow Data"]


@pytest.fixture(scope="module", params=["annual", "quarterly"])
def wb(request):
    return load_workbook(io.BytesIO(ing.build_company_template(
        company_id=39, company_name="T", currency="USD", statement_units="actual",
        ownership="private", standard="us_gaap", frequency=request.param,
        last_historical_year=2025)))


def test_every_column_has_an_explicit_width_and_the_header_fits(wb):
    bad = []
    for name, hdr_row in LIST_SHEETS:
        ws = wb[name]
        for i in range(len(ing.COLUMN_HELP[name])):
            L = get_column_letter(1 + i)
            w = ws.column_dimensions[L].width
            h = ws.cell(row=hdr_row, column=1 + i).value
            if w is None:
                bad.append(f"{name}!{L} has no explicit width")
            elif h and len(str(h)) + 1 > w:
                bad.append(f"{name}!{L} header {len(str(h))} chars > width {w}")
    assert not bad, bad


def test_every_column_has_a_row1_instruction(wb):
    for name, _ in LIST_SHEETS:
        ws = wb[name]
        for i in range(len(ing.COLUMN_HELP[name])):
            c = ws.cell(row=1, column=1 + i)
            assert c.value, f"{name}!{get_column_letter(1+i)}1 has no instruction"


def test_row1_is_locked_and_never_shaded(wb):
    """Shading means unlocked input (v7.2). An instruction is neither."""
    bad = []
    for name, _ in LIST_SHEETS:
        ws = wb[name]
        for i in range(len(ing.COLUMN_HELP[name])):
            c = ws.cell(row=1, column=1 + i)
            if c.fill is not None and c.fill.fill_type == "solid":
                bad.append(f"{name}!{c.coordinate} is shaded")
            if c.protection is not None and c.protection.locked is False:
                bad.append(f"{name}!{c.coordinate} is unlocked")
    assert not bad, bad


def test_row1_height_is_computed_and_sane(wb):
    """A height is only meaningful if it reflects the wrapped text. Six lines is
    the point at which the INSTRUCTION is too long, not the row too short."""
    for name, _ in LIST_SHEETS:
        ws = wb[name]
        h = ws.row_dimensions[1].height
        assert h is not None, f"{name} row 1 has no computed height"
        lines = round((h - 6) / ing.ROW1_LINE_HEIGHT)
        assert 1 <= lines <= 5, f"{name} row 1 wraps to {lines} lines (h={h})"


def test_statement_sheets_keep_their_matrix_layout(wb):
    """Row 1 there is the sheet TITLE and rows 2-4 the banner/period block —
    applying the list treatment would overwrite structure, not improve it."""
    for name in STATEMENT_SHEETS:
        ws = wb[name]
        assert ws["A3"].value == "Period Type (Historical / Forecast)"
        assert ws["A4"].value == "Period (year)"


def test_resize_is_permitted_while_cells_stay_locked(wb):
    """The flags only stopped a client widening a column to READ it."""
    for ws in wb.worksheets:
        if not ws.protection.sheet:
            continue
        assert ws.protection.formatColumns is False, f"{ws.title}: column resize blocked"
        assert ws.protection.formatRows is False, f"{ws.title}: row resize blocked"


def test_inserting_rows_and_columns_stays_prohibited(wb):
    """The parser addresses rows and columns by POSITION, so an inserted one
    silently shifts meaning. Reported before touching, and deliberately kept."""
    ws = wb[ing.KPI_SHEET]
    assert ws.protection.insertRows is True
    assert ws.protection.insertColumns is True


def test_subtotal_cells_remain_locked(wb):
    ws = wb["Income Statement"]
    found = [c for row in ws.iter_rows(min_row=5) for c in row
             if isinstance(c.value, str) and c.value.startswith("=")]
    assert found, "no subtotal formulas found to check"
    assert all(c.protection.locked for c in found)


def test_sheet_level_rules_live_on_instructions_not_per_column(wb):
    """Duplicating a sheet-level rule above seven columns guarantees drift."""
    text = "\n".join(str(wb["Instructions"].cell(row=r, column=1).value or "")
                     for r in range(1, 45))
    assert "pre-formatted rows" in text
    assert "warnings and skipped" in text
    for name, _ in LIST_SHEETS:
        joined = " ".join(ing.COLUMN_HELP[name])
        assert "200" not in joined, f"{name}: row-cap leaked into per-column text"
        assert "never stop" not in joined, f"{name}: skip-rule leaked into per-column text"
