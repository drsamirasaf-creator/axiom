"""v13 — the cost-behaviour and capacity extension (T4.1).

⭐⭐ THE LABELS MUST EXIST BEFORE ANYTHING CAN DECLINE IN THEM. T3 declines
today with `cost_behaviour (fixed/variable split)` — an ENGINE TOKEN, on a page
a client reads, which is exactly the defect the naming lane closed on the ratio
surface. The naming resolver cannot fix it: it maps tokens to columns the
WORKBOOK CONTAINS, and no cost-behaviour column exists. So the template
extension comes first and the decline vocabulary follows it.

⭐ NO ANALYTICS IN THIS LANE. Contribution, break-even, operating leverage and
constrained mix are T4.2. What is asserted here is that the data can arrive,
that it arrives at the grain a controller can answer, and that supplying none of
it is not an error.
"""
import io
import os
import tempfile

os.environ.setdefault("DATABASE_URL",
                      "sqlite:///" + tempfile.mktemp(prefix="v13-", suffix=".db"))

import pytest
from openpyxl import load_workbook

from services.api.modules.financials import templates
from services.api.modules.financials import template_policy as policy


@pytest.fixture(scope="module")
def book():
    return load_workbook(io.BytesIO(templates.build_template("us_gaap")))


# ── 1 · the version ────────────────────────────────────────────────────────

def test_the_version_is_bumped():
    """⭐ A NEW SHEET IS A NEW TEMPLATE, and the stamp is what tells a reader
    which sheet they are holding. It is NOT a gate — §7.37 — and nothing below
    tests it as one."""
    assert policy.VERSION_MAJOR == 14
    assert policy.GENERIC_VERSION == "v14"


# ── 2 · cost behaviour, at the ruled grain ─────────────────────────────────

def test_the_cost_behaviour_sheet_exists_with_its_client_facing_labels(book):
    assert templates.COST_BEHAVIOUR_SHEET in book.sheetnames
    ws = book[templates.COST_BEHAVIOUR_SHEET]
    header = [ws.cell(row=policy.COST_BEHAVIOUR_HEADER_ROW, column=c).value
              for c in range(1, len(policy.COST_BEHAVIOUR_COLUMNS) + 1)]
    assert header == [label for label, _hint in policy.COST_BEHAVIOUR_COLUMNS]


def test_the_grain_is_the_cost_pool_not_the_product_line():
    """⭐⭐ THE RULING. A controller knows the support pool is largely fixed and
    freight is variable. Asking for a fixed/variable split of EVERY LINE's cost
    asks them to perform the allocation AXIOM exists to perform — and it is the
    grain the shared pools T1/T2 already allocate by driver.

    ⭐ There is no `Product` or `Segment` column here, and its absence is the
    ruling made structural rather than written in a note."""
    labels = [label for label, _hint in policy.COST_BEHAVIOUR_COLUMNS]
    assert "Cost Pool" in labels
    for per_line in ("Product", "Segment", "Product Line", "Line Code"):
        assert per_line not in labels, (
            f"{per_line!r} puts the behaviour split at the wrong grain")


def test_all_four_behaviour_classes_are_offered():
    assert policy.COST_BEHAVIOUR_CLASSES == (
        "fixed", "variable", "semi-variable", "step-fixed")


def test_semi_variable_and_step_fixed_cannot_collapse_into_fixed_and_variable():
    """⭐⭐ THE COLLAPSE IS THE DEFECT, AND COLUMNS ARE WHAT PREVENT IT. A
    semi-variable pool needs its two portions; a step-fixed pool needs the
    activity level at which it steps and the size of the step. Without those
    columns a client can only pick the nearest of `fixed`/`variable`, and a
    step-fixed cost averaged into a smooth one produces a SMOOTH OPTIMUM WHERE
    THE REAL ONE JUMPS — the capacity decision T4.2 exists to get right."""
    labels = [label for label, _hint in policy.COST_BEHAVIOUR_COLUMNS]
    assert "Fixed Portion" in labels and "Variable Portion" in labels
    assert "Step Threshold" in labels and "Step Size" in labels


def test_the_behaviour_column_offers_exactly_the_four_classes(book):
    """The dropdown is the vocabulary; a free-text column would collect
    'mostly fixed' and 'depends'."""
    ws = book[templates.COST_BEHAVIOUR_SHEET]
    formulas = {dv.formula1 for dv in ws.data_validations.dataValidation}
    assert any("BEHAVIOUR" in (f or "") for f in formulas), formulas


# ── 3 · capacity: a declared ceiling ───────────────────────────────────────

def test_the_capacity_sheet_collects_the_constraint_and_the_consumption(book):
    """⭐⭐ §8h·2 — "we cannot sell more than 8,000 units" is a declaration a
    controller can defend. AXIOM collects it; it never infers it, and it never
    estimates a demand RESPONSE."""
    assert templates.CAPACITY_SHEET in book.sheetnames
    assert set(policy.CAPACITY_MEASURES) == {
        "capacity_available", "consumption_per_unit", "maximum_sales_units"}


def test_the_capacity_sheet_carries_a_line_code_because_consumption_is_per_line(book):
    labels = [label for label, _hint in policy.CAPACITY_COLUMNS]
    assert "Line Code" in labels and "Resource" in labels


# ── 4 · prior versions parse unchanged ─────────────────────────────────────

def test_a_workbook_without_the_new_sheets_parses_identically(book):
    """⭐⭐ PARTIAL SUPPLY IS NEVER AN ERROR, AND THAT IS WHAT ADDITIVE MEANS. A
    v12 workbook has no cost-behaviour sheet; that is a FACT ABOUT THE FILE, not
    a failure of it. Same discipline as v9→v10 and v11→v12.

    ⭐ THE COMPARISON IS AGAINST THE SAME WORKBOOK WITHOUT THE SHEETS, not
    against an empty error list. A blank download has never parsed cleanly — it
    carries no company data — so asserting `errors == []` would have tested the
    template's emptiness rather than this lane's change. What must be true is
    that the two sheets introduce NOTHING.
    """
    full = templates.build_template("us_gaap")
    wb = load_workbook(io.BytesIO(full))
    for name in (templates.COST_BEHAVIOUR_SHEET, templates.CAPACITY_SHEET):
        del wb[name]
    buf = io.BytesIO()
    wb.save(buf)

    _ds_with, errs_with = templates.parse_workbook(full)
    _ds_without, errs_without = templates.parse_workbook(buf.getvalue())
    assert errs_with == errs_without, (
        "the new sheets change how a workbook parses: "
        f"{[e for e in errs_with if e not in errs_without]}")


def test_the_new_sheets_left_entirely_blank_are_not_an_error(book):
    """The sheets ship in the download. A client who fills neither must upload
    exactly as cleanly as one who never saw them."""
    _dataset, errors = templates.parse_workbook(templates.build_template("us_gaap"))
    # the download carries no company data, so it has never parsed clean; what
    # must be absent is any error ABOUT THE NEW SHEETS
    blamed = [e for e in errors
              if templates.COST_BEHAVIOUR_SHEET in str(e)
              or templates.CAPACITY_SHEET in str(e)
              or "behaviour" in str(e).lower() or "capacity" in str(e).lower()]
    assert blamed == [], f"a blank new sheet is being reported as an error: {blamed}"


# ── 5 · the decline vocabulary now names a column ──────────────────────────

def test_the_contribution_decline_names_the_client_column_not_an_engine_token():
    """⭐⭐ THE DEFECT THE NAMING LANE CLOSED, ON A NEW SURFACE. It read
    `supply cost_behaviour (fixed/variable split) to compute contribution_profit`
    — two engine tokens and a parenthetical, on a page a CFO reads."""
    from services.api.modules.financials import dimensional_analytics as A
    h = A.margin_hierarchy(revenue=100.0, direct_cost=60.0)
    cp = h["contribution_profit"]
    assert cp["available"] is False
    # ⭐ THE RAW TOKENS SURVIVE in `missing_measures` — the machine field, the
    # same way the ratio surface keeps `formula` beside `formula_display`. What
    # must carry no token is what a PERSON reads.
    assert cp["missing_measures"] == ["cost_behaviour (fixed/variable split)"]
    text = " ".join(cp["needs_columns"]) + " " + cp["unlocks"]
    for token in ("cost_behaviour", "contribution_profit", "direct_cost"):
        assert token not in text, f"engine token {token!r} in a client sentence"
    assert templates.COST_BEHAVIOUR_SHEET in text, text
    assert "Cost Behaviour" in text


def test_the_named_column_actually_exists_on_the_sheet(book):
    """⭐ A DECLINE THAT NAMES A COLUMN NOBODY CAN FIND IS WORSE THAN A TOKEN —
    the client goes looking. This is the assertion that keeps the sentence and
    the workbook in step."""
    from services.api.modules.financials import dimensional_analytics as A
    cp = A.margin_hierarchy(revenue=100.0)["contribution_profit"]
    ws = book[templates.COST_BEHAVIOUR_SHEET]
    header = {ws.cell(row=policy.COST_BEHAVIOUR_HEADER_ROW, column=c).value
              for c in range(1, len(policy.COST_BEHAVIOUR_COLUMNS) + 1)}
    named = [h for h in header if h and f"'{h}'" in cp["unlocks"]]
    assert named, f"the decline names no column that exists: {cp['unlocks']}"


# ── 6 · the data dictionary teaches what each field unlocks ────────────────

def test_every_new_column_is_explained_in_the_data_dictionary(book):
    """⭐ Derived from the same list the sheet is built from, so a column added
    without an explanation is impossible rather than merely discouraged."""
    dd = book["Data Dictionary"]
    text = " ".join(str(c.value) for row in dd.iter_rows() for c in row
                    if c.value is not None)
    for label, hint in (list(policy.COST_BEHAVIOUR_COLUMNS)
                        + list(policy.CAPACITY_COLUMNS)):
        assert label in text, f"{label!r} is not in the Data Dictionary"
        assert hint[:24] in text, f"{label!r} has no explanation"
