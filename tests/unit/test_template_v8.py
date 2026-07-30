"""v8 template: the shape the founder approved, proven by round-trip.

⭐ THE POINT OF EACH ADDITION, SO A FUTURE READER DOES NOT "SIMPLIFY" IT AWAY:

  · The five non-current components exist so operating-side invested capital can
    be built from NET PP&E. Building it from "Total Non-Current Assets" would
    capitalise goodwill and non-operating investments into an operating base,
    and the financing-vs-operating delta would then measure the aggregation
    rather than the unclassified items it is supposed to signal.
  · `noncurrent_assets` is DERIVED, not entered. Eight consumers read it and
    every stored dataset carries it; asking the customer for both the parts and
    the total invites a total that disagrees with its parts.
  · The Opening column is BALANCE SHEET ONLY. Flow statements have no opening
    balance. It is what makes rule 3's average basis computable for the earliest
    period — without it, year one falls back to BOP on every dataset forever.
  · Dropdowns are DEFINED NAMES pointing at a VISIBLE sheet. CORE §7.36/§7.37:
    the participant template stamped its version as a defined name over a HIDDEN
    sheet, Excel and Sheets dropped it, and legitimate uploads were rejected.
    Names are used for lists only, never as a stamp.
"""
import io
import os
import tempfile

os.environ.setdefault("DATABASE_URL",
                      "sqlite:///" + tempfile.mktemp(prefix="v8-", suffix=".db"))

import pytest
from openpyxl import load_workbook

from services.api.modules.financials import engines, templates as T

STANDARDS = ["us_gaap", "ifrs"]


def _wb(standard):
    return load_workbook(io.BytesIO(T.build_template(standard)))


@pytest.mark.parametrize("standard", STANDARDS)
def test_the_six_new_rows_are_present_and_labelled(standard):
    wb = _wb(standard)
    ws = wb[T.LABELS[standard]["sheets"]["balance_sheet"]]
    labels = {ws[f"A{r}"].value for r in range(5, 5 + len(engines.BS_KEYS))}
    for key in engines.BS_NONCURRENT_COMPONENTS + ["other_noncurrent_liabilities"]:
        assert T.LABELS[standard]["lines"][key] in labels, f"{key} row missing"


@pytest.mark.parametrize("standard", STANDARDS)
def test_dropdowns_are_defined_names_over_a_visible_sheet(standard):
    wb = _wb(standard)
    for name in ("OWNERSHIP", "PERIODTYPE", "PERIODTYPE_BS"):
        assert name in wb.defined_names, f"{name} not defined"
    assert "Lists" in wb.sheetnames
    assert wb["Lists"].sheet_state == "visible", \
        ("the Lists sheet must not be hidden — Excel and Google Sheets re-scope "
         "or drop defined names pointing at hidden sheets (CORE §7.36)")


@pytest.mark.parametrize("standard", STANDARDS)
def test_the_balance_sheet_offers_opening_and_the_flow_statements_do_not(standard):
    wb = _wb(standard)
    lab = T.LABELS[standard]["sheets"]
    assert "Opening" in wb[lab["balance_sheet"]]["A3"].value
    for block in ("income_statement", "cash_flow"):
        assert "Opening" not in (wb[lab[block]]["A3"].value or ""), \
            f"{block} is a flow statement; an opening column is meaningless"


def test_the_policy_tax_rate_has_its_own_cell():
    fields = [f for f, _l, _a in T.COMPANY_ROWS]
    assert "tax_rate" in fields and "tax_rate_policy" in fields, \
        ("rule 1's precedence is admin > template POLICY > implied EFFECTIVE; "
         "one cell cannot express it")


def test_the_version_is_stamped_but_the_family_is_the_gate():
    wb = _wb("us_gaap")
    assert wb["Instructions"]["A1"].value.startswith("AXIOM-FIN-TEMPLATE v8")
    # a v1 file must still parse — CORE §7.37, version is never a gate
    wb["Instructions"]["A1"] = "AXIOM-FIN-TEMPLATE v1 us_gaap"
    b = io.BytesIO(); wb.save(b)
    _ds, issues = T.parse_workbook(b.getvalue())
    assert not any("not an AXIOM" in (i.get("error") or "") for i in issues)


def _fill(standard, with_opening, with_components):
    """A minimal, valid two-year workbook."""
    wb = _wb(standard)
    lab = T.LABELS[standard]
    for block, keys in T.BLOCK_KEYS.items():
        ws = wb[lab["sheets"][block]]
        cols = ["B", "C"]
        for col, yr in zip(cols, (2023, 2024)):
            ws[f"{col}3"] = "Historical"
            ws[f"{col}4"] = yr
        if block == "balance_sheet" and with_opening:
            ws["D3"] = "Opening"
            ws["D4"] = 2022
        for r, key in enumerate(keys, start=5):
            optional = key in engines.BS_OPTIONAL_KEYS
            if optional and not with_components:
                continue
            for col in cols:
                ws[f"{col}{r}"] = 20.0 if optional else 100.0
            if block == "balance_sheet" and with_opening:
                ws[f"D{r}"] = 20.0 if optional else 100.0
    cw = wb["Company"]
    vals = {"name": "V8 Co", "ownership": "private", "currency": "USD",
            "tax_rate": 0.25, "tax_rate_policy": 0.21, "risk_free_rate": 0.04,
            "market_risk_premium": 0.055, "cost_of_debt": 0.06,
            "shares_outstanding": 1000, "unlevered_industry_beta": 1.0,
            "target_debt_to_equity": 0.5, "size_premium": 0.02,
            "specific_risk_premium": 0.01, "dlom": 0.2}
    for r, (field, _l, _a) in enumerate(T.COMPANY_ROWS, start=2):
        if field in vals:
            cw[f"B{r}"] = vals[field]
    b = io.BytesIO(); wb.save(b)
    return b.getvalue()


def test_a_pre_v8_shaped_upload_still_parses():
    """⭐ THE MIGRATION. No components, no opening column — every existing
    customer's next upload."""
    ds, issues = T.parse_workbook(_fill("us_gaap", False, False))
    assert ds is not None, issues
    assert "opening" not in ds["periods"]


def test_the_opening_column_lands_in_its_own_slot():
    ds, issues = T.parse_workbook(_fill("us_gaap", True, True))
    assert ds is not None, issues
    assert ds["periods"]["opening"] == 2022
    assert 2022 not in ds["periods"]["historical"], \
        ("the opening balance is not a reporting period — it has no income "
         "statement and no cash flow, and folding it into historical would "
         "misalign every consumer that zips the three blocks")


def test_the_noncurrent_total_is_derived_and_foots():
    ds, issues = T.parse_workbook(_fill("us_gaap", False, True))
    assert ds is not None, issues
    bs = ds["balance_sheet"]
    for y in ("2023", "2024"):
        parts = sum(bs[k][y] for k in engines.BS_NONCURRENT_COMPONENTS)
        assert bs["noncurrent_assets"][y] == pytest.approx(parts), \
            "the derived total must equal its components exactly"
